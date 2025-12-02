from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Query, Header, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
import asyncio
import json
import time
import traceback
import logging
from logging import Handler, LogRecord
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
import os
from contextlib import contextmanager
import pathlib
import shutil
import io
import pandas as pd
from datetime import datetime, timedelta
import httpx
from queue import Queue
import threading
import uuid
from collections import defaultdict
from functools import wraps
from .excel_grid_llm import explain_cell_value, ExplainCellRequest

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# ---------- Rate Limiting -------------------------------------
# Simple in-memory rate limiter
_rate_limit_store: Dict[str, List[float]] = defaultdict(list)
_rate_limit_lock = threading.Lock()

def rate_limit(max_requests: int = 100, window_seconds: int = 60):
    """
    Rate limiting decorator for API endpoints.
    Args:
        max_requests: Maximum number of requests allowed
        window_seconds: Time window in seconds
    """
    def decorator(func):
        @wraps(func)
        async def wrapper(request: Request, *args, **kwargs):
            # Get client IP
            client_ip = request.client.host if request.client else "unknown"
            
            current_time = time.time()
            window_start = current_time - window_seconds
            
            with _rate_limit_lock:
                # Clean old entries
                _rate_limit_store[client_ip] = [
                    req_time for req_time in _rate_limit_store[client_ip]
                    if req_time > window_start
                ]
                
                # Check rate limit
                if len(_rate_limit_store[client_ip]) >= max_requests:
                    logging.warning(f"Rate limit exceeded for IP: {client_ip}")
                    raise HTTPException(
                        status_code=429,
                        detail=f"Rate limit exceeded. Maximum {max_requests} requests per {window_seconds} seconds."
                    )
                
                # Record this request
                _rate_limit_store[client_ip].append(current_time)
            
            return await func(request, *args, **kwargs)
        return wrapper
    return decorator

# ---------- Optimized Excel Reading Helper -------------------------------------
def read_excel_optimized(file_path: pathlib.Path, sheet_name: Optional[str] = None, 
                         chunk_size: Optional[int] = None, usecols: Optional[List[str]] = None):
    """
    Optimized Excel file reading with optional chunking and column selection.
    Args:
        file_path: Path to Excel file
        sheet_name: Sheet name or index (None for first sheet)
        chunk_size: If provided, read in chunks (useful for large files)
        usecols: List of column names to read (reduces memory usage)
    Returns:
        DataFrame or list of DataFrames if chunking
    """
    try:
        if chunk_size and usecols:
            # Read in chunks with column selection for very large files
            chunks = pd.read_excel(
                file_path, 
                sheet_name=sheet_name, 
                engine='openpyxl',
                usecols=usecols,
                chunksize=chunk_size
            )
            return list(chunks)
        elif usecols:
            # Read with column selection only
            return pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                engine='openpyxl',
                usecols=usecols
            )
        else:
            # Standard read
            return pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                engine='openpyxl'
            )
    except ValueError as e:
        # Try alternative sheet names
        if sheet_name:
            for alt_name in ['Raw_Data', 'Raw Data', 0]:
                if alt_name != sheet_name:
                    try:
                        return pd.read_excel(file_path, sheet_name=alt_name, engine='openpyxl', usecols=usecols)
                    except ValueError:
                        continue
        raise e

# === Import your MCP orchestrator tools (already in app/mcp_scripts) ===
import sys
# Make the project root (parent of the `app/` directory) importable so
# `import app.mcp_scripts...` resolves correctly. Appending the `app/`
# directory itself causes Python to look for `app` *inside* that folder
# (i.e. app/app), which fails with ModuleNotFoundError.
sys.path.append(str(pathlib.Path(__file__).parent.parent.resolve()))

# Import analysis engine
from app.analysis import create_analysis_engine
# Import LLM config (must be after sys.path.append)
from app.llm_config import get_llm_config, EXCEL_FILE_NAME
# Import policy analyzer for advanced analysis
from app.policy_analyzer import detect_policy
from app.query_router_scope import validate_query_scope, get_out_of_scope_message
from app.query_router import classify_query, get_data_requirements, log_data_requirements, QueryIntent

from app.mcp_scripts.Parent_Rule_mcp_Server import (
    run_pipeline_address_expand,
    run_pipeline_services_applications,
    run_pipeline_all,
    run_export_curated_to_new_excel,
)
from app.mcp_scripts.Scoring import score_over_permissive_points_inline_from_excel

# build intermediates
from app.mcp_scripts.Rule_base_combined import combine_rule_bases
from app.mcp_scripts.Address_mcp import (
    load_address_data_from_parent_folder,
    expand_groups_and_append_all_addresses,
    clear_store as clear_address_store,
)
from app.mcp_scripts.Services_mcp import (
    load_service_data_from_parent_folder,
    expand_service_groups_consolidated,
    clear_store as clear_services_store,
)
from app.mcp_scripts.Application_mcp import (
    load_application_data_from_parent_folder,
    expand_application_groups_map_attribute,
    clear_store as clear_applications_store,
)
# stitch into Final_output.xlsx
from app.mcp_scripts.Parent_Rule_mcp_Server import (
    expand_addresses_only,
    expand_services_only,
    expand_applications_only,
)

APP_DIR = pathlib.Path(__file__).resolve().parent                # .../app
MCP_DEFAULT_EXP = APP_DIR / "expansions"                         # where MCP scripts write by default
MCP_DEFAULT_EXP.mkdir(parents=True, exist_ok=True)

@contextmanager
def pushd(new_dir: pathlib.Path):
    prev = os.getcwd()
    os.chdir(str(new_dir))
    try:
        yield
    finally:
        os.chdir(prev)

def _copy_if_exists(src: pathlib.Path, dst: pathlib.Path):
    """Copy src -> dst if src exists; ensure dst parent exists."""
    if src.exists():
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(str(src), str(dst))
# --------------------------------------------------------------------------------------
# App & CORS
# --------------------------------------------------------------------------------------

app = FastAPI(title="Palo Alto MCP Pipeline API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------- Event broadcasting for SSE clients -------------------------------------
# Each subscriber is an (asyncio.Queue, loop) pair; producers (any thread) call
# broadcast_event(...) which schedules queue.put_nowait on the queue's loop.
_subscribers: list[tuple[asyncio.Queue, asyncio.AbstractEventLoop]] = []
_sub_lock = asyncio.Lock()

async def register_subscriber(q: asyncio.Queue):
    loop = asyncio.get_running_loop()
    async with _sub_lock:
        _subscribers.append((q, loop))

async def unregister_subscriber(q: asyncio.Queue):
    async with _sub_lock:
        for i, (qq, _l) in enumerate(list(_subscribers)):
            if qq is q:
                _subscribers.pop(i)
                break

def broadcast_event(payload: dict):
    """Thread-safe broadcast: schedule putting payload (dict) into all subscriber queues."""
    text = json.dumps(payload)
    # Use a snapshot to avoid holding lock while scheduling
    subs = list(_subscribers)
    for q, loop in subs:
        try:
            loop.call_soon_threadsafe(q.put_nowait, text)
        except Exception as e:
            # Log failures but don't break the broadcast for other subscribers
            logging.warning(f"Failed to broadcast to subscriber: {e}")

def _send_tool_progress(tool: str, percent: int | None = None, message: str | None = None):
    payload: dict = {"tool": tool}
    if isinstance(percent, int):
        payload["percent"] = percent
    if message is not None:
        payload["message"] = message
    broadcast_event(payload)


# --------------------------------------------------------------------------------------
# Canonical folders (single 'expansions' folder, no timestamps)
# --------------------------------------------------------------------------------------

DATA_ROOT = pathlib.Path(__file__).resolve().parents[1] / "data"
DATA_ROOT.mkdir(parents=True, exist_ok=True)

# Where the UI uploads the entire SourceFiles/ tree
SOURCE_DIR = DATA_ROOT / "SourceFiles"
SOURCE_DIR.mkdir(parents=True, exist_ok=True)

# Where all pipeline outputs go
EXPANSIONS_DIR = DATA_ROOT / "expansions"
EXPANSIONS_DIR.mkdir(parents=True, exist_ok=True)

# ============================================================================
# Advanced Analysis Support (Activity Log Handler, Session Context, etc.)
# ============================================================================

# Thread-local storage for current session_id
_thread_local = threading.local()

# In-memory store for generated analysis files (id -> path)
ANALYSIS_FILES: Dict[str, str] = {}

# In-memory store for analysis progress (session_id -> progress data)
ANALYSIS_PROGRESS: Dict[str, Dict] = {}

# SSE event queues for each session (session_id -> Queue)
SSE_QUEUES: Dict[str, Queue] = {}

class ActivityLogHandler(Handler):
    """
    Custom logging handler that sends logs to activity_log in addition to console.
    Automatically captures all logging.info(), logging.warning(), logging.error() calls.
    """
    def emit(self, record: LogRecord):
        """Emit a log record to activity_log if session_id is available."""
        try:
            # Check if we have a session_id in thread-local storage
            session_id = getattr(_thread_local, 'session_id', None)
            
            if not session_id:
                return  # No active session, skip
            
            if session_id not in ANALYSIS_PROGRESS:
                return  # Session not found in progress store
            
            # Format the log message (just the message, not the full formatted log)
            msg = record.getMessage()
            
            # Skip if it's an activity log entry itself (avoid recursion/infinite loop)
            if "[Activity Log #" in msg or "Activity Log #" in msg:
                return
            
            # Determine log type from level
            log_type = "info"
            if record.levelno >= logging.ERROR:
                log_type = "error"
            elif record.levelno >= logging.WARNING:
                log_type = "warning"
            elif record.levelno >= logging.INFO:
                log_type = "info"
            
            # Get or create activity_log list
            if "activity_log" not in ANALYSIS_PROGRESS[session_id]:
                ANALYSIS_PROGRESS[session_id]["activity_log"] = []
            
            log_list = ANALYSIS_PROGRESS[session_id]["activity_log"]
            
            # Create log entry with timestamp
            log_entry = {
                "timestamp": datetime.now().strftime("%H:%M:%S"),
                "message": msg,  # Just the message, not the full formatted log
                "type": log_type
            }
            
            # Append to list (CRITICAL: Use the same list reference)
            log_list.append(log_entry)
            current_count = len(log_list)
            
            # Send to SSE queue if it exists
            if session_id in SSE_QUEUES:
                try:
                    SSE_QUEUES[session_id].put({
                        "type": "log",
                        "data": log_entry
                    })
                except Exception:
                    pass  # Ignore queue errors
            
            # Keep max 200 entries
            if len(log_list) > 200:
                log_list.pop(0)
        except Exception:
            # Log the error to console but don't break logging
            pass

class SessionContext:
    """Context manager to set session_id for logging handler"""
    def __init__(self, session_id: str):
        self.session_id = session_id
        self.prev_session_id = None
    
    def __enter__(self):
        self.prev_session_id = getattr(_thread_local, 'session_id', None)
        _thread_local.session_id = self.session_id
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.prev_session_id is not None:
            _thread_local.session_id = self.prev_session_id
        else:
            if hasattr(_thread_local, 'session_id'):
                delattr(_thread_local, 'session_id')
        return False

# Add custom handler to root logger (after ANALYSIS_PROGRESS is defined)
root_logger = logging.getLogger()
activity_handler = ActivityLogHandler()
activity_handler.setLevel(logging.INFO)
root_logger.addHandler(activity_handler)

def _find_static_file() -> pathlib.Path | None:
    """
    Find Final_output_scored_static.xlsx in data/expansions directory.
    Returns the Path if found, None otherwise.
    """
    static_file_path = EXPANSIONS_DIR / "Final_output_scored_static.xlsx"
    
    try:
        resolved_path = static_file_path.resolve()
        if resolved_path.exists() and resolved_path.is_file():
            logging.info(f"✅ Found static file at: {resolved_path}")
            return resolved_path
        else:
            logging.info(f"❌ Static file not found at: {resolved_path}")
            return None
    except Exception as e:
        logging.warning(f"⚠️  Error checking {static_file_path}: {e}")
        return None

# Direct Excel analysis - no SQLite needed

# Initialize analysis engine
analysis_engine = create_analysis_engine(DATA_ROOT)

# --------------------------------------------------------------------------------------
# Schemas
# --------------------------------------------------------------------------------------

class AddressExpandParams(BaseModel):
    # These are optional; we will supply sane defaults in the route
    final_output_path: Optional[str] = None
    source_col: Optional[str] = None
    dest_col: Optional[str] = None
    sheet_name: Optional[str] = None
    separator: Optional[str] = None

class ServicesAppsParams(BaseModel):
    final_output_path: Optional[str] = None
    service_col: Optional[str] = None
    application_col: Optional[str] = None
    sheet_name: Optional[str] = None
    separator: Optional[str] = None

class RunAllParams(BaseModel):
    # You can extend with more knobs if needed
    final_output_path: Optional[str] = None
    source_col: Optional[str] = None
    dest_col: Optional[str] = None
    service_col: Optional[str] = None
    application_col: Optional[str] = None
    sheet_name: Optional[str] = None
    separator: Optional[str] = None

class CuratedParams(BaseModel):
    final_output_path: Optional[str] = None
    output_path: Optional[str] = None
    input_sheet: Optional[str] = None
    output_sheet: Optional[str] = None

class CopyTabRequest(BaseModel):
    tab_name: str
    file_path: Optional[str] = None

# --------------------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------------------
def _build_intermediates_from_sourcefiles() -> dict:
    """
    Produce these under DATA/expansions/ by first letting the MCP scripts write to app/expansions/,
    then copying them into data/expansions/:
      - address_groups_merged.xlsx
      - service_groups_merged.xlsx
      - application_groups_mapped.xlsx
      - rule_base_combined.xlsx
    """
    # Make sure app/expansions exists (MCP default output location)
    MCP_DEFAULT_EXP.mkdir(parents=True, exist_ok=True)

    # 1) Let MCP scripts run with their expected CWD so they write into app/expansions/
    with pushd(APP_DIR):
        # Address objects & groups
        load_address_data_from_parent_folder(
            parent_folder=str(SOURCE_DIR),
            addresses_subdir="Addresses",
            groups_subdir="Address Groups",
        )
        expand_groups_and_append_all_addresses(output_column=None, purge_old=True)

        # Services objects & groups
        load_service_data_from_parent_folder(
            parent_folder=str(SOURCE_DIR),
            services_subdir="Services",
            groups_subdir="Service Groups",
            recursive=True,
        )
        # writes app/expansions/service_groups_merged.xlsx
        expand_service_groups_consolidated(save_filename="service_groups_merged.xlsx", purge_old=True)

        # Applications & groups
        load_application_data_from_parent_folder(
            parent_folder=str(SOURCE_DIR),
            apps_subdir="Applications",
            groups_subdir="Application Groups",
            recursive=True,
        )
        expand_application_groups_map_attribute(target_field="Standard Ports", purge_old=True)

        # Rulebase combined – if your combiner accepts explicit output_path,
        # pass app/expansions so files are co-located with the other intermediates.
        combine_rule_bases(
            input_paths=str((SOURCE_DIR / "Rulebase").resolve()),
            output_path=str((MCP_DEFAULT_EXP / "rule_base_combined.xlsx").resolve()),
            recursive=True,
            purge_old=True,
        )

    # 2) Copy from app/expansions -> data/expansions (canonical place we read from later)
    addr_src = MCP_DEFAULT_EXP / "address_groups_merged.xlsx"
    svc_src  = MCP_DEFAULT_EXP / "service_groups_merged.xlsx"
    app_src  = MCP_DEFAULT_EXP / "application_groups_mapped.xlsx"
    rule_src = MCP_DEFAULT_EXP / "rule_base_combined.xlsx"

    addr_dst = EXPANSIONS_DIR / "address_groups_merged.xlsx"
    svc_dst  = EXPANSIONS_DIR / "service_groups_merged.xlsx"
    app_dst  = EXPANSIONS_DIR / "application_groups_mapped.xlsx"
    rule_dst = EXPANSIONS_DIR / "rule_base_combined.xlsx"

    _copy_if_exists(addr_src, addr_dst)
    _copy_if_exists(svc_src,  svc_dst)
    _copy_if_exists(app_src,  app_dst)
    _copy_if_exists(rule_src, rule_dst)

    # 3) Verify they exist; fail clearly if anything is missing
    missing = []
    for lbl, p in [
        ("address_groups_merged.xlsx", addr_dst),
        ("service_groups_merged.xlsx", svc_dst),
        ("application_groups_mapped.xlsx", app_dst),
        ("rule_base_combined.xlsx", rule_dst),
    ]:
        if not p.exists():
            missing.append(lbl)

    if missing:
        raise HTTPException(
            status_code=500,
            detail=f"Intermediates not produced: {', '.join(missing)}. "
                   f"Check MCP scripts wrote into {MCP_DEFAULT_EXP}."
        )

    return {
        "address_result_path": str(addr_dst.resolve()),
        "services_result_path": str(svc_dst.resolve()),
        "applications_result_path": str(app_dst.resolve()),
        "rule_result_path": str(rule_dst.resolve()),
    }


def _safe_within_data(p: pathlib.Path) -> bool:
    """Restrict /api/download to only serve files under DATA_ROOT."""
    try:
        p.resolve().relative_to(DATA_ROOT.resolve())
        return True
    except Exception:
        return False

def _save_into_source_tree(
    files: list[UploadFile],
    paths: list[str] | None = None
) -> list[dict]:
    """
    Save to data/SourceFiles/, preserving subfolders from `paths`.
    Removes an optional leading 'SourceFiles/' prefix the browser may include.
    """
    saved: list[dict] = []
    paths = paths or []

    for i, up in enumerate(files):
        # choose the relative path
        rel = paths[i] if i < len(paths) and paths[i] else (up.filename or "")
        rel = rel.replace("\\", "/").lstrip("/")              # normalize
        if "../" in rel:                                      # very light traversal guard
            rel = rel.split("../")[-1]
        if rel.startswith("SourceFiles/"):                    # <<< KEY FIX
            rel = rel[len("SourceFiles/"):]                   # strip the prefix

        subpath = rel if "/" in rel else os.path.basename(rel) or "upload.bin"

        dest = (SOURCE_DIR / subpath).resolve()
        dest.parent.mkdir(parents=True, exist_ok=True)
        with open(dest, "wb") as fh:
            shutil.copyfileobj(up.file, fh)

        saved.append({"name": os.path.basename(subpath), "path": str(dest)})

    return saved



REQUIRED_SUBDIRS = [
    "Addresses",
    "Address Groups",
    "Services",
    "Service Groups",
    "Applications",
    "Application Groups",
    "Rulebase",
]

def _validate_source_tree() -> None:
    """Check that each required subfolder has at least one CSV/XLSX file."""
    missing = []
    for sub in REQUIRED_SUBDIRS:
        # Accept either the exact folder name (with spaces) or a common
        # variant where spaces were replaced by underscores (e.g.,
        # "Address Groups" vs "Address_Groups"). This makes uploads
        # resilient to different OS/tools that normalize names differently.
        subdir_space = SOURCE_DIR / sub
        subdir_underscore = SOURCE_DIR / sub.replace(" ", "_")

        def _has_data(p: pathlib.Path) -> bool:
            return p.exists() and (any(p.glob("*.csv")) or any(p.glob("*.xlsx")))

        if not (_has_data(subdir_space) or _has_data(subdir_underscore)):
            missing.append(sub)
    if missing:
        raise HTTPException(
            status_code=400,
            detail=(
                "Missing input files in subfolders: "
                + ", ".join(missing)
                + f" under {SOURCE_DIR}"
            ),
        )

# --------------------------------------------------------------------------------------
# Routes
# --------------------------------------------------------------------------------------
@app.post("/api/upload")
@rate_limit(max_requests=20, window_seconds=60)  # 20 uploads per minute
async def api_upload(
    request: Request,
    files: List[UploadFile] = File(...),
    paths: Optional[List[str]] = Form(None),
):
    """
    Save uploaded files under data/SourceFiles/ preserving relative paths.
    This endpoint is used by the frontend to upload files without immediately
    running the pipeline.
    """
    # Clear source directory and stores to ensure fresh files for each client
    shutil.rmtree(SOURCE_DIR, ignore_errors=True)
    SOURCE_DIR.mkdir(parents=True, exist_ok=True)
    
    # Clear all global stores to prevent data contamination between clients
    clear_address_store()
    clear_services_store()
    clear_applications_store()
    
    saved = _save_into_source_tree(files, paths)
    return JSONResponse({"uploaded": saved})


@app.get("/api/run/events")
async def api_run_events():
    """Subscribe to server-side broadcast events and stream them to the client.
    If no events arrive, send periodic heartbeat messages so the client stays alive.
    """
    q: asyncio.Queue = asyncio.Queue()
    await register_subscriber(q)

    async def event_generator():
        try:
            # initial connected message
            yield f"data: {json.dumps({'message': 'connected'})}\n\n"
            while True:
                try:
                    # wait up to 6s for a real event
                    text = await asyncio.wait_for(q.get(), timeout=6.0)
                    # text is already JSON string
                    yield f"data: {text}\n\n"
                except asyncio.TimeoutError:
                    # heartbeat
                    hb = json.dumps({"message": "heartbeat", "ts": int(time.time())})
                    yield f"data: {hb}\n\n"
        except asyncio.CancelledError:
            return
        finally:
            try:
                await unregister_subscriber(q)
            except Exception:
                pass

    return StreamingResponse(event_generator(), media_type="text/event-stream")


@app.get("/api/preview")
def api_preview(path: str = Query(..., description="Absolute path within ./data to preview"), format: str = Query("html")):
    """Return a small HTML preview of the first sheet (xlsx) or start of CSV.

    The `path` must point to a file under the `data/` directory (same safety
    rules as `/api/download`). `format` currently supports `html` (returns
    text/html). This endpoint is intentionally minimal and returns the first
    sheet / first few rows to keep responses small.
    """
    p = pathlib.Path(path)
    if not _safe_within_data(p) or not p.exists() or not p.is_file():
        raise HTTPException(status_code=404, detail="File not found or not allowed")

    try:
        suffix = p.suffix.lower()
        if suffix in (".xlsx", ".xls"):
            # read first sheet only
            df = pd.read_excel(p, sheet_name=0, engine="openpyxl")
        elif suffix in (".csv", ".txt"):
            df = pd.read_csv(p, nrows=200)
        else:
            # unsupported preview type
            raise HTTPException(status_code=400, detail=f"Unsupported preview file type: {suffix}")

        # take a small slice to avoid huge responses
        small = df.head(200)
        html = small.to_html(index=False, classes="preview-table")
        return JSONResponse(content=html, media_type="text/html")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Preview failed: {str(exc)}")

@app.post("/api/upload-security-control")
async def upload_security_control(file: UploadFile = File(...)):
    """
    PDF upload endpoint is disabled. Security controls are now determined from rule boolean flags using ComplianceGrid mapping.
    """
    logging.warning("⚠️ PDF upload endpoint called but is disabled - using ComplianceGrid mapping instead")
    return JSONResponse(
        status_code=410,  # 410 Gone - resource is no longer available
        content={
            "success": False,
            "error": "PDF upload is no longer supported. Security controls are automatically determined from rule boolean flags using the ComplianceGrid mapping."
        }
    )

@app.post("/api/analyze-security-controls")
async def analyze_security_controls(request: Request):
    """
    PDF analysis endpoint is disabled. Security controls are now determined from rule boolean flags using ComplianceGrid mapping.
    """
    logging.warning("⚠️ PDF analysis endpoint called but is disabled - using ComplianceGrid mapping instead")
    return JSONResponse(
        status_code=410,  # 410 Gone - resource is no longer available
        content={
            "success": False,
            "error": "PDF analysis is no longer supported. Security controls are automatically determined from rule boolean flags using the ComplianceGrid mapping."
        }
    )

@app.post("/api/upload-and-run")
@rate_limit(max_requests=10, window_seconds=60)  # 10 upload-and-run per minute (more resource intensive)
async def api_upload_and_run(
    request: Request,
    files: List[UploadFile] = File(...),
    run_curated: bool = True,
    paths: Optional[List[str]] = Form(None),
):
    # 1) Save incoming files into data/SourceFiles/ (preserve subfolders)
    #    Clear source directory to ensure fresh files for each client
    shutil.rmtree(SOURCE_DIR, ignore_errors=True)
    SOURCE_DIR.mkdir(parents=True, exist_ok=True)
    
    # Clear all global stores to prevent data contamination between clients
    clear_address_store()
    clear_services_store()
    clear_applications_store()
    
    saved = _save_into_source_tree(files, paths)

    # let clients know upload finished
    _send_tool_progress("upload", 100, "Uploaded files saved")

    try:
        # 2) Validate expected subfolders/files exist
        _send_tool_progress("validate", 0, "Validating source tree")
        _validate_source_tree()
        _send_tool_progress("validate", 100, "Validation complete")

        # 3) Run the full pipeline into data/expansions/Final_output.xlsx
        final_target = str((EXPANSIONS_DIR / "Final_output.xlsx").resolve())

        # signal start of build intermediates
        _send_tool_progress("build_intermediates", 0, "Building intermediates")
        inter = _build_intermediates_from_sourcefiles()
        _send_tool_progress("build_intermediates", 100, "Intermediates ready")

        # Expand Addresses → rules_expanded in final_xlsx
        _send_tool_progress("address_expand", 0, "Expanding addresses")
        expand_addresses_only(
            rule_result_path=inter["rule_result_path"],
            address_result_path=inter["address_result_path"],
            final_output_path=final_target,
            source_col="Source Address",
            dest_col="Destination Address",
            sheet_name="rules",
            separator="; ",
        )
        _send_tool_progress("address_expand", 100, "Address expansion complete")

        # Apply Services → same final_xlsx
        _send_tool_progress("services", 0, "Expanding services")
        expand_services_only(
            rule_result_path=inter["rule_result_path"],
            services_result_path=inter["services_result_path"],
            final_output_path=final_target,
            service_col="Service",
            sheet_name="rules",
            separator="; ",
        )
        _send_tool_progress("services", 100, "Services applied")

        # Apply Applications → same final_xlsx
        _send_tool_progress("applications", 0, "Expanding applications")
        expand_applications_only(
            rule_result_path=inter["rule_result_path"],
            applications_result_path=inter["applications_result_path"],
            final_output_path=final_target,
            application_col="Application",
            sheet_name="rules",
            separator="; ",
        )
        _send_tool_progress("applications", 100, "Applications applied")

        payload = {
            "uploaded": saved,
            "full": {"final_output_path": final_target},
            "downloads": {"final_output": f"/api/download?path={final_target}"}
        }

        # 4) Optionally export curated into data/expansions/Final_output_curated.xlsx
        if run_curated:
            _send_tool_progress("export_curated", 0, "Exporting curated output")
            curated_target = str((EXPANSIONS_DIR / "Final_output_curated.xlsx").resolve())
            run_export_curated_to_new_excel(
                final_output_path=final_target,
                output_path=curated_target,
                input_sheet="rules_expanded",
                output_sheet="rules_curated",
                purge_old=True,
            )
            payload["curated"] = {"output_path": curated_target}
            payload["downloads"]["final_output_curated"] = f"/api/download?path={curated_target}"
            _send_tool_progress("export_curated", 100, "Curated export complete")

        # done
        _send_tool_progress("pipeline", 100, "Pipeline complete")
        return JSONResponse(payload)

    except Exception as exc:
        tb = traceback.format_exc()
        # log full traceback to server logs
        logging.error("Exception in /api/upload-and-run:\n%s", tb)
        # notify clients via SSE and return an informative 500 response
        _send_tool_progress("pipeline", None, f"Error: {str(exc)}")
        return JSONResponse(status_code=500, content={
            "detail": "Internal server error during upload-and-run",
            "error": str(exc),
            # include a limited-length trace to help debugging
            "trace": tb[:4000]
        })

@app.post("/api/run/address-expand")
def api_run_address_expand(params: AddressExpandParams):
    # Check if SourceFiles directory exists and has content
    if not SOURCE_DIR.exists() or not any(SOURCE_DIR.iterdir()):
        raise HTTPException(
            status_code=400,
            detail="No source files found. Please upload files first using /api/upload or /api/upload-and-run endpoint."
        )
    
    _validate_source_tree()

    final_target = params.final_output_path or str((EXPANSIONS_DIR / "Final_output.xlsx").resolve())
    kwargs = {k: v for k, v in params.model_dump().items() if v not in (None, "")}
    kwargs.update({
        "parent_folder": str(SOURCE_DIR),
        "addresses_subdir": "Addresses",
        "address_groups_subdir": "Address Groups",
        "rulebase_subdir": "Rulebase",
        "final_output_path": final_target,
    })

    res = run_pipeline_address_expand(**kwargs)
    res = res or {}
    res["final_output_path"] = final_target
    return JSONResponse(res)

@app.post("/api/run/services-applications")
def api_run_services_applications(params: ServicesAppsParams):
    # Check if SourceFiles directory exists and has content
    if not SOURCE_DIR.exists() or not any(SOURCE_DIR.iterdir()):
        raise HTTPException(
            status_code=400,
            detail="No source files found. Please upload files first using /api/upload or /api/upload-and-run endpoint."
        )
    
    _validate_source_tree()

    final_target = params.final_output_path or str((EXPANSIONS_DIR / "Final_output.xlsx").resolve())
    kwargs = {k: v for k, v in params.model_dump().items() if v not in (None, "")}
    kwargs.update({
        "parent_folder": str(SOURCE_DIR),
        "services_subdir": "Services",
        "service_groups_subdir": "Service Groups",
        "applications_subdir": "Applications",
        "application_groups_subdir": "Application Groups",
        "final_output_path": final_target,
    })

    # services & applications applied into the same Final_output.xlsx
    res = run_pipeline_services_applications(**kwargs)
    res = res or {}
    res["final_output_path"] = final_target
    return JSONResponse(res)
@app.post("/api/run/all")
def api_run_all():
    try:
        # Check if SourceFiles directory exists and has content
        if not SOURCE_DIR.exists() or not any(SOURCE_DIR.iterdir()):
            raise HTTPException(
                status_code=400,
                detail="No source files found. Please upload files first using /api/upload or /api/upload-and-run endpoint."
            )
        
        _validate_source_tree()  # make sure SourceFiles/* exist

        # 1) Build intermediates under data/expansions/
        _send_tool_progress("build_intermediates", 0, "Building intermediates")
        inter = _build_intermediates_from_sourcefiles()
        _send_tool_progress("build_intermediates", 100, "Intermediates ready")

        # 2) Write into a single Final_output.xlsx
        final_xlsx = str((EXPANSIONS_DIR / "Final_output.xlsx").resolve())

        # Expand Addresses → rules_expanded in final_xlsx
        _send_tool_progress("address_expand", 0, "Expanding addresses")
        expand_addresses_only(
            rule_result_path=inter["rule_result_path"],
            address_result_path=inter["address_result_path"],
            final_output_path=final_xlsx,
            source_col="Source Address",
            dest_col="Destination Address",
            sheet_name="rules",
            separator="; ",
        )
        _send_tool_progress("address_expand", 100, "Address expansion complete")

        # Apply Services → same final_xlsx
        _send_tool_progress("services", 0, "Expanding services")
        expand_services_only(
            rule_result_path=inter["rule_result_path"],
            services_result_path=inter["services_result_path"],
            final_output_path=final_xlsx,
            service_col="Service",
            sheet_name="rules",
            separator="; ",
        )
        _send_tool_progress("services", 100, "Services applied")

        # Apply Applications → same final_xlsx
        _send_tool_progress("applications", 0, "Expanding applications")
        expand_applications_only(
            rule_result_path=inter["rule_result_path"],
            applications_result_path=inter["applications_result_path"],
            final_output_path=final_xlsx,
            application_col="Application",
            sheet_name="rules",
            separator="; ",
        )
        _send_tool_progress("applications", 100, "Applications applied")

        return JSONResponse({
            "final_output_path": final_xlsx,
            "address_phase": {
                "address_result_path": inter["address_result_path"],
                "rule_result_path": inter["rule_result_path"],
                "final_output_path": final_xlsx,
            },
            "services_applications_phase": {
                "services_result_path": inter["services_result_path"],
                "applications_result_path": inter["applications_result_path"],
                "final_output_path": final_xlsx,
            },
        })
    except Exception as exc:
        tb = traceback.format_exc()
        logging.error("Exception in /api/run/all:\n%s", tb)
        _send_tool_progress("pipeline", None, f"Error: {str(exc)}")
        return JSONResponse(status_code=500, content={
            "detail": "Internal server error during run/all",
            "error": str(exc),
            "trace": tb[:4000]
        })



@app.post("/api/run/export-curated")
def api_export_curated(params: CuratedParams):
    """
    Export curated workbook from the current Final_output.xlsx.
    Writes to data/expansions/Final_output_curated.xlsx.
    """
    final_source = pathlib.Path(params.final_output_path or (EXPANSIONS_DIR / "Final_output.xlsx")).resolve()
    if not final_source.exists():
        raise HTTPException(400, detail=f"Final_output.xlsx not found at {final_source}. Run 'Run All' first.")

    curated_target = pathlib.Path(params.output_path or (EXPANSIONS_DIR / "Final_output_curated.xlsx")).resolve()

    kwargs = {k: v for k, v in params.model_dump().items() if v not in (None, "")}
    kwargs.update({
        "final_output_path": str(final_source),
        "output_path": str(curated_target),
        # sensible defaults; you can override via body if desired
        "input_sheet": params.input_sheet or "rules_expanded",
        "output_sheet": params.output_sheet or "rules_curated",
        "purge_old": True,
    })

    res = run_export_curated_to_new_excel(**kwargs)
    res = res or {}
    res["output_path"] = str(curated_target)
    res["source_final_input"] = str(final_source)
    return JSONResponse(res)

@app.post("/api/run/scoring")
def api_run_scoring():
    """Run scoring on Final_output_curated.xlsx."""
    curated_file = EXPANSIONS_DIR / "Final_output_curated.xlsx"
    if not curated_file.exists():
        raise HTTPException(status_code=404, detail="Final_output_curated.xlsx not found. Please run curation first.")
    
    try:
        # Create a progress callback wrapper that converts Scoring.py format to _send_tool_progress format
        # Scoring.py callback format: (message: str, percent: int)
        # _send_tool_progress format: (tool: str, percent: int, message: str)
        def scoring_progress_callback(message: str, percent: int):
            """Wrapper to convert scoring progress callbacks to tool progress format."""
            _send_tool_progress("scoring", percent, message)
            # Also log for debugging
            logging.info(f"📊 Scoring progress: {percent}% - {message}")
        
        _send_tool_progress("scoring", 0, "Starting scoring process")
        
        # Run scoring with progress callback
        output_path = EXPANSIONS_DIR / "Final_output_scored.xlsx"
        result = score_over_permissive_points_inline_from_excel(
            input_path=str(curated_file),
            sheet_name="rules_curated",
            output_path=str(output_path),
            separator="; ",
            progress_callback=scoring_progress_callback
        )
        
        _send_tool_progress("scoring", 100, "Scoring completed")
        
        return JSONResponse({
            "status": "success",
            "message": "Scoring completed successfully",
            "output_file": str(output_path),
            "result": result
        })
    except Exception as e:
        logging.error(f"Error running scoring: {e}")
        _send_tool_progress("scoring", None, f"Error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Scoring failed: {str(e)}")



@app.post("/api/clear-cache")
def api_clear_cache():
    """Clear download cache."""
    cache_buster = int(time.time())
    return JSONResponse({
        "status": "success",
        "message": "Cache cleared",
        "cache_buster": cache_buster,
        "timestamp": cache_buster
    })

@app.post("/api/import-to-sqlite")
async def api_import_to_sqlite(file: Optional[UploadFile] = File(None)):
    """Import the scored Excel file and return analysis data directly."""
    if file:
        try:
            import tempfile
            
            # Save uploaded file temporarily
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                content = file.file.read()
                tmp_file.write(content)
                tmp_file_path = pathlib.Path(tmp_file.name)
            
            try:
                # Get analysis data directly from Excel
                analysis_data = analysis_engine.get_analysis_data_from_excel(tmp_file_path)
                
                return JSONResponse({
                    "message": "Analysis completed successfully",
                    "analysis_data": analysis_data
                })
            finally:
                # Clean up temporary file
                if tmp_file_path.exists():
                    tmp_file_path.unlink()
                    
        except Exception as e:
            logging.error(f"Error analyzing uploaded file: {e}")
            raise HTTPException(status_code=500, detail=f"Failed to analyze file: {str(e)}")
    else:
        # Use existing scored file - find the most recent one
        scored_file = None
        possible_files = [
            EXPANSIONS_DIR / "Final_output_scored.xlsx",
            EXPANSIONS_DIR / "Final_output_scored_static.xlsx",
            EXPANSIONS_DIR / "over_permissive_scored.xlsx",
            EXPANSIONS_DIR / "Final_output_scored - Copy.xlsx",
        ]
        
        # Find the first existing file
        for file_path in possible_files:
            if file_path.exists():
                scored_file = file_path
                break
        
        if not scored_file:
            raise HTTPException(
                status_code=404, 
                detail="Scored Excel file not found. Please upload a file or run scoring first."
            )
        
        try:
            analysis_data = analysis_engine.get_analysis_data_from_excel(scored_file)
            return JSONResponse({
                "message": "Analysis completed successfully",
                "analysis_data": analysis_data,
                "file_used": str(scored_file)
            })
        except Exception as e:
            logging.error(f"Error analyzing existing file: {e}")
            raise HTTPException(status_code=500, detail=f"Failed to analyze file: {str(e)}")

@app.get("/api/analysis-data")
def api_get_analysis_data():
    """Get analysis data directly from Excel file."""
    # Try to find the most recent scored file (prefer dynamic file, then static)
    scored_file = None
    possible_files = [
        EXPANSIONS_DIR / "Final_output_scored.xlsx",  # Dynamic file from scoring
        EXPANSIONS_DIR / "Final_output_scored_static.xlsx",  # Static file from scoring
        EXPANSIONS_DIR / "over_permissive_scored.xlsx",  # Default output name
        EXPANSIONS_DIR / "Final_output_scored - Copy.xlsx",  # Fallback for legacy
    ]
    
    # Find the first existing file (prioritized by order above)
    for file_path in possible_files:
        if file_path.exists():
            scored_file = file_path
            break
    
    if not scored_file:
        raise HTTPException(
            status_code=404, 
            detail="Scored Excel file not found. Please run scoring first. Looking for: Final_output_scored.xlsx"
        )
    
    try:
        analysis_data = analysis_engine.get_analysis_data_from_excel(scored_file)
        # Add file info to help with cache-busting
        file_stat = scored_file.stat()
        analysis_data["_metadata"] = {
            "file_path": str(scored_file),
            "file_name": scored_file.name,
            "last_modified": file_stat.st_mtime,
            "file_size": file_stat.st_size
        }
        return JSONResponse(analysis_data)
    except Exception as e:
        logging.error(f"Error getting analysis data: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to get analysis data: {str(e)}")

@app.get("/api/load-static-file")
def api_load_static_file():
    """Load static file (Final_output_scored_static.xlsx) and return analysis data for dashboard."""
    # Try to find the static file
    static_file_path = EXPANSIONS_DIR / "Final_output_scored_static.xlsx"
    
    if not static_file_path.exists():
        # Also check for Final_output_scored.xlsx as fallback
        fallback_file = EXPANSIONS_DIR / "Final_output_scored.xlsx"
        if fallback_file.exists():
            static_file_path = fallback_file
            logging.info(f"✅ Using fallback file: {static_file_path}")
        else:
            logging.warning(f"❌ Static file not found at: {EXPANSIONS_DIR / 'Final_output_scored_static.xlsx'}")
            raise HTTPException(
                status_code=404,
                detail="Final_output_scored_static.xlsx not found. Please run scoring to generate it."
            )
    
    try:
        logging.info(f"📖 Loading static file: {static_file_path}")
        # Use the dashboard-specific method that returns the correct format
        dashboard_data = analysis_engine.get_dashboard_data_from_excel(static_file_path)
        
        return JSONResponse(dashboard_data)
    except Exception as e:
        logging.error(f"Error loading static file: {e}")
        logging.exception("Full traceback:")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to load static file: {str(e)}"
        )

@app.get("/api/excel-grid-data")
@rate_limit(max_requests=100, window_seconds=60)  # 100 requests per minute
async def api_excel_grid_data(request: Request):
    """Load static file and return Excel grid data with computed columns."""
    # Try to find the static file
    static_file_path = EXPANSIONS_DIR / "Final_output_scored_static.xlsx"
    
    if not static_file_path.exists():
        # Also check for Final_output_scored.xlsx as fallback
        fallback_file = EXPANSIONS_DIR / "Final_output_scored.xlsx"
        if fallback_file.exists():
            static_file_path = fallback_file
            logging.info(f"✅ Using fallback file: {static_file_path}")
        else:
            # List available Excel files in EXPANSIONS_DIR for better error message
            available_files = []
            if EXPANSIONS_DIR.exists():
                for file in EXPANSIONS_DIR.glob("*.xlsx"):
                    available_files.append(file.name)
            
            error_detail = "Final_output_scored_static.xlsx not found. Please run scoring to generate it."
            if available_files:
                error_detail += f" Available files in expansions directory: {', '.join(sorted(available_files))}"
            else:
                error_detail += f" No Excel files found in {EXPANSIONS_DIR}"
            
            logging.warning(f"❌ Static file not found at: {EXPANSIONS_DIR / 'Final_output_scored_static.xlsx'}")
            logging.warning(f"Available files: {available_files}")
            raise HTTPException(
                status_code=404,
                detail=error_detail
            )
    
    try:
        logging.info(f"📖 Loading Excel grid data from: {static_file_path}")
        grid_data = analysis_engine.get_excel_grid_data_from_excel(static_file_path)
        
        return JSONResponse(grid_data)
    except Exception as e:
        logging.error(f"Error loading excel grid data: {e}")
        logging.exception("Full traceback:")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to load excel grid data: {str(e)}"
        )

@app.post("/api/upload-application-sheet")
async def api_upload_application_sheet(file: UploadFile = File(...)):
    """Upload Application ID sheet and return available tabs."""
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx or .xls) are supported")
    
    try:
        # Save uploaded file
        upload_dir = EXPANSIONS_DIR / "app_id_uploads"
        upload_dir.mkdir(parents=True, exist_ok=True)
        file_path = upload_dir / file.filename
        
        with open(file_path, "wb") as f:
            content = await file.read()
            f.write(content)
        
        # Read Excel file to get sheet names
        df_sheets = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
        tabs = list(df_sheets.keys())
        
        return JSONResponse({
            "success": True,
            "tabs": tabs,
            "file_path": str(file_path)
        })
    except Exception as e:
        logging.error(f"Error uploading application sheet: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to upload application sheet: {str(e)}")

@app.post("/api/copy-tab-to-application-id")
def api_copy_tab_to_application_id(request: CopyTabRequest):
    """Copy a tab from uploaded file to Final_output_curated.xlsx as Application ID sheet."""
    try:
        # Source file
        if request.file_path:
            source_file = pathlib.Path(request.file_path)
        else:
            # Look for the most recent uploaded file
            upload_dir = EXPANSIONS_DIR / "app_id_uploads"
            if not upload_dir.exists():
                raise HTTPException(status_code=404, detail="No uploaded file found")
            files = list(upload_dir.glob("*.xlsx")) + list(upload_dir.glob("*.xls"))
            if not files:
                raise HTTPException(status_code=404, detail="No uploaded file found")
            source_file = max(files, key=lambda p: p.stat().st_mtime)
        
        # Target file
        target_file = EXPANSIONS_DIR / "Final_output_curated.xlsx"
        
        if not target_file.exists():
            raise HTTPException(status_code=404, detail="Final_output_curated.xlsx not found. Please run curation first.")
        
        if not source_file.exists():
            raise HTTPException(status_code=404, detail=f"Source file not found: {source_file}")
        
        # Read the requested sheet from source
        try:
            df = pd.read_excel(source_file, sheet_name=request.tab_name, engine='openpyxl')
        except ValueError:
            raise HTTPException(status_code=404, detail=f"Sheet '{request.tab_name}' not found in uploaded file")
        
        # Append to target Excel file
        with pd.ExcelWriter(target_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name='Application ID', index=False)
        
        logging.info(f"Copied sheet '{request.tab_name}' to Application ID in {target_file}")
        
        return JSONResponse({
            "success": True,
            "message": f"Successfully copied '{request.tab_name}' to Application ID sheet"
        })
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error copying tab to Application ID: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to copy tab: {str(e)}")

@app.get("/api/download")
def download(path: str = Query(..., description="Absolute path within ./data to download")):
    p = pathlib.Path(path)
    if not _safe_within_data(p) or not p.exists() or not p.is_file():
        raise HTTPException(status_code=404, detail="File not found or not allowed")
    return FileResponse(path=str(p), filename=p.name)


@app.post("/api/excel-sheet-info")
async def api_get_excel_sheet_info(file: UploadFile = File(...)):
    """Get information about sheets in an uploaded Excel file."""
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx or .xls) are supported")
    
    try:
        import tempfile
        
        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            content = file.file.read()
            tmp_file.write(content)
            tmp_file_path = pathlib.Path(tmp_file.name)
        
        try:
            # Get sheet info
            sheet_info = analysis_engine.get_excel_sheet_info(tmp_file_path)
            return JSONResponse(sheet_info)
        finally:
            # Clean up temporary file
            if tmp_file_path.exists():
                tmp_file_path.unlink()
                
    except Exception as e:
        logging.error(f"Error getting Excel sheet info: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to get Excel sheet info: {str(e)}")

@app.get("/api/check-excel-file-ai")
async def check_excel_file_ai():
    """
    Check if Final_output_scored_static.xlsx file exists for AI Analysis.
    Returns status indicating if file is found.
    """
    static_file_path = EXPANSIONS_DIR / EXCEL_FILE_NAME
    
    if static_file_path.exists():
        logging.info(f"✅ Excel file found at: {static_file_path}")
        return JSONResponse({
            "success": True,
            "found": True,
            "message": "connected",
            "path": str(static_file_path)
        })
    else:
        logging.warning(f"❌ Excel file not found: {EXCEL_FILE_NAME}")
        return JSONResponse({
            "success": True,
            "found": False,
            "message": "File not found, Refresh the page"
        })

class NLMQueryRequest(BaseModel):
    query: str

class VoiceScopeValidationRequest(BaseModel):
    query: str

# Security Control Mapping (from ComplianceGrid.tsx)
SECURITY_CONTROL_MAPPING = {
    'Overpermissive_Source': {
        'CIS': ['11.3', '12.3', '9.1'],
        'PCI': ['1.2.1', '1.2.2', '1.2.3', '1.2.4'],
        'NIST': ['AC-4', 'AC-6', 'SC-7']
    },
    'Overpermissive_Destination': {
        'CIS': ['11.3', '12.3', '9.1'],
        'PCI': ['1.2.1', '1.2.2', '1.2.3', '1.2.4'],
        'NIST': ['AC-4', 'AC-6', 'SC-7']
    },
    'Service_App': {
        'CIS': ['11.3', '12.7', '9.2', '12.2'],
        'PCI': ['1.2.2', '1.2.3', '1.2.4', '2.2.4'],
        'NIST': ['SC-7', 'SI-4', 'SC-8', 'SC-13']
    },
    'Insecure_Ports': {
        'CIS': ['9.2', '12.2'],
        'PCI': ['1.2.2', '2.2.4'],
        'NIST': ['SC-7', 'SC-8', 'SC-13']
    },
    'Risky_Inbound': {
        'CIS': ['11.3', '12.7'],
        'PCI': ['1.2.2', '1.2.3', '1.2.4'],
        'NIST': ['SC-7', 'SI-4']
    },
    'Risky_Outbound': {
        'CIS': ['11.3', '12.7'],
        'PCI': ['1.2.2', '1.2.3', '1.2.4'],
        'NIST': ['SC-7', 'SI-4']
    },
    'Rule_Usage': {
        'CIS': ['11.2', '12.5'],
        'PCI': ['1.2.1', '1.2.2'],
        'NIST': ['CM-3', 'CM-8']
    },
    'Rule_Description': {
        'CIS': ['11.1', '12.5'],
        'PCI': ['1.2.1', '6.4.2'],
        'NIST': ['PM-9', 'CM-5']
    },
    'Source_User': {
        'CIS': ['14.1', '11.4'],
        'PCI': ['7.2.1', '7.2.4', '1.2.1'],
        'NIST': ['AC-6', 'AC-3']
    },
    'Security_Profile': {
        'CIS': ['10.1', '10.2', '10.5'],
        'PCI': ['5.2.1', '5.2.2', '5.2.3', '5.3'],
        'NIST': ['SI-3', 'SI-4']
    },
    'Shadow_Rule': {
        'CIS': ['11.2', '12.1'],
        'PCI': ['1.2.1', '1.2.2'],
        'NIST': ['CM-3', 'CM-5', 'AC-4']
    },
    'Redundant_Rule': {
        'CIS': ['11.2', '12.5'],
        'PCI': ['1.2.1', '1.2.2'],
        'NIST': ['CM-3', 'CM-5']
    },
    'Consolidation': {
        'CIS': ['11.2', '12.4'],
        'PCI': ['1.2.1', '1.2.2', '2.2.1'],
        'NIST': ['CM-8', 'CM-3']
    }
}

def get_rule_security_control_failures(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Analyze each rule and determine which security controls it fails based on boolean columns.
    Returns a dictionary mapping rule names to their failed controls.
    """
    rule_control_failures = {}
    
    # Find Rule_Name column
    rule_name_col = None
    for col_name in ['Rule_Name', 'Rule Name', 'rule_name', 'Name']:
        if col_name in df.columns:
            rule_name_col = col_name
            break
    
    if not rule_name_col:
        logging.warning("⚠️ Rule_Name column not found, cannot map rules to security controls")
        return {}
    
    # Iterate through each rule
    for idx, row in df.iterrows():
        rule_name = str(row.get(rule_name_col, f'Rule_{idx}'))
        if pd.isna(rule_name) or rule_name.strip() == '':
            rule_name = f'Rule_{idx}'
        
        failed_controls = {
            'CIS': set(),
            'PCI': set(),
            'NIST': set()
        }
        
        # Check each column in the mapping
        for column_name, control_mapping in SECURITY_CONTROL_MAPPING.items():
            # Check if column exists and is True
            if column_name in df.columns:
                column_value = row.get(column_name)
                # Handle boolean values (True/False) or string "True"/"False"
                is_true = False
                if isinstance(column_value, bool):
                    is_true = column_value
                elif isinstance(column_value, str):
                    is_true = column_value.lower() in ['true', '1', 'yes']
                elif pd.notna(column_value):
                    is_true = bool(column_value)
                
                if is_true:
                    # Add all controls for this column
                    if 'CIS' in control_mapping:
                        failed_controls['CIS'].update(control_mapping['CIS'])
                    if 'PCI' in control_mapping:
                        failed_controls['PCI'].update(control_mapping['PCI'])
                    if 'NIST' in control_mapping:
                        failed_controls['NIST'].update(control_mapping['NIST'])
        
        # Convert sets to sorted lists
        rule_control_failures[rule_name] = {
            'CIS': sorted(list(failed_controls['CIS'])),
            'PCI': sorted(list(failed_controls['PCI'])),
            'NIST': sorted(list(failed_controls['NIST']))
        }
    
    return rule_control_failures

def get_policy_analyzer_data_from_excel(excel_file: pathlib.Path) -> Dict[str, Any]:
    """
    Extract policy analyzer results from Excel file if columns exist.
    Returns counts and sample rules for each category.
    """
    if not excel_file.exists():
        return {
            "found": False,
            "shadowRules": {"count": 0, "sampleRules": []},
            "partialShadowRules": {"count": 0, "sampleRules": []},
            "redundantRules": {"count": 0, "sampleRules": []},
            "generalizationRisks": {"count": 0, "sampleRules": []},
            "correlationRisks": {"count": 0, "sampleRules": []},
            "consolidationCandidates": {"count": 0, "sampleRules": []}
        }
    
    try:
        # Read Excel file - try 'Raw Data' first, then 'Raw_Data', then first sheet
        try:
            df = pd.read_excel(excel_file, sheet_name='Raw Data', engine='openpyxl')
        except ValueError:
            try:
                df = pd.read_excel(excel_file, sheet_name='Raw_Data', engine='openpyxl')
            except ValueError:
                # Fallback to first sheet
                df = pd.read_excel(excel_file, sheet_name=0, engine='openpyxl')
        
        # Check if policy analyzer columns exist
        has_shadow = 'Shadow_Rule' in df.columns
        has_partial_shadow = 'Partial_Shadow_Rule' in df.columns
        has_redundant = 'Redundant_Rule' in df.columns
        has_generalization = 'Generalization_Risk' in df.columns
        has_correlation = 'Correlation_Risk' in df.columns
        has_consolidation = 'Consolidation_Candidate' in df.columns
        
        if not any([has_shadow, has_partial_shadow, has_redundant, has_generalization, has_correlation, has_consolidation]):
            # Policy analyzer hasn't been run on this file
            return {
                "found": False,
                "shadowRules": {"count": 0, "sampleRules": []},
                "partialShadowRules": {"count": 0, "sampleRules": []},
                "redundantRules": {"count": 0, "sampleRules": []},
                "generalizationRisks": {"count": 0, "sampleRules": []},
                "correlationRisks": {"count": 0, "sampleRules": []},
                "consolidationCandidates": {"count": 0, "sampleRules": []}
            }
        
        result = {"found": True}
        
        # Shadow Rules
        if has_shadow:
            shadow_df = df[df['Shadow_Rule'] == True].copy()
            shadow_count = len(shadow_df)
            shadow_samples = []
            if shadow_count > 0:
                # Get top 5-10 sample rules with reasons
                for idx, row in shadow_df.head(10).iterrows():
                    rule_name = str(row.get('Rule_Name', 'Unknown'))
                    reason = str(row.get('Shadow_Reason', 'Fully covered by earlier rule'))
                    shadow_samples.append({
                        "name": rule_name,
                        "reason": reason
                    })
            result["shadowRules"] = {"count": shadow_count, "sampleRules": shadow_samples}
        else:
            result["shadowRules"] = {"count": 0, "sampleRules": []}
        
        # Partial Shadow Rules
        if has_partial_shadow:
            partial_shadow_df = df[df['Partial_Shadow_Rule'] == True].copy()
            partial_shadow_count = len(partial_shadow_df)
            partial_shadow_samples = []
            if partial_shadow_count > 0:
                for idx, row in partial_shadow_df.head(10).iterrows():
                    rule_name = str(row.get('Rule_Name', 'Unknown'))
                    reason = str(row.get('Shadow_Reason', 'Partially covered by earlier rule'))
                    partial_shadow_samples.append({
                        "name": rule_name,
                        "reason": reason
                    })
            result["partialShadowRules"] = {"count": partial_shadow_count, "sampleRules": partial_shadow_samples}
        else:
            result["partialShadowRules"] = {"count": 0, "sampleRules": []}
        
        # Redundant Rules
        if has_redundant:
            redundant_df = df[df['Redundant_Rule'] == True].copy()
            redundant_count = len(redundant_df)
            redundant_samples = []
            if redundant_count > 0:
                for idx, row in redundant_df.head(10).iterrows():
                    rule_name = str(row.get('Rule_Name', 'Unknown'))
                    reason = str(row.get('Redundancy_Reason', 'Same action and fully covered by earlier rule'))
                    redundant_samples.append({
                        "name": rule_name,
                        "reason": reason
                    })
            result["redundantRules"] = {"count": redundant_count, "sampleRules": redundant_samples}
        else:
            result["redundantRules"] = {"count": 0, "sampleRules": []}
        
        # Generalization Risks
        if has_generalization:
            generalization_df = df[df['Generalization_Risk'] == True].copy()
            generalization_count = len(generalization_df)
            generalization_samples = []
            if generalization_count > 0:
                for idx, row in generalization_df.head(10).iterrows():
                    rule_name = str(row.get('Rule_Name', 'Unknown'))
                    reason = str(row.get('Generalization_Reason', 'Generalizes later rule'))
                    generalization_samples.append({
                        "name": rule_name,
                        "reason": reason
                    })
            result["generalizationRisks"] = {"count": generalization_count, "sampleRules": generalization_samples}
        else:
            result["generalizationRisks"] = {"count": 0, "sampleRules": []}
        
        # Correlation Risks
        if has_correlation:
            correlation_df = df[df['Correlation_Risk'] == True].copy()
            correlation_count = len(correlation_df)
            correlation_samples = []
            if correlation_count > 0:
                for idx, row in correlation_df.head(10).iterrows():
                    rule_name = str(row.get('Rule_Name', 'Unknown'))
                    reason = str(row.get('Correlation_Reason', 'Partially overlaps with different actions'))
                    correlation_samples.append({
                        "name": rule_name,
                        "reason": reason
                    })
            result["correlationRisks"] = {"count": correlation_count, "sampleRules": correlation_samples}
        else:
            result["correlationRisks"] = {"count": 0, "sampleRules": []}
        
        # Consolidation Candidates
        if has_consolidation:
            consolidation_df = df[df['Consolidation_Candidate'] == True].copy()
            consolidation_count = len(consolidation_df)
            consolidation_samples = []
            if consolidation_count > 0:
                # Group by Consolidation_Key to show related rules together
                consolidation_groups = {}
                for idx, row in consolidation_df.iterrows():
                    key = str(row.get('Consolidation_Key', 'Unknown'))
                    rule_name = str(row.get('Rule_Name', 'Unknown'))
                    reason = str(row.get('Consolidation_Reason', 'Can be consolidated'))
                    if key not in consolidation_groups:
                        consolidation_groups[key] = {
                            "rules": [],
                            "reason": reason
                        }
                    consolidation_groups[key]["rules"].append(rule_name)
                
                # Convert to sample format (top 10 groups)
                for key, group_data in list(consolidation_groups.items())[:10]:
                    consolidation_samples.append({
                        "rules": group_data["rules"],
                        "reason": group_data["reason"]
                    })
            result["consolidationCandidates"] = {"count": consolidation_count, "sampleRules": consolidation_samples}
        else:
            result["consolidationCandidates"] = {"count": 0, "sampleRules": []}
        
        return result
        
    except Exception as e:
        logging.error(f"Error extracting policy analyzer data: {e}")
        logging.exception("Full traceback:")
        # Return empty structure on error
        return {
            "found": False,
            "shadowRules": {"count": 0, "sampleRules": []},
            "partialShadowRules": {"count": 0, "sampleRules": []},
            "redundantRules": {"count": 0, "sampleRules": []},
            "generalizationRisks": {"count": 0, "sampleRules": []},
            "correlationRisks": {"count": 0, "sampleRules": []},
            "consolidationCandidates": {"count": 0, "sampleRules": []}
        }

def get_detailed_rule_examples_from_excel(excel_file: pathlib.Path, max_examples_per_category: int = 5) -> Dict[str, Any]:
    """
    Extract detailed rule examples from Excel file with full rule attributes.
    Returns examples for different categories to help LLM provide specific responses.
    """
    if not excel_file.exists():
        logging.error(f"❌ Excel file not found: {excel_file}")
        return {"found": False, "examples": {}}
    
    logging.info(f"📂 Reading Excel file: {excel_file.name}")
    
    try:
        # Read Excel file - try 'Raw Data' first, then 'Raw_Data', then first sheet
        try:
            df = pd.read_excel(excel_file, sheet_name='Raw Data', engine='openpyxl')
            logging.info(f"✅ Read sheet 'Raw Data'")
        except ValueError:
            try:
                df = pd.read_excel(excel_file, sheet_name='Raw_Data', engine='openpyxl')
                logging.info(f"✅ Read sheet 'Raw_Data'")
            except ValueError:
                df = pd.read_excel(excel_file, sheet_name=0, engine='openpyxl')
                logging.info(f"✅ Read first sheet (index 0)")
        
        examples = {}
        
        # Log available columns to help debug
        logging.info(f"📋 Available columns in Excel ({len(df.columns)} total): {list(df.columns)[:20]}...")  # Log first 20 columns
        
        # Find Rule_Name column with variations
        rule_name_col = None
        possible_names = ['Rule_Name', 'Rule Name', 'rule_name', 'RULE_NAME', 'Name', 'RuleName']
        for col_name in possible_names:
            if col_name in df.columns:
                rule_name_col = col_name
                logging.info(f"✅ Found rule name column: '{col_name}'")
                break
        
        if rule_name_col is None:
            logging.warning(f"⚠️ Rule_Name column not found in standard names. Searching for similar columns...")
            # Try to find any column with 'rule' and 'name' in it (case-insensitive)
            for col in df.columns:
                col_lower = str(col).lower()
                if 'rule' in col_lower and 'name' in col_lower:
                    rule_name_col = col
                    logging.info(f"✅ Found similar column: '{col}'")
                    break
        
        if rule_name_col is None:
            logging.error(f"❌ Could not find Rule_Name column. Available columns: {list(df.columns)}")
            logging.error(f"❌ Will use 'Unknown' for all rule names. This will prevent LLM from referencing actual rule names.")
            rule_name_col = 'Unknown_Column'  # Will trigger fallback
        
        # Helper to format a rule row into a readable string
        def format_rule(row, include_score=True):
            # Try to get rule name from the identified column
            if rule_name_col and rule_name_col != 'Unknown_Column' and rule_name_col in row.index:
                rule_name = str(row[rule_name_col]) if pd.notna(row[rule_name_col]) else 'Unknown'
            else:
                # Fallback: try common variations
                rule_name = str(row.get('Rule_Name', row.get('Name', row.get('Rule Name', 'Unknown'))))
            
            # Log if we're using fallback (only log once per category to avoid spam)
            if rule_name == 'Unknown' and rule_name_col == 'Unknown_Column':
                pass  # Already logged above
            
            source = str(row.get('Source Address', row.get('Src Address', 'N/A')))
            dest = str(row.get('Destination Address', row.get('Dst Address', 'N/A')))
            service = str(row.get('Service', 'N/A'))
            action = str(row.get('Action', 'N/A'))
            score = ""
            if include_score and 'Score_Total' in row:
                try:
                    score_val = pd.to_numeric(row['Score_Total'], errors='coerce')
                    if pd.notna(score_val):
                        score = f", Score: {int(score_val)}"
                except:
                    pass
            
            return f"{rule_name}: Source={source}, Destination={dest}, Service={service}, Action={action}{score}"
        
        # 1. Risky Inbound Internet Rules
        if 'Risky_Inbound' in df.columns:
            risky_inbound = df[df['Risky_Inbound'].astype(str).str.contains('True', case=False, na=False)]
            if len(risky_inbound) > 0:
                examples['riskyInbound'] = [format_rule(row) for _, row in risky_inbound.head(max_examples_per_category).iterrows()]
        
        # 2. Risky Outbound Internet Rules
        if 'Risky_Outbound' in df.columns:
            risky_outbound = df[df['Risky_Outbound'].astype(str).str.contains('True', case=False, na=False)]
            if len(risky_outbound) > 0:
                examples['riskyOutbound'] = [format_rule(row) for _, row in risky_outbound.head(max_examples_per_category).iterrows()]
        
        # 3. Source Any Rules
        if 'Src_IsAny' in df.columns:
            src_any = df[df['Src_IsAny'].astype(str).str.contains('True', case=False, na=False)]
            if len(src_any) > 0:
                examples['sourceAny'] = [format_rule(row) for _, row in src_any.head(max_examples_per_category).iterrows()]
        
        # 4. Destination Any Rules
        if 'Dst_IsAny' in df.columns:
            dst_any = df[df['Dst_IsAny'].astype(str).str.contains('True', case=False, na=False)]
            if len(dst_any) > 0:
                examples['destinationAny'] = [format_rule(row) for _, row in dst_any.head(max_examples_per_category).iterrows()]
        
        # 5. Insecure Ports
        if 'Service_Insecure_Match' in df.columns:
            insecure = df[df['Service_Insecure_Match'].astype(str).str.contains('True', case=False, na=False)]
            if len(insecure) > 0:
                examples['insecurePorts'] = [format_rule(row) for _, row in insecure.head(max_examples_per_category).iterrows()]
        
        # 6. High Risk Rules (Score >= 100)
        if 'Score_Total' in df.columns:
            score_col = pd.to_numeric(df['Score_Total'], errors='coerce')
            high_risk = df[score_col >= 100].copy()
            if len(high_risk) > 0:
                high_risk = high_risk.sort_values('Score_Total', ascending=False, na_position='last')
                examples['highRisk'] = [format_rule(row) for _, row in high_risk.head(max_examples_per_category).iterrows()]
        
        # 7. Critical Risk Rules (Score >= 175)
        if 'Score_Total' in df.columns:
            score_col = pd.to_numeric(df['Score_Total'], errors='coerce')
            critical_risk = df[score_col >= 175].copy()
            if len(critical_risk) > 0:
                critical_risk = critical_risk.sort_values('Score_Total', ascending=False, na_position='last')
                examples['criticalRisk'] = [format_rule(row) for _, row in critical_risk.head(max_examples_per_category).iterrows()]
        
        # 8. Service Broad Rules
        if 'Service_Any_OR_RangeGt1000' in df.columns:
            service_broad = df[df['Service_Any_OR_RangeGt1000'].astype(str).str.contains('True', case=False, na=False)]
            if len(service_broad) > 0:
                examples['serviceBroad'] = [format_rule(row) for _, row in service_broad.head(max_examples_per_category).iterrows()]
        
        # 9. Missing Security Profile
        if 'Profile_Scoring' in df.columns:
            missing_profile = df[df['Profile_Scoring'].astype(str).str.contains('True', case=False, na=False)]
            if len(missing_profile) > 0:
                examples['missingProfile'] = [format_rule(row) for _, row in missing_profile.head(max_examples_per_category).iterrows()]
        
        # 10. Missing Log Forwarding
        if 'Options_Scoring' in df.columns:
            missing_log = df[df['Options_Scoring'].astype(str).str.contains('True', case=False, na=False)]
            if len(missing_log) > 0:
                examples['missingLogForwarding'] = [format_rule(row) for _, row in missing_log.head(max_examples_per_category).iterrows()]
        
        # 11. Shadow Rules (Policy Analyzer)
        if 'Shadow_Rule' in df.columns:
            shadow_rules = df[df['Shadow_Rule'] == True].copy()
            if len(shadow_rules) > 0:
                shadow_examples = []
                for _, row in shadow_rules.head(max_examples_per_category).iterrows():
                    rule_str = format_rule(row)
                    reason = str(row.get('Shadow_Reason', 'Fully covered by earlier rule'))
                    shadow_examples.append(f"{rule_str} | Reason: {reason}")
                examples['shadowRules'] = shadow_examples
        
        # 12. Partial Shadow Rules (Policy Analyzer)
        if 'Partial_Shadow_Rule' in df.columns:
            partial_shadow = df[df['Partial_Shadow_Rule'] == True].copy()
            if len(partial_shadow) > 0:
                partial_shadow_examples = []
                for _, row in partial_shadow.head(max_examples_per_category).iterrows():
                    rule_str = format_rule(row)
                    reason = str(row.get('Shadow_Reason', 'Partially covered by earlier rule'))
                    partial_shadow_examples.append(f"{rule_str} | Reason: {reason}")
                examples['partialShadowRules'] = partial_shadow_examples
        
        # 13. Redundant Rules (Policy Analyzer)
        if 'Redundant_Rule' in df.columns:
            redundant = df[df['Redundant_Rule'] == True].copy()
            if len(redundant) > 0:
                redundant_examples = []
                for _, row in redundant.head(max_examples_per_category).iterrows():
                    rule_str = format_rule(row)
                    reason = str(row.get('Redundancy_Reason', 'Same action and fully covered'))
                    redundant_examples.append(f"{rule_str} | Reason: {reason}")
                examples['redundantRules'] = redundant_examples
        
        # 14. Generalization Risks (Policy Analyzer)
        if 'Generalization_Risk' in df.columns:
            generalization = df[df['Generalization_Risk'] == True].copy()
            if len(generalization) > 0:
                generalization_examples = []
                for _, row in generalization.head(max_examples_per_category).iterrows():
                    rule_str = format_rule(row)
                    reason = str(row.get('Generalization_Reason', 'Generalizes later rule'))
                    generalization_examples.append(f"{rule_str} | Reason: {reason}")
                examples['generalizationRisks'] = generalization_examples
        
        # 15. Correlation Risks (Policy Analyzer)
        if 'Correlation_Risk' in df.columns:
            correlation = df[df['Correlation_Risk'] == True].copy()
            if len(correlation) > 0:
                correlation_examples = []
                for _, row in correlation.head(max_examples_per_category).iterrows():
                    rule_str = format_rule(row)
                    reason = str(row.get('Correlation_Reason', 'Overlaps with different actions'))
                    correlation_examples.append(f"{rule_str} | Reason: {reason}")
                examples['correlationRisks'] = correlation_examples
        
        # 16. Consolidation Candidates (Policy Analyzer)
        if 'Consolidation_Candidate' in df.columns:
            consolidation = df[df['Consolidation_Candidate'] == True].copy()
            if len(consolidation) > 0:
                # Group by Consolidation_Key
                consolidation_groups = {}
                for _, row in consolidation.iterrows():
                    key = str(row.get('Consolidation_Key', 'Unknown'))
                    rule_str = format_rule(row, include_score=False)  # Don't include score for consolidation
                    reason = str(row.get('Consolidation_Reason', 'Can be consolidated'))
                    if key not in consolidation_groups:
                        consolidation_groups[key] = {
                            "rules": [],
                            "reason": reason
                        }
                    consolidation_groups[key]["rules"].append(rule_str)
                
                # Format consolidation groups
                consolidation_examples = []
                for key, group_data in list(consolidation_groups.items())[:max_examples_per_category]:
                    rules_list = group_data["rules"]
                    rules_str = "; ".join(rules_list[:3])
                    if len(rules_list) > 3:
                        rules_str += f" (+{len(rules_list) - 3} more rules)"
                    consolidation_examples.append(f"Group: {rules_str} | Reason: {group_data['reason']}")
                examples['consolidationCandidates'] = consolidation_examples
        
        return {"found": True, "examples": examples}
        
    except Exception as e:
        logging.error(f"Error extracting detailed rule examples: {e}")
        logging.exception("Full traceback:")
        return {"found": False, "examples": {}}

def get_reordering_suggestions_from_excel(excel_file: pathlib.Path, max_suggestions: int = 10) -> Dict[str, Any]:
    """
    Extract reordering suggestions from Excel file if Suggested_Order column exists.
    Returns suggestions with rule names, from/to positions, and reasons.
    """
    if not excel_file.exists():
        logging.error(f"❌ Excel file not found: {excel_file}")
        return {"found": False, "suggestions": []}
    
    try:
        # Read Excel file
        try:
            df = pd.read_excel(excel_file, sheet_name='Raw Data', engine='openpyxl')
        except ValueError:
            try:
                df = pd.read_excel(excel_file, sheet_name='Raw_Data', engine='openpyxl')
            except ValueError:
                df = pd.read_excel(excel_file, sheet_name=0, engine='openpyxl')
        
        # Check if Suggested_Order column exists
        if 'Suggested_Order' not in df.columns or 'Rule_Order' not in df.columns:
            logging.info(f"ℹ️ Suggested_Order or Rule_Order column not found. Reordering suggestions not available.")
            return {"found": False, "suggestions": []}
        
        # Find Rule_Name column
        rule_name_col = None
        for col_name in ['Rule_Name', 'Rule Name', 'rule_name', 'Name']:
            if col_name in df.columns:
                rule_name_col = col_name
                break
        
        if rule_name_col is None:
            logging.warning(f"⚠️ Rule_Name column not found for reordering suggestions")
            return {"found": False, "suggestions": []}
        
        # Find differences between Rule_Order and Suggested_Order
        suggestions = []
        for idx, row in df.iterrows():
            rule_order = pd.to_numeric(row.get('Rule_Order', 0), errors='coerce')
            suggested_order = pd.to_numeric(row.get('Suggested_Order', 0), errors='coerce')
            
            if pd.notna(rule_order) and pd.notna(suggested_order) and rule_order != suggested_order:
                rule_name = str(row.get(rule_name_col, 'Unknown'))
                reason = str(row.get('Shadow_Reason', row.get('Redundancy_Reason', row.get('Consolidation_Reason', 
                    row.get('Generalization_Reason', 'Rule should be reordered for better policy structure')))))
                
                suggestions.append({
                    "rule": rule_name,
                    "from": int(rule_order),
                    "to": int(suggested_order),
                    "reason": reason[:200]  # Limit reason length
                })
        
        # Sort by suggested order
        suggestions.sort(key=lambda x: x['to'])
        
        logging.info(f"✅ Found {len(suggestions)} reordering suggestions")
        return {"found": True, "suggestions": suggestions[:max_suggestions]}
        
    except Exception as e:
        logging.error(f"Error extracting reordering suggestions: {e}")
        logging.exception("Full traceback:")
        return {"found": False, "suggestions": []}

@app.post("/api/voice/validate-scope")
async def validate_voice_scope(request: VoiceScopeValidationRequest):
    """
    Validate if a voice query is within scope before sending to LLM.
    Provides fast frontend validation.
    """
    try:
        validation_result = validate_query_scope(request.query)
        return JSONResponse({
            "success": True,
            "in_scope": validation_result["in_scope"],
            "reason": validation_result["reason"],
            "confidence": validation_result["confidence"],
            "message": get_out_of_scope_message() if not validation_result["in_scope"] else None
        })
    except Exception as e:
        logging.error(f"❌ Error validating scope: {str(e)}")
        logging.exception("Full traceback:")
        raise HTTPException(status_code=500, detail=f"Scope validation failed: {e}")

@app.post("/api/nlm-query")
@rate_limit(max_requests=30, window_seconds=60)  # 30 requests per minute
async def nlm_query(request: Request, query_request: NLMQueryRequest):
    """
    Process Natural Language queries using LLM model.
    Uses AnalysisEngine to provide structured dashboard and analysis data.
    """
    query = query_request.query.strip()
    
    if not query:
        raise HTTPException(status_code=400, detail="Query is required")
    
    # Optional: Quick scope validation (LLM will also validate as secondary check)
    validation_result = validate_query_scope(query)
    if not validation_result["in_scope"] and validation_result["confidence"] == "high":
        # High confidence out-of-scope, return early
        logging.info(f"🚫 Query rejected by scope validation: {validation_result['reason']}")
        return JSONResponse({
            "success": False,
            "response": get_out_of_scope_message(),
            "scope_rejected": True,
            "reason": validation_result["reason"]
        })
    
    # Check if Excel file exists
    static_file_path = EXPANSIONS_DIR / EXCEL_FILE_NAME
    if not static_file_path.exists():
        raise HTTPException(status_code=404, detail="Excel file not found. Please refresh the page.")
    
    try:
        # ===== QUERY ROUTING: Classify query and determine data requirements =====
        query_intent = classify_query(query)
        data_reqs = get_data_requirements(query_intent, query)
        log_data_requirements(data_reqs, query)
        
        # ===== CONDITIONAL DATA LOADING: Load only what's needed =====
        
        # Always load basic analysis data (lightweight)
        analysis_data = None
        dashboard_data = None
        if data_reqs.needs_analysis_data:
            logging.debug(f"📊 Loading analysis data...")
            analysis_data = analysis_engine.get_analysis_data_from_excel(static_file_path)
        
        if data_reqs.needs_dashboard_data:
            logging.debug(f"📊 Loading dashboard data...")
            dashboard_data = analysis_engine.get_dashboard_data_from_excel(static_file_path)
        
        # Conditionally load policy analyzer data
        policy_analyzer_data = {"found": False}
        if data_reqs.needs_policy_data:
            logging.debug(f"🔍 Loading policy analyzer data...")
            policy_analyzer_data = get_policy_analyzer_data_from_excel(static_file_path)
        
        # Conditionally load rule examples (only for specific categories if specified)
        rule_examples = {"found": False, "examples": {}}
        if data_reqs.needs_rule_examples:
            logging.debug(f"📋 Loading rule examples...")
            try:
                # If specific categories requested, load only those
                if data_reqs.specific_categories:
                    logging.debug(f"📋 Loading specific categories: {data_reqs.specific_categories}")
                    # Load all examples first, then filter
                    all_examples = get_detailed_rule_examples_from_excel(static_file_path, max_examples_per_category=2)
                    if all_examples.get('found'):
                        filtered_examples = {}
                        for category in data_reqs.specific_categories:
                            if category in all_examples.get('examples', {}):
                                filtered_examples[category] = all_examples['examples'][category]
                        rule_examples = {"found": len(filtered_examples) > 0, "examples": filtered_examples}
                        logging.debug(f"✅ Loaded {len(filtered_examples)} specific categories")
                else:
                    # Load all examples
                    rule_examples = get_detailed_rule_examples_from_excel(static_file_path, max_examples_per_category=2)
                    if rule_examples.get('found'):
                        example_categories = list(rule_examples.get('examples', {}).keys())
                        total_examples = sum(len(examples) for examples in rule_examples.get('examples', {}).values())
                        logging.debug(f"✅ Loaded {total_examples} rule examples across {len(example_categories)} categories")
            except Exception as e:
                logging.error(f"❌ Error loading rule examples: {e}")
                rule_examples = {"found": False, "examples": {}}
        
        # Conditionally load reordering suggestions
        reordering_suggestions = {"found": False, "suggestions": []}
        if data_reqs.needs_reordering:
            logging.debug(f"🔀 Loading reordering suggestions...")
            reordering_suggestions = get_reordering_suggestions_from_excel(static_file_path, max_suggestions=3)
        
        # Conditionally load firewall distribution
        firewall_analysis_distribution = []
        if data_reqs.needs_firewall_distribution:
            logging.debug(f"📊 Loading firewall distribution...")
            try:
                # Get risk distribution from dashboard data (already loaded)
                risk_dist = dashboard_data.get('firewallRiskDistribution', []) if dashboard_data else []
                
                # Get policy analysis distribution from Excel
                try:
                    df_temp = pd.read_excel(static_file_path, engine='openpyxl', nrows=1)
                    firewall_cols = ['Source_File', 'Source File', 'Firewall', 'Device', 'Source']
                    has_firewall_col = any(col in df_temp.columns for col in firewall_cols)
                    
                    if has_firewall_col:
                        df_full = pd.read_excel(static_file_path, engine='openpyxl')
                        policy_dist = analysis_engine.get_firewall_analysis_distribution(df_full)
                    else:
                        policy_dist = []
                        logging.warning(f"⚠️ No firewall column found. Checked: {firewall_cols}")
                except Exception as e:
                    logging.warning(f"⚠️ Could not load policy distribution: {e}")
                    policy_dist = []
                
                # Merge risk and policy distributions
                merged_data = {}
                
                logging.info(f"📊 Risk dist count: {len(risk_dist)}, Policy dist count: {len(policy_dist)}")
                if risk_dist:
                    logging.debug(f"📊 Sample risk data: {risk_dist[0]}")
                if policy_dist:
                    logging.debug(f"📊 Sample policy data: {policy_dist[0]}")
                
                # Add risk data first
                for item in risk_dist:
                    fw = item.get('firewall')
                    if fw:
                        merged_data[fw] = item.copy()
                
                # Merge policy data
                for item in policy_dist:
                    fw = item.get('firewall')
                    if fw:
                        if fw in merged_data:
                            logging.debug(f"📊 Merging policy data for {fw}: {item}")
                            merged_data[fw].update(item)
                        else:
                            merged_data[fw] = item
                
                # Convert back to list and sort by High + Critical Risk count (descending)
                firewall_analysis_distribution = list(merged_data.values())
                firewall_analysis_distribution.sort(key=lambda x: x.get('high', 0) + x.get('critical', 0), reverse=True)
                
                if firewall_analysis_distribution:
                    logging.info(f"✅ Loaded merged distribution for {len(firewall_analysis_distribution)} firewalls")
                    logging.debug(f"🔥 First firewall data: {firewall_analysis_distribution[0]}")
                else:
                    logging.warning("⚠️ Firewall distribution is empty!")
                    if dashboard_data:
                        logging.warning(f"📋 Dashboard data keys: {list(dashboard_data.keys())}")
                        logging.warning(f"📋 Risk dist from dashboard: {risk_dist}")
            except Exception as e:
                logging.warning(f"⚠️ Could not load firewall distribution: {e}")
                logging.exception("Full traceback:")
        
        # Conditionally load security control failures
        rule_control_failures = {}
        if data_reqs.needs_security_controls:
            logging.debug(f"🔒 Loading security control failures...")
            try:
                try:
                    df_for_controls = pd.read_excel(static_file_path, sheet_name='Raw Data', engine='openpyxl')
                except ValueError:
                    try:
                        df_for_controls = pd.read_excel(static_file_path, sheet_name='Raw_Data', engine='openpyxl')
                    except ValueError:
                        df_for_controls = pd.read_excel(static_file_path, sheet_name=0, engine='openpyxl')
                
                rule_control_failures = get_rule_security_control_failures(df_for_controls)
                
                if rule_control_failures:
                    logging.debug(f"✅ Loaded security control failures for {len(rule_control_failures)} rules")
            except Exception as e:
                logging.warning(f"⚠️ Could not load security control failures: {e}")
        
        # Log consolidated summary instead of individual steps
        policy_summary = ""
        if policy_analyzer_data.get('found'):
            policy_summary = f"Policy: Shadow={policy_analyzer_data.get('shadowRules', {}).get('count', 0)}, Redundant={policy_analyzer_data.get('redundantRules', {}).get('count', 0)}"
        
        security_summary = ""
        if rule_control_failures and len(rule_control_failures) > 0:
            # Count rules with failures
            rules_with_failures = sum(1 for failures in rule_control_failures.values() 
                                     if len(failures.get('CIS', [])) > 0 or 
                                        len(failures.get('PCI', [])) > 0 or 
                                        len(failures.get('NIST', [])) > 0)
            security_summary = f"Security Controls: {rules_with_failures} rules with control failures (from ComplianceGrid mapping)"
        
        total_rules = analysis_data.get('totalRules', 0) if analysis_data else 0
        logging.info(f"📊 Data loaded | Intent: {query_intent.value.upper()} | Rules: {total_rules} | {policy_summary} | {security_summary}")
        
        # Get LLM config (needed for both COUNT and non-COUNT queries)
        llm_config = get_llm_config()
        
        # ===== INTENT-SPECIFIC PROMPT BUILDING =====
        # For COUNT queries, use minimal prompt to fit within token limits
        if query_intent == QueryIntent.COUNT:
            from app.prompt_builder import build_minimal_prompt_for_count
            prompt = build_minimal_prompt_for_count(query, analysis_data, dashboard_data)
            logging.info(f"✅ Using minimal COUNT prompt ({len(prompt)} chars)")
        elif query_intent == QueryIntent.SPECIFIC_RULES:
            from app.prompt_builder import build_minimal_prompt_for_specific_rules
            prompt = build_minimal_prompt_for_specific_rules(query, analysis_data, rule_examples, data_reqs.specific_categories)
            logging.info(f"✅ Using minimal SPECIFIC prompt ({len(prompt)} chars)")
        elif query_intent == QueryIntent.OVERVIEW:
            from app.prompt_builder import build_minimal_prompt_for_overview
            prompt = build_minimal_prompt_for_overview(query, analysis_data, dashboard_data, rule_examples)
            logging.info(f"✅ Using minimal OVERVIEW prompt ({len(prompt)} chars)")
        elif query_intent == QueryIntent.COMPARISON:
            from app.prompt_builder import build_minimal_prompt_for_comparison
            prompt = build_minimal_prompt_for_comparison(query, analysis_data, firewall_analysis_distribution)
            logging.info(f"✅ Using minimal COMPARISON prompt ({len(prompt)} chars)")
        else:
            # For other queries, build comprehensive data summary
            # Build comprehensive data summary for AI
            data_summary = f"""
=== FIREWALL POLICY ANALYSIS SUMMARY ===

📊 OVERALL STATISTICS:
- Total Rules: {analysis_data.get('totalRules', 0):,}
- Average Score: {analysis_data.get('averageScore', 0):.2f}
- High Risk Rules: {analysis_data.get('highRisk', 0):,}
- Medium Risk Rules: {analysis_data.get('mediumRisk', 0):,}
- Low Risk Rules: {analysis_data.get('lowRisk', 0):,}
- Insecure Ports: {analysis_data.get('insecurePortCount', 0):,}
- Source User Not Used: {analysis_data.get('sourceUserNotUsed', 0):,}

🛡️ RISK DASHBOARD:
"""
            if dashboard_data.get('riskDashboard') and dashboard_data['riskDashboard'].get('found'):
                risk = dashboard_data['riskDashboard']
                data_summary += f"""
- Critical Risk (175-200): {risk.get('critical', 0):,} rules
- High Risk (100-174): {risk.get('high', 0):,} rules
- Medium Risk (50-99): {risk.get('medium', 0):,} rules
- Low Risk (1-49): {risk.get('low', 0):,} rules
- No Risk (0): {risk.get('none', 0):,} rules
"""
            else:
                data_summary += "- Risk Dashboard data not available\n"
        
            data_summary += """
    🚨 OVERPERMISSIVE RULES:
    """
            if dashboard_data.get('overpermissive'):
                op = dashboard_data['overpermissive']
                data_summary += f"""
    Source Issues:
    - Source Any: {op.get('sourceAny', {}).get('count', 0):,} rules"""
                if rule_examples.get('found') and rule_examples.get('examples', {}).get('sourceAny'):
                    data_summary += "\n  Example Source Any Rules:\n"
                    for i, example in enumerate(rule_examples['examples']['sourceAny'], 1):
                        data_summary += f"  {i}. {example}\n"
                data_summary += f"""
    - Source CIDR ≤16: {op.get('srcCidrLe16', {}).get('count', 0):,} rules
    - Source CIDR 17-22: {op.get('srcCidr17_22', {}).get('count', 0):,} rules

    Destination Issues:
    - Destination Any: {op.get('destinationAny', {}).get('count', 0):,} rules"""
                if rule_examples.get('found') and rule_examples.get('examples', {}).get('destinationAny'):
                    data_summary += "\n  Example Destination Any Rules:\n"
                    for i, example in enumerate(rule_examples['examples']['destinationAny'], 1):
                        data_summary += f"  {i}. {example}\n"
                data_summary += f"""
    - Destination CIDR ≤16: {op.get('dstCidrLe16', {}).get('count', 0):,} rules
    - Destination CIDR 17-22: {op.get('dstCidr17_22', {}).get('count', 0):,} rules

    Service & Application:
    - Service is Broad: {op.get('serviceBroad', {}).get('count', 0):,} rules"""
                if rule_examples.get('found') and rule_examples.get('examples', {}).get('serviceBroad'):
                    data_summary += "\n  Example Service Broad Rules:\n"
                    for i, example in enumerate(rule_examples['examples']['serviceBroad'], 1):
                        data_summary += f"  {i}. {example}\n"
                data_summary += f"""
    - Application is Broad: {op.get('appBroad', {}).get('count', 0):,} rules
    - Insecure Ports: {op.get('insecurePorts', {}).get('count', 0):,} rules"""
                if rule_examples.get('found') and rule_examples.get('examples', {}).get('insecurePorts'):
                    data_summary += "\n  Example Insecure Port Rules:\n"
                    for i, example in enumerate(rule_examples['examples']['insecurePorts'], 1):
                        data_summary += f"  {i}. {example}\n"
                data_summary += f"""

    Contextual Risk:
    - Risky Inbound Internet: {op.get('riskyInbound', {}).get('count', 0):,} rules"""
                if rule_examples.get('found') and rule_examples.get('examples', {}).get('riskyInbound'):
                    data_summary += "\n  Example Risky Inbound Rules:\n"
                    for i, example in enumerate(rule_examples['examples']['riskyInbound'], 1):
                        data_summary += f"  {i}. {example}\n"
                data_summary += f"""
    - Risky Outbound Internet: {op.get('riskyOutbound', {}).get('count', 0):,} rules"""
                if rule_examples.get('found') and rule_examples.get('examples', {}).get('riskyOutbound'):
                    data_summary += "\n  Example Risky Outbound Rules:\n"
                    for i, example in enumerate(rule_examples['examples']['riskyOutbound'], 1):
                        data_summary += f"  {i}. {example}\n"
                data_summary += f"""
    - Migrate to App-ID (Insecure): {op.get('migrateInsecure', {}).get('count', 0):,} rules
    - Migrate Other Ports: {op.get('migrateOtherPorts', {}).get('count', 0):,} rules

    Hygiene Risk:
    - Source Zone Any: {op.get('srcZoneIsAny', {}).get('count', 0):,} rules
    - Destination Zone Any: {op.get('dstZoneIsAny', {}).get('count', 0):,} rules
    - Unused Rules: {op.get('ruleUsageScoring', {}).get('count', 0):,} rules
    - Missing Security Profile: {op.get('profileScoring', {}).get('count', 0):,} rules"""
                if rule_examples.get('found') and rule_examples.get('examples', {}).get('missingProfile'):
                    data_summary += "\n  Example Missing Security Profile Rules:\n"
                    for i, example in enumerate(rule_examples['examples']['missingProfile'], 1):
                        data_summary += f"  {i}. {example}\n"
                data_summary += f"""
    - Missing Log Forwarding: {op.get('optionsScoring', {}).get('count', 0):,} rules"""
                if rule_examples.get('found') and rule_examples.get('examples', {}).get('missingLogForwarding'):
                    data_summary += "\n  Example Missing Log Forwarding Rules:\n"
                    for i, example in enumerate(rule_examples['examples']['missingLogForwarding'], 1):
                        data_summary += f"  {i}. {example}\n"
                data_summary += f"""
    - Missing Description: {op.get('ruleUsageDescriptionScoring', {}).get('count', 0):,} rules
    - Source User Not Used: {op.get('sourceUserScoring', {}).get('count', 0):,} rules
    """
        
            # Combined Risk Data
            if dashboard_data.get('combinedRiskInbound') and dashboard_data['combinedRiskInbound'].get('found'):
                cri = dashboard_data['combinedRiskInbound']
                data_summary += f"""
    ⚠️ COMBINED RISK - RISKY INBOUND + OVERPERMISSIVE:
    - Critical (3 overpermissive flags): {cri.get('critical', 0):,} rules
    - High (2 overpermissive flags): {cri.get('high', 0):,} rules
    - Medium (1 overpermissive flag): {cri.get('medium', 0):,} rules
    """
        
            if dashboard_data.get('combinedRiskOutbound') and dashboard_data['combinedRiskOutbound'].get('found'):
                cro = dashboard_data['combinedRiskOutbound']
                data_summary += f"""
    ⚠️ COMBINED RISK - RISKY OUTBOUND + OVERPERMISSIVE:
    - Critical (3 overpermissive flags): {cro.get('critical', 0):,} rules
    - High (2 overpermissive flags): {cro.get('high', 0):,} rules
    - Medium (1 overpermissive flag): {cro.get('medium', 0):,} rules
    """
        
            # Top Risky Rules
            if analysis_data.get('riskyRules') and len(analysis_data['riskyRules']) > 0:
                data_summary += f"""
    🔴 TOP 10 RISKIEST RULES:
    """
                for i, rule in enumerate(analysis_data['riskyRules'][:10], 1):
                    data_summary += f"{i}. {rule.get('name', 'Unknown')} - Score: {rule.get('score', 0)} ({rule.get('riskLevel', 'Unknown')} Risk)\n"
        
            # Add detailed examples for high and critical risk rules
            if rule_examples.get('found'):
                if rule_examples.get('examples', {}).get('criticalRisk'):
                    data_summary += "\n📋 Example Critical Risk Rules (Score >= 175):\n"
                    for i, example in enumerate(rule_examples['examples']['criticalRisk'], 1):
                        data_summary += f"  {i}. {example}\n"
                if rule_examples.get('examples', {}).get('highRisk'):
                    data_summary += "\n📋 Example High Risk Rules (Score >= 100):\n"
                    for i, example in enumerate(rule_examples['examples']['highRisk'], 1):
                        data_summary += f"  {i}. {example}\n"
        
            # Internet Rules Sample
            if analysis_data.get('internetRules') and len(analysis_data['internetRules']) > 0:
                data_summary += f"""
    🌐 INTERNET RISK RULES (Sample of {min(5, len(analysis_data['internetRules']))}):
    """
                for i, rule in enumerate(analysis_data['internetRules'][:5], 1):
                    inbound = "Yes" if rule.get('riskyInbound') else "No"
                    outbound = "Yes" if rule.get('riskyOutbound') else "No"
                    data_summary += f"{i}. {rule.get('name', 'Unknown')} - Inbound: {inbound}, Outbound: {outbound}, Score: {rule.get('score', 0)}\n"
        
            # Advanced Policy Analysis (Policy Analyzer Results)
            if policy_analyzer_data.get('found'):
                data_summary += f"""
    🔍 ADVANCED POLICY ANALYSIS:
    """
                # Shadow Rules
                shadow_data = policy_analyzer_data.get('shadowRules', {})
                shadow_count = shadow_data.get('count', 0)
                data_summary += f"- Shadow Rules: {shadow_count:,} (fully covered by earlier rules with different actions)\n"
                if rule_examples.get('found') and rule_examples.get('examples', {}).get('shadowRules'):
                    data_summary += "  Example Shadow Rules (with full rule details):\n"
                    for i, example in enumerate(rule_examples['examples']['shadowRules'], 1):
                        data_summary += f"  {i}. {example}\n"
                elif shadow_count > 0 and shadow_data.get('sampleRules'):
                    data_summary += "  Sample Shadow Rules:\n"
                    for i, rule in enumerate(shadow_data['sampleRules'][:5], 1):
                        data_summary += f"  {i}. {rule.get('name', 'Unknown')} - {rule.get('reason', '')[:100]}\n"
            
                # Partial Shadow Rules
                partial_shadow_data = policy_analyzer_data.get('partialShadowRules', {})
                partial_shadow_count = partial_shadow_data.get('count', 0)
                data_summary += f"- Partial Shadow Rules: {partial_shadow_count:,} (partially covered by earlier rules)\n"
                if rule_examples.get('found') and rule_examples.get('examples', {}).get('partialShadowRules'):
                    data_summary += "  Example Partial Shadow Rules (with full rule details):\n"
                    for i, example in enumerate(rule_examples['examples']['partialShadowRules'], 1):
                        data_summary += f"  {i}. {example}\n"
                elif partial_shadow_count > 0 and partial_shadow_data.get('sampleRules'):
                    data_summary += "  Sample Partial Shadow Rules:\n"
                    for i, rule in enumerate(partial_shadow_data['sampleRules'][:5], 1):
                        data_summary += f"  {i}. {rule.get('name', 'Unknown')} - {rule.get('reason', '')[:100]}\n"
            
                # Redundant Rules
                redundant_data = policy_analyzer_data.get('redundantRules', {})
                redundant_count = redundant_data.get('count', 0)
                data_summary += f"- Redundant Rules: {redundant_count:,} (same action, fully covered by earlier rule)\n"
                if rule_examples.get('found') and rule_examples.get('examples', {}).get('redundantRules'):
                    data_summary += "  Example Redundant Rules (with full rule details):\n"
                    for i, example in enumerate(rule_examples['examples']['redundantRules'], 1):
                        data_summary += f"  {i}. {example}\n"
                elif redundant_count > 0 and redundant_data.get('sampleRules'):
                    data_summary += "  Sample Redundant Rules:\n"
                    for i, rule in enumerate(redundant_data['sampleRules'][:5], 1):
                        data_summary += f"  {i}. {rule.get('name', 'Unknown')} - {rule.get('reason', '')[:100]}\n"
            
                # Generalization Risks
                generalization_data = policy_analyzer_data.get('generalizationRisks', {})
                generalization_count = generalization_data.get('count', 0)
                data_summary += f"- Generalization Risks: {generalization_count:,} (broader rules generalizing specific ones)\n"
                if rule_examples.get('found') and rule_examples.get('examples', {}).get('generalizationRisks'):
                    data_summary += "  Example Generalization Risks (with full rule details):\n"
                    for i, example in enumerate(rule_examples['examples']['generalizationRisks'], 1):
                        data_summary += f"  {i}. {example}\n"
                elif generalization_count > 0 and generalization_data.get('sampleRules'):
                    data_summary += "  Sample Generalization Risks:\n"
                    for i, rule in enumerate(generalization_data['sampleRules'][:5], 1):
                        data_summary += f"  {i}. {rule.get('name', 'Unknown')} - {rule.get('reason', '')[:100]}\n"
            
                # Correlation Risks
                correlation_data = policy_analyzer_data.get('correlationRisks', {})
                correlation_count = correlation_data.get('count', 0)
                data_summary += f"- Correlation Risks: {correlation_count:,} (overlapping rules with different actions)\n"
                if rule_examples.get('found') and rule_examples.get('examples', {}).get('correlationRisks'):
                    data_summary += "  Example Correlation Risks (with full rule details):\n"
                    for i, example in enumerate(rule_examples['examples']['correlationRisks'], 1):
                        data_summary += f"  {i}. {example}\n"
                elif correlation_count > 0 and correlation_data.get('sampleRules'):
                    data_summary += "  Sample Correlation Risks:\n"
                    for i, rule in enumerate(correlation_data['sampleRules'][:5], 1):
                        data_summary += f"  {i}. {rule.get('name', 'Unknown')} - {rule.get('reason', '')[:100]}\n"
            
                # Consolidation Candidates
                consolidation_data = policy_analyzer_data.get('consolidationCandidates', {})
                consolidation_count = consolidation_data.get('count', 0)
                data_summary += f"- Consolidation Candidates: {consolidation_count:,} (rules that can be merged)\n"
                if rule_examples.get('found') and rule_examples.get('examples', {}).get('consolidationCandidates'):
                    data_summary += "  Example Consolidation Groups (with full rule details):\n"
                    for i, example in enumerate(rule_examples['examples']['consolidationCandidates'], 1):
                        data_summary += f"  {i}. {example}\n"
                elif consolidation_count > 0 and consolidation_data.get('sampleRules'):
                    data_summary += "  Sample Consolidation Groups:\n"
                    for i, group in enumerate(consolidation_data['sampleRules'][:5], 1):
                        rules_list = group.get('rules', [])
                        rules_str = ', '.join(rules_list[:3])
                        if len(rules_list) > 3:
                            rules_str += f" (+{len(rules_list) - 3} more)"
                        data_summary += f"  {i}. Rules: {rules_str} - {group.get('reason', '')[:100]}\n"
            else:
                data_summary += """
    🔍 ADVANCED POLICY ANALYSIS:
    - Advanced policy analysis has not been run on this file. Run the Advanced Analysis to get shadow rules, redundancies, generalizations, correlations, and consolidation candidates.
    """
        
            # Rule Reordering Suggestions
            if reordering_suggestions.get('found') and len(reordering_suggestions.get('suggestions', [])) > 0:
                data_summary += f"""
    🔀 RULE REORDERING SUGGESTIONS:
    - Total suggestions: {len(reordering_suggestions.get('suggestions', []))} rules should be reordered
    - Example reordering suggestions (with actual rule names):
    """
                for i, suggestion in enumerate(reordering_suggestions['suggestions'][:10], 1):
                    data_summary += f"  {i}. Rule '{suggestion['rule']}': Move from position {suggestion['from']} to position {suggestion['to']} | Reason: {suggestion['reason']}\n"
            else:
                data_summary += """
    🔀 RULE REORDERING SUGGESTIONS:
    - Reordering suggestions are not available. Run the Advanced Analysis and generate reordering suggestions to see recommended rule order changes.
    """
        
            # Firewall-Level Analysis Distribution (from AdvancedAnalysis)
            if firewall_analysis_distribution and len(firewall_analysis_distribution) > 0:
                data_summary += f"""
    📊 FIREWALL-LEVEL ANALYSIS DISTRIBUTION:
    - Total Firewalls Analyzed: {len(firewall_analysis_distribution)}
    - Firewalls sorted by total advanced analysis issues (Shadow + Redundant + Generalization + Correlation + Consolidation)

    Top Firewalls by Total Issues:
    """
                # Show top 10 firewalls by total issues
                sorted_firewalls = sorted(firewall_analysis_distribution, key=lambda x: x.get('total', 0), reverse=True)[:10]
                for i, fw in enumerate(sorted_firewalls, 1):
                    firewall_name = fw.get('firewall', 'Unknown').replace('.csv', '').replace('.xlsx', '')
                    data_summary += f"  {i}. {firewall_name}:\n"
                    data_summary += f"     - Shadow (Full + Partial): {fw.get('totalShadow', 0):,}\n"
                    data_summary += f"     - Redundant: {fw.get('redundant', 0):,}\n"
                    data_summary += f"     - Generalization: {fw.get('generalization', 0):,}\n"
                    data_summary += f"     - Correlation: {fw.get('correlation', 0):,}\n"
                    data_summary += f"     - Consolidation: {fw.get('consolidation', 0):,}\n"
                    data_summary += f"     - Total Issues: {fw.get('total', 0):,}\n"
            else:
                data_summary += """
    📊 FIREWALL-LEVEL ANALYSIS DISTRIBUTION:
    - Firewall-level distribution data not available. Run Advanced Analysis to see per-firewall breakdown of shadow, redundant, generalization, correlation, and consolidation issues.
    """
        
            # Security Control Assessment - now using ComplianceGrid mapping only
            # PDF extraction has been removed
            # Security Control Failures by Rule (from ComplianceGrid mapping)
            if rule_control_failures and len(rule_control_failures) > 0:
                data_summary += f"""
    🔒 SECURITY CONTROL FAILURES BY RULE (Based on ComplianceGrid Mapping):
    - Total Rules Analyzed: {len(rule_control_failures):,}
    - This mapping shows which security controls each rule fails based on its boolean flags
    - Controls are automatically determined from rule attributes (Overpermissive_Source, Risky_Inbound, etc.)

    Rules Failing Security Controls:
    """
                # Group rules by framework for better organization
                rules_with_cis_failures = []
                rules_with_pci_failures = []
                rules_with_nist_failures = []
                rules_with_all_frameworks = []
            
                for rule_name, failures in rule_control_failures.items():
                    has_cis = len(failures['CIS']) > 0
                    has_pci = len(failures['PCI']) > 0
                    has_nist = len(failures['NIST']) > 0
                
                    if has_cis and has_pci and has_nist:
                        rules_with_all_frameworks.append({
                            'name': rule_name,
                            'failures': failures
                        })
                    elif has_cis:
                        rules_with_cis_failures.append({
                            'name': rule_name,
                            'failures': failures
                        })
                    elif has_pci:
                        rules_with_pci_failures.append({
                            'name': rule_name,
                            'failures': failures
                        })
                    elif has_nist:
                        rules_with_nist_failures.append({
                            'name': rule_name,
                            'failures': failures
                        })
            
                # Show rules that fail all frameworks first
                if rules_with_all_frameworks:
                    data_summary += f"\nRules Failing ALL Frameworks (CIS, NIST, PCI): {len(rules_with_all_frameworks)}\n"
                    for rule_info in rules_with_all_frameworks[:20]:  # Limit to 20 for prompt size
                        rule_name = rule_info['name']
                        failures = rule_info['failures']
                        cis_list = ', '.join(failures['CIS'])
                        pci_list = ', '.join(failures['PCI'])
                        nist_list = ', '.join(failures['NIST'])
                        data_summary += f"  - {rule_name}:\n"
                        data_summary += f"    CIS: {cis_list}\n"
                        data_summary += f"    PCI: {pci_list}\n"
                        data_summary += f"    NIST: {nist_list}\n"
                    if len(rules_with_all_frameworks) > 20:
                        data_summary += f"  ... and {len(rules_with_all_frameworks) - 20} more rules failing all frameworks\n"
            
                # Show CIS failures
                if rules_with_cis_failures:
                    data_summary += f"\nRules Failing CIS Controls: {len(rules_with_cis_failures)}\n"
                    for rule_info in rules_with_cis_failures[:15]:
                        rule_name = rule_info['name']
                        cis_list = ', '.join(rule_info['failures']['CIS'])
                        data_summary += f"  - {rule_name}: CIS {cis_list}\n"
                    if len(rules_with_cis_failures) > 15:
                        data_summary += f"  ... and {len(rules_with_cis_failures) - 15} more rules\n"
            
                # Show PCI failures
                if rules_with_pci_failures:
                    data_summary += f"\nRules Failing PCI Controls: {len(rules_with_pci_failures)}\n"
                    for rule_info in rules_with_pci_failures[:15]:
                        rule_name = rule_info['name']
                        pci_list = ', '.join(rule_info['failures']['PCI'])
                        data_summary += f"  - {rule_name}: PCI {pci_list}\n"
                    if len(rules_with_pci_failures) > 15:
                        data_summary += f"  ... and {len(rules_with_pci_failures) - 15} more rules\n"
            
                # Show NIST failures
                if rules_with_nist_failures:
                    data_summary += f"\nRules Failing NIST Controls: {len(rules_with_nist_failures)}\n"
                    for rule_info in rules_with_nist_failures[:15]:
                        rule_name = rule_info['name']
                        nist_list = ', '.join(rule_info['failures']['NIST'])
                        data_summary += f"  - {rule_name}: NIST {nist_list}\n"
                    if len(rules_with_nist_failures) > 15:
                        data_summary += f"  ... and {len(rules_with_nist_failures) - 15} more rules\n"
            
                data_summary += """
    CRITICAL: When answering questions about security controls:
    - Use the exact control IDs listed above (e.g., "CIS 11.3", "NIST AC-4", "PCI 1.2.1")
    - For each rule mentioned, list ALL the controls it fails (CIS, NIST, PCI if applicable)
    - If a rule fails controls from all three frameworks, list all of them
    - Be specific: "Rule 'X' fails CIS 11.3, 12.3, 9.1; PCI 1.2.1-1.2.4; NIST AC-4, AC-6, SC-7"
    - This mapping is based on the ComplianceGrid mapping and is automatically determined from rule boolean flags
    """
            else:
                data_summary += """
    🔒 SECURITY CONTROL FAILURES BY RULE:
    - Security control failure analysis not available
    - Rules are analyzed based on boolean flags (Overpermissive_Source, Risky_Inbound, etc.)
    """
        
            # Detect if query is specific (not asking for general overview)
            is_specific_query = any(keyword in query.lower() for keyword in [
                'risky inbound', 'risky outbound', 'insecure port', 'top risky', 
                'show me', 'list', 'how many', 'what are', 'which rules', 'rules with',
                'shadow', 'redundant', 'generalization', 'correlation', 'consolidation',
                'shadowed', 'redundancy', 'generalize', 'overlap', 'merge',
                'reorder', 'reordering', 'rule order', 'suggested order', 'move rule', 'rule position', 'reorder rule'
            ])
        
            if is_specific_query:
                # For specific queries, add instruction to stay focused
                focus_instruction = """
    CRITICAL: The user asked a SPECIFIC question. Answer ONLY what they asked.
    - Do NOT provide overall statistics or other categories
    - Do NOT add "helpful" context about other metrics
    - Focus ONLY on the specific data requested
    - If they ask about "risky inbound internet", only discuss those specific rules and their direct attributes
    - Do NOT mention unrelated categories like "Source Any overall" or "Destination Any overall" unless the query explicitly asks for comparisons
    """
            else:
                focus_instruction = """
    Provide a comprehensive analysis based on the data above.
    """
        
            # Special handling for redundant/shadow/partial shadow rules position change questions
            position_change_instruction = ""
            query_lower = query.lower()
            if any(keyword in query_lower for keyword in ['redundant', 'shadow', 'partial shadow', 'partial']) and \
               any(keyword in query_lower for keyword in ['position', 'change', 'move', 'reorder', 'order']):
                position_change_instruction = """
    ⚠️ CRITICAL FOR REDUNDANT/SHADOW/PARTIAL SHADOW RULES POSITION QUESTIONS:
    - If the user asks about redundant rules, shadow rules, or partial shadow rules needing position changes, you MUST:
      1. List ALL redundant/shadow/partial shadow rules from the examples provided above (do NOT just say "reordering suggestions are not available")
      2. Explain that redundant rules are fully covered by earlier rules, so they should be:
         - REMOVED (recommended, since they're redundant and never evaluated), OR
         - MOVED DOWN in the rule order (if they must be kept for some reason)
      3. Explain that shadow rules are fully covered by earlier rules, so they should be:
         - REMOVED (recommended, since they're shadowed and never evaluated), OR
         - MOVED UP in the rule order (if they need to be evaluated before the covering rule)
      4. Explain that partial shadow rules are partially covered by earlier rules, so they should be:
         - REMOVED (recommended if fully redundant), OR
         - MOVED UP in the rule order (to be evaluated before the partially covering rule), OR
         - MODIFIED to remove the overlapping portion (if they need to be kept)
      5. For EACH redundant/shadow/partial shadow rule listed in the examples, specify what position change is needed (remove or move)
      6. Do NOT reference the "RULE REORDERING SUGGESTIONS" section being unavailable - use the redundant/shadow/partial shadow rule examples that ARE available
    - Redundant, shadow, and partial shadow rules inherently need position changes because they are covered (fully or partially) by earlier rules
    - The absence of explicit reordering suggestions does NOT prevent you from answering about redundant/shadow/partial shadow rules
    """
        
            prompt = f"""You are an expert firewall rule analyst with access to comprehensive firewall policy analysis data.

    🚨 CRITICAL DOMAIN RESTRICTION: 
    - You MUST ONLY answer questions about FIREWALL RULES and FIREWALL POLICY ANALYSIS
    - If the user asks about anything unrelated to firewall rules (e.g., countries, general knowledge, other topics), you MUST respond: **"I can only answer questions about firewall rules and firewall policy analysis. Please ask me about firewall rules, security policies, risk analysis, or rule configurations.
    Connect with PwC team to know the right query to ask !!!"**
    - Do NOT answer general knowledge questions or questions outside the firewall domain
    - Your expertise is LIMITED to firewall rule analysis based on the data provided below

    🚫 PDF/DOCUMENT QUESTIONS ARE NOT SUPPORTED:
    - If the user asks questions about PDF files, documents, benchmarks (CIS, NIST, ISO, PCI-DSS), or uploaded files, you MUST respond: **"You are asking out of scope question"**
    - PDF files are NOT used in this system - security controls are determined from rule boolean flags using the ComplianceGrid mapping
    - Do NOT answer questions like:
      * "What is in the PDF?"
      * "What does the CIS benchmark say?"
      * "Explain the NIST controls"
      * "Tell me about the uploaded document"
      * "What controls are in the PDF?"
    - These questions are OUT OF SCOPE - respond with: "You are asking out of scope question"

    IMPORTANT: The data summary below includes detailed rule examples with ACTUAL FIREWALL RULE NAMES. 
    You MUST use these actual rule names in your responses, not generic descriptions.
    When discussing rules, always cite the specific rule names from the examples (e.g., "Rule_Name_123", "FW_Rule_ABC").

    ⚠️ CRITICAL: If you do NOT see rule examples in the data below, your response will be incomplete. 
    Look for sections like "Example Source Any Rules:", "Example Shadow Rules:", etc. and USE THOSE ACTUAL RULE NAMES in your response.

    {data_summary}

    USER QUERY: {query}

    {focus_instruction}
    {position_change_instruction}

    INSTRUCTIONS:
    1. **FIRST CHECK: Is this question about PDF files, documents, benchmarks, or uploaded files? If YES, respond with: "You are asking out of scope question"**
       - **Note: Security controls are determined from rule boolean flags using ComplianceGrid mapping, NOT from PDF files**
    2. **SECOND CHECK: Is this question about firewall rules? If NO, respond with: "I can only answer questions about firewall rules and firewall policy analysis. Please ask me about firewall rules, security policies, risk analysis, or rule configurations. Connect with PwC team to know the right query to ask !!!"**
    3. Answer ONLY what the user specifically asked - do not provide additional context
    3. If the query asks about a specific category, focus ONLY on that category
    4. Do NOT mention overall statistics or unrelated metrics unless explicitly requested
    5. Use the EXACT numbers provided above, but only those directly relevant to the query
    6. **CRITICAL: ALWAYS use actual firewall rule names from the examples provided above when discussing specific rules**
    7. **When answering questions about rule categories, you MUST cite specific rule names from the examples (e.g., "Rule_Name_123", "FW_Rule_ABC", etc.)**
    8. **Do NOT provide generic answers - always reference actual rule names and their specific attributes (Source, Destination, Service, Action, Score) from the examples**
    9. **Example format: Instead of "Some rules have Source=Any", say "Rules like 'Rule_Name_123' and 'FW_Rule_ABC' have Source=Any"**
    10. When discussing specific rules, use the detailed rule examples provided above (they include Rule Name, Source, Destination, Service, Action, and Score attributes)
    11. Reference actual rule names and attributes from the examples when answering questions about specific rule categories
    12. Format your response using the following REQUIRED structure (MUST INCLUDE ALL SECTIONS):

    RESPONSE STRUCTURE (MUST INCLUDE ALL SECTIONS):

    **1. Summary**
    - What was found (exact counts and numbers from the data)
    - Key findings from the analysis
    - Brief overview of the issue
    - **MUST include specific rule names from the examples above (e.g., "Rule_Name_123", "FW_Rule_ABC")**

    **2. Root Cause Analysis - "Why do these rules have these values?"**
    - Technical explanation of why these rules are flagged
    - Configuration issues that caused the problem
    - Design decisions or patterns that led to this situation
    - Specific attributes that contribute to the risk
    - Explain the underlying reasons, not just what is wrong
    - **MUST reference specific rule names from the examples (e.g., "Rule 'FW_Rule_XYZ' has Source=Any because...")**

    **3. Immediate Actions Required - "What action must the user take?"**
    - List of immediate actions (numbered, prioritized)
    - What needs to be done first (critical items)
    - What needs to be done next (high priority)
    - Specific tasks with clear ownership
    - Focus on urgent actions that must be taken immediately
    - **MUST specify which rules need action (e.g., "Review and fix rule 'Rule_Name_123' which has Source=Any")**

    **4. Remediation Plan - "What is the remediation plan?"**
    - Step-by-step plan to fix the issues:
      a. Assessment phase (review and document current state)
      b. Planning phase (design new configurations)
      c. Implementation phase (make changes)
      d. Validation phase (verify fixes)
    - Specific configuration changes for each rule type
    - Testing and validation steps
    - Rollback procedures if needed
    - Timeline or priority order for remediation
    - **MUST include specific rule names that need remediation (e.g., "Rule 'FW_Rule_ABC' should be changed from Source=Any to Source=10.0.0.0/8")**

    **5. Recommendations - "What are the recommendations?"**
    - Best practices to implement
    - Policy recommendations
    - Process improvements
    - Security hardening measures
    - Monitoring and alerting suggestions
    - Long-term strategic recommendations
    - How to prevent similar issues in the future
    - **MUST reference specific rule examples when making recommendations (e.g., "Rules like 'Rule_Name_123' should follow this pattern...")**

    **6. Security Controls - "Which security controls are the rules failing?"**
    - **CRITICAL: This section MUST ALWAYS be included in your response**
    - **Use the "🔒 SECURITY CONTROL FAILURES BY RULE" section from the data above to identify which security controls are being failed**
    - **This mapping is automatically determined from rule boolean flags (Overpermissive_Source, Risky_Inbound, etc.) using the ComplianceGrid mapping**
    - **Provide a CONCISE summary of which security controls are failing:**
      - **List the security control IDs that are being failed** (e.g., "CIS 11.3, 12.3, 9.1", "PCI 1.2.1-1.2.4", "NIST AC-4, AC-6, SC-7")
      - **Include brief information about what each control category represents** (e.g., "CIS 11.3 relates to firewall rule restrictions", "NIST AC-4 relates to information flow enforcement")
      - **Group by framework** (CIS, NIST, PCI) and list the control IDs that are failing
    - **DO NOT list individual rules in this section** - this section should only state which controls are failing, not which specific rules fail them
    - **DO NOT provide detailed information about each rule** - keep it concise and focused on the security controls themselves
    - **Format example: "The rules are failing the following security controls:**
      - **CIS:** 11.3 (Firewall rule restrictions), 12.3 (Network access controls), 9.1 (Port restrictions)
      - **PCI:** 1.2.1-1.2.4 (Firewall configuration requirements)
      - **NIST:** AC-4 (Information flow enforcement), AC-6 (Least privilege), SC-7 (Boundary protection)"
    - **If the "🔒 SECURITY CONTROL FAILURES BY RULE" section is not available:**
      - State: "Security control failure analysis not available"
      - Explain that rules are analyzed based on boolean flags to determine control failures

    **7. Suggestions - "List ALL specific rules with their names and reasons"**
    - **CRITICAL: This section MUST list EVERY SINGLE rule from the examples provided above**
    - **MANDATORY REQUIREMENT: You MUST list ALL rules shown in the "Example [Category] Rules:" sections above - do NOT skip any, do NOT summarize, do NOT say "and X more rules"**
    - **For EACH and EVERY rule mentioned in the examples above, you MUST provide:**
      - **Rule Name:** [The actual rule name from the examples, e.g., 'FW_Rule_ABC', 'Rule_Name_123']
      - **Why/Reason:** [The specific reason why this rule is flagged, using the reason from the data provided]
      - **Details:** [Additional context like current position, suggested position, or specific attributes]
    - **If the data shows "Example [Category] Rules:" with a numbered list (1, 2, 3, etc.), you MUST list ALL of them in your Suggestions section**
    - **Format for different question types (list ALL rules individually - NO EXCEPTIONS):**
      - **For Overpermissive Rules:**
        - **Source Any Rules:** List each rule with why Source=Any is risky (e.g., "Rule 'FW_Rule_ABC' has Source=Any which allows traffic from any source, increasing attack surface...")
        - **Destination Any Rules:** List each rule with why Destination=Any is risky (e.g., "Rule 'Rule_Name_XYZ' has Destination=Any which allows traffic to any destination...")
        - **Service Broad Rules:** List each rule with why the service is too broad (e.g., "Rule 'FW_Rule_123' uses broad service range which exposes multiple ports...")
        - **Insecure Ports:** List each rule with which insecure port it uses and why it's risky (e.g., "Rule 'Rule_Name_ABC' uses insecure port HTTP (80) which transmits data unencrypted...")
      - **For Contextual Risk Rules:**
        - **Risky Inbound Internet Rules:** List each rule with why it's risky for inbound internet traffic (e.g., "Rule 'FW_Rule_XYZ' is risky inbound because it allows internet traffic to internal resources...")
        - **Risky Outbound Internet Rules:** List each rule with why it's risky for outbound internet traffic (e.g., "Rule 'Rule_Name_123' is risky outbound because it allows internal systems to reach any internet destination...")
      - **For Hygiene Rules:**
        - **Missing Security Profile Rules:** List each rule with why missing security profile is a problem (e.g., "Rule 'FW_Rule_ABC' is missing security profile which means no threat protection is applied...")
        - **Missing Log Forwarding Rules:** List each rule with why missing logging is a problem (e.g., "Rule 'Rule_Name_XYZ' is missing log forwarding which prevents security monitoring...")
      - **For Policy Analyzer Rules:**
        - **Shadow Rules:** List each shadow rule with why it's shadowed (e.g., "Rule 'FW_Rule_123' is shadowed because it's fully covered by earlier rule 'FW_Rule_ABC'...")
        - **Partial Shadow Rules:** List each partial shadow rule with why it's partially shadowed
        - **Redundant Rules:** List each redundant rule with why it's redundant (e.g., "Rule 'Rule_Name_XYZ' is redundant because it has the same action and is fully covered by earlier rule...")
        - **Generalization Risks:** List each rule with why it generalizes other rules
        - **Correlation Risks:** List each rule with why it correlates with other rules
        - **Consolidation Candidates:** List each consolidation group with which rules can be merged and why
      - **For Reordering Suggestions:** List each rule with current position, suggested position, direction (ABOVE/BELOW), and why (e.g., "Rule 'Rule_Name_XYZ': Move from position 45 to 12 (ABOVE) because it's shadowed and should be evaluated earlier...")
      - **For High/Critical Risk Rules:** List each rule with its score and why it's high risk (e.g., "Rule 'FW_Rule_ABC' has score 185 (Critical Risk) because it combines Source=Any, Destination=Internet, and insecure ports...")
      - **For any other category:** Follow the same pattern - list each rule with its name and specific reason
    - **ABSOLUTELY CRITICAL: Do NOT summarize - list ALL rules individually from the examples provided**
    - **If you see "Example [Category] Rules:" with 30 rules listed, you MUST list all 30 rules in your Suggestions section**
    - **If you see "Example [Category] Rules:" with 5 rules listed, you MUST list all 5 rules in your Suggestions section**
    - **Do NOT say "here are some examples" or "here are a few rules" - say "here are ALL the rules from the examples provided"**
    - **Use the exact rule names from the examples provided in the data above**
    - **Include the full reason/explanation for each rule from the data**
    - **If the question asks about a specific category, list ALL rules in that category with their names and reasons**
    - **Count the number of rules in the "Example [Category] Rules:" section and ensure you list that EXACT number of rules**

    CRITICAL REQUIREMENT: Throughout your entire response, you MUST cite actual rule names from the examples provided above. Do NOT use generic phrases like "some rules" or "these rules" without naming specific rules. Always say "Rule 'FW_Rule_ABC'" or "Rules like 'Rule_Name_123' and 'Rule_Name_456'".

    7. Each section should be specific and actionable
    8. Use exact numbers from the data provided
    9. Avoid generic advice - be specific to the firewall rules context
    10. Make recommendations practical and implementable

    ANSWER:"""
        
            # CRITICAL: Check prompt size and truncate if needed to fit in context window
            from app.llm_config import LLM_MAX_PROMPT_CHARS
        
            if len(prompt) > LLM_MAX_PROMPT_CHARS:
                logging.warning(f"⚠️ Prompt too large ({len(prompt):,} chars), truncating to fit in {LLM_MAX_PROMPT_CHARS:,} char limit")
            
                # Calculate how much we need to truncate data_summary
                prompt_overhead = len(prompt) - len(data_summary)  # Everything except data_summary
                max_data_summary_size = LLM_MAX_PROMPT_CHARS - prompt_overhead - 1000  # Leave 1000 chars buffer
            
                if max_data_summary_size > 0 and len(data_summary) > max_data_summary_size:
                    # Truncate data_summary
                    truncated_data_summary = data_summary[:max_data_summary_size] + "\n\n[... Data truncated to fit context window. Showing summary statistics and limited examples ...]"
                
                    # Rebuild prompt with truncated data
                    prompt = f"""You are an expert firewall rule analyst with access to comprehensive firewall policy analysis data.

    🚨 CRITICAL DOMAIN RESTRICTION: 
    - You MUST ONLY answer questions about FIREWALL RULES and FIREWALL POLICY ANALYSIS
    - If the user asks about anything unrelated to firewall rules (e.g., countries, general knowledge, other topics), you MUST respond: **"I can only answer questions about firewall rules and firewall policy analysis. Please ask me about firewall rules, security policies, risk analysis, or rule configurations.
    Connect with PwC team to know the right query to ask !!!"**
    - Do NOT answer general knowledge questions or questions outside the firewall domain
    - Your expertise is LIMITED to firewall rule analysis based on the data provided below

    🚫 PDF/DOCUMENT QUESTIONS ARE NOT SUPPORTED:
    - If the user asks questions about PDF files, documents, benchmarks (CIS, NIST, ISO, PCI-DSS), or uploaded files, you MUST respond: **"You are asking out of scope question"**
    - PDF files are NOT used in this system - security controls are determined from rule boolean flags using the ComplianceGrid mapping
    - Do NOT answer questions like:
      * "What is in the PDF?"
      * "What does the CIS benchmark say?"
      * "Explain the NIST controls"
      * "Tell me about the uploaded document"
      * "What controls are in the PDF?"
    - These questions are OUT OF SCOPE - respond with: "You are asking out of scope question"

    IMPORTANT: The data summary below includes detailed rule examples with ACTUAL FIREWALL RULE NAMES. 
    You MUST use these actual rule names in your responses, not generic descriptions.
    When discussing rules, always cite the specific rule names from the examples (e.g., "Rule_Name_123", "FW_Rule_ABC").

    ⚠️ CRITICAL: If you do NOT see rule examples in the data below, your response will be incomplete. 
    Look for sections like "Example Source Any Rules:", "Example Shadow Rules:", etc. and USE THOSE ACTUAL RULE NAMES in your response.

    {truncated_data_summary}

    USER QUERY: {query}

    {focus_instruction}
    {position_change_instruction}

    INSTRUCTIONS:
    1. **FIRST CHECK: Is this question about PDF files, documents, benchmarks, or uploaded files? If YES, respond with: "You are asking out of scope question"**
       - **Note: Security controls are determined from rule boolean flags using ComplianceGrid mapping, NOT from PDF files**
    2. **SECOND CHECK: Is this question about firewall rules? If NO, respond with: "I can only answer questions about firewall rules and firewall policy analysis. Please ask me about firewall rules, security policies, risk analysis, or rule configurations. Connect with PwC team to know the right query to ask !!!"**
    3. Answer ONLY what the user specifically asked - do not provide additional context
    3. If the query asks about a specific category, focus ONLY on that category
    4. Do NOT mention overall statistics or unrelated metrics unless explicitly requested
    5. Use the EXACT numbers provided above, but only those directly relevant to the query
    6. **CRITICAL: ALWAYS use actual firewall rule names from the examples provided above when discussing specific rules**
    7. **When answering questions about rule categories, you MUST cite specific rule names from the examples (e.g., "Rule_Name_123", "FW_Rule_ABC", etc.)**
    8. **Do NOT provide generic answers - always reference actual rule names and their specific attributes (Source, Destination, Service, Action, Score) from the examples**
    9. **Example format: Instead of "Some rules have Source=Any", say "Rules like 'Rule_Name_123' and 'FW_Rule_ABC' have Source=Any"**
    10. When discussing specific rules, use the detailed rule examples provided above (they include Rule Name, Source, Destination, Service, Action, and Score attributes)
    11. Reference actual rule names and attributes from the examples when answering questions about specific rule categories
    12. Format your response concisely due to limited context window

    ANSWER:"""
                    logging.info(f"✅ Prompt truncated from {len(prompt):,} to {len(prompt):,} chars")
        
        # Call LLM API
        logging.info(f"🤖 Query Intent: {query_intent.value.upper()} | Sending to LLM ({llm_config['model']})")
        logging.info(f"📏 Prompt size: {len(prompt):,} chars | Query: {len(query)} chars")
        
        # Warn if prompt is very large
        if len(prompt) > 100000:
            logging.warning(f"⚠️ Very large prompt ({len(prompt):,} chars), this may take longer to process")
        
        # Create timeout object with explicit connect and read timeouts
        timeout_config = httpx.Timeout(
            connect=10.0,  # 10 seconds to establish connection
            read=llm_config['timeout'],  # Configurable read timeout
            write=10.0,  # 10 seconds to write
            pool=5.0  # 5 seconds to get connection from pool
        )
        
        try:
            # Use Google Gemini API or fallback to Ollama/OpenAI
            async with httpx.AsyncClient(timeout=timeout_config) as client:
                # Check if using Gemini API (base_url contains generativelanguage.googleapis.com)
                if "generativelanguage.googleapis.com" in llm_config['base_url']:
                    # Google Gemini API format
                    model_name = llm_config['model']
                    gemini_url = f"{llm_config['base_url']}/models/{model_name}:generateContent?key={llm_config['api_key']}"
                    
                    gemini_payload = {
                        "contents": [{
                            "parts": [{
                                "text": prompt
                            }]
                        }],
                        "generationConfig": {
                            "temperature": llm_config['temperature'],
                            "maxOutputTokens": llm_config['max_tokens']
                        }
                    }
                    
                    headers = {
                        "Content-Type": "application/json"
                    }
                    
                    logging.info(f"🤖 Calling Google Gemini API: {model_name}")
                    logging.info(f"🔗 Gemini URL: {gemini_url.split('?')[0]}...")  # Don't log full URL with key
                    
                    response = await client.post(gemini_url, json=gemini_payload, headers=headers)
                    
                    if response.status_code == 200:
                        result = response.json()
                        logging.info(f"📥 Raw Gemini response structure: {list(result.keys())}")
                        
                        # Extract response from Gemini format
                        candidates = result.get("candidates", [])
                        if candidates and len(candidates) > 0:
                            candidate = candidates[0]
                            
                            # Check for finish reason (safety filters, etc.)
                            finish_reason = candidate.get("finishReason", "")
                            
                            # Handle MAX_TOKENS specifically - response was cut off due to token limit
                            if finish_reason == "MAX_TOKENS":
                                content = candidate.get("content", {})
                                parts = content.get("parts", [])
                                if parts and len(parts) > 0:
                                    llm_response = parts[0].get("text", "")
                                    if llm_response and llm_response.strip() != "":
                                        # Return partial response with a warning
                                        logging.warning(f"⚠️ Gemini response hit MAX_TOKENS limit. Returning partial response ({len(llm_response)} chars). Consider increasing max_tokens.")
                                        
                                        # Build context with policy analyzer data
                                        context = {
                                            "totalRules": analysis_data.get('totalRules', 0),
                                            "averageScore": analysis_data.get('averageScore', 0),
                                            "riskLevels": {
                                                "critical": dashboard_data.get('riskDashboard', {}).get('critical', 0) if dashboard_data.get('riskDashboard', {}).get('found') else 0,
                                                "high": dashboard_data.get('riskDashboard', {}).get('high', 0) if dashboard_data.get('riskDashboard', {}).get('found') else 0,
                                                "medium": dashboard_data.get('riskDashboard', {}).get('medium', 0) if dashboard_data.get('riskDashboard', {}).get('found') else 0,
                                                "low": dashboard_data.get('riskDashboard', {}).get('low', 0) if dashboard_data.get('riskDashboard', {}).get('found') else 0,
                                                "none": dashboard_data.get('riskDashboard', {}).get('none', 0) if dashboard_data.get('riskDashboard', {}).get('found') else 0
                                            },
                                            "topRiskyRules": analysis_data.get('riskyRules', [])[:5] if analysis_data.get('riskyRules') else []
                                        }
                                        
                                        # Add policy analyzer data to context if available
                                        if policy_analyzer_data.get('found'):
                                            context["policyAnalysis"] = {
                                                "found": True,
                                                "shadowRules": {
                                                    "count": policy_analyzer_data.get('shadowRules', {}).get('count', 0),
                                                    "sampleRules": policy_analyzer_data.get('shadowRules', {}).get('sampleRules', [])[:5]
                                                },
                                                "partialShadowRules": {
                                                    "count": policy_analyzer_data.get('partialShadowRules', {}).get('count', 0),
                                                    "sampleRules": policy_analyzer_data.get('partialShadowRules', {}).get('sampleRules', [])[:5]
                                                },
                                                "redundantRules": {
                                                    "count": policy_analyzer_data.get('redundantRules', {}).get('count', 0),
                                                    "sampleRules": policy_analyzer_data.get('redundantRules', {}).get('sampleRules', [])[:5]
                                                },
                                                "generalizationRisks": {
                                                    "count": policy_analyzer_data.get('generalizationRisks', {}).get('count', 0),
                                                    "sampleRules": policy_analyzer_data.get('generalizationRisks', {}).get('sampleRules', [])[:5]
                                                },
                                                "correlationRisks": {
                                                    "count": policy_analyzer_data.get('correlationRisks', {}).get('count', 0),
                                                    "sampleRules": policy_analyzer_data.get('correlationRisks', {}).get('sampleRules', [])[:5]
                                                },
                                                "consolidationCandidates": {
                                                    "count": policy_analyzer_data.get('consolidationCandidates', {}).get('count', 0),
                                                    "sampleRules": policy_analyzer_data.get('consolidationCandidates', {}).get('sampleRules', [])[:5]
                                                }
                                            }
                                        else:
                                            context["policyAnalysis"] = {
                                                "found": False
                                            }
                                        
                                        return JSONResponse({
                                            "success": True,
                                            "response": str(llm_response) + "\n\n[Note: Response was truncated due to token limit. Consider increasing max_tokens for complete responses.]",
                                            "model": llm_config['model'],
                                            "context": context
                                        })
                                else:
                                    # No content available even though MAX_TOKENS was hit
                                    usage_metadata = result.get("usageMetadata", {})
                                    total_tokens = usage_metadata.get("totalTokenCount", 0)
                                    max_tokens = llm_config.get('max_tokens', 'unknown')
                                    error_msg = f"Gemini API hit MAX_TOKENS limit ({max_tokens} tokens) before generating any content. Total tokens used: {total_tokens}. Please increase max_tokens in your LLM configuration."
                                    logging.error(f"❌ {error_msg}")
                                    logging.error(f"📥 Usage metadata: {usage_metadata}")
                                    raise HTTPException(
                                        status_code=500,
                                        detail=error_msg
                                    )
                            
                            # Handle other finish reasons (safety filters, etc.)
                            if finish_reason and finish_reason != "STOP":
                                # Response was blocked or stopped for a reason
                                safety_ratings = candidate.get("safetyRatings", [])
                                safety_details = ", ".join([f"{r.get('category', 'Unknown')}: {r.get('probability', 'Unknown')}" for r in safety_ratings])
                                error_msg = f"Gemini API blocked response. Finish reason: {finish_reason}"
                                if safety_details:
                                    error_msg += f". Safety ratings: {safety_details}"
                                logging.error(f"❌ {error_msg}")
                                logging.error(f"📥 Full candidate structure: {candidate}")
                                raise HTTPException(
                                    status_code=500,
                                    detail=error_msg
                                )
                            
                            content = candidate.get("content", {})
                            parts = content.get("parts", [])
                            if parts and len(parts) > 0:
                                llm_response = parts[0].get("text", "")
                                
                                # Additional check: if text is None or empty string
                                if not llm_response or llm_response.strip() == "":
                                    logging.error(f"❌ Empty text in Gemini response. Candidate: {candidate}")
                                    logging.error(f"📥 Full response structure: {result}")
                                    raise HTTPException(
                                        status_code=500,
                                        detail="Gemini API returned empty text content. The response may have been blocked by safety filters."
                                    )
                            else:
                                logging.error(f"❌ No parts in Gemini response. Candidate: {candidate}")
                                logging.error(f"📥 Full response structure: {result}")
                                llm_response = ""
                        else:
                            # Check for error in response
                            error_msg = result.get("error", {})
                            if error_msg:
                                error_message = error_msg.get("message", "Unknown error")
                                logging.error(f"❌ Gemini API error: {error_message}")
                                raise HTTPException(
                                    status_code=500,
                                    detail=f"Gemini API error: {error_message}"
                                )
                            
                            # Log the full response for debugging
                            logging.error(f"❌ No candidates in Gemini response. Full response: {result}")
                            llm_response = ""
                        
                        if llm_response and llm_response.strip() != "":
                            logging.info(f"✅ Gemini response received ({len(str(llm_response))} chars)")
                            
                            # Build context with policy analyzer data
                            context = {
                                    "totalRules": analysis_data.get('totalRules', 0),
                                    "averageScore": analysis_data.get('averageScore', 0),
                                    "riskLevels": {
                                        "critical": dashboard_data.get('riskDashboard', {}).get('critical', 0) if dashboard_data.get('riskDashboard', {}).get('found') else 0,
                                        "high": dashboard_data.get('riskDashboard', {}).get('high', 0) if dashboard_data.get('riskDashboard', {}).get('found') else 0,
                                        "medium": dashboard_data.get('riskDashboard', {}).get('medium', 0) if dashboard_data.get('riskDashboard', {}).get('found') else 0,
                                        "low": dashboard_data.get('riskDashboard', {}).get('low', 0) if dashboard_data.get('riskDashboard', {}).get('found') else 0,
                                        "none": dashboard_data.get('riskDashboard', {}).get('none', 0) if dashboard_data.get('riskDashboard', {}).get('found') else 0
                                    },
                                    "topRiskyRules": analysis_data.get('riskyRules', [])[:5] if analysis_data.get('riskyRules') else []
                                }
                            
                            # Add policy analyzer data to context if available
                            if policy_analyzer_data.get('found'):
                                context["policyAnalysis"] = {
                                    "found": True,
                                    "shadowRules": {
                                        "count": policy_analyzer_data.get('shadowRules', {}).get('count', 0),
                                        "sampleRules": policy_analyzer_data.get('shadowRules', {}).get('sampleRules', [])[:5]
                                    },
                                    "partialShadowRules": {
                                        "count": policy_analyzer_data.get('partialShadowRules', {}).get('count', 0),
                                        "sampleRules": policy_analyzer_data.get('partialShadowRules', {}).get('sampleRules', [])[:5]
                                    },
                                    "redundantRules": {
                                        "count": policy_analyzer_data.get('redundantRules', {}).get('count', 0),
                                        "sampleRules": policy_analyzer_data.get('redundantRules', {}).get('sampleRules', [])[:5]
                                    },
                                    "generalizationRisks": {
                                        "count": policy_analyzer_data.get('generalizationRisks', {}).get('count', 0),
                                        "sampleRules": policy_analyzer_data.get('generalizationRisks', {}).get('sampleRules', [])[:5]
                                    },
                                    "correlationRisks": {
                                        "count": policy_analyzer_data.get('correlationRisks', {}).get('count', 0),
                                        "sampleRules": policy_analyzer_data.get('correlationRisks', {}).get('sampleRules', [])[:5]
                                    },
                                    "consolidationCandidates": {
                                        "count": policy_analyzer_data.get('consolidationCandidates', {}).get('count', 0),
                                        "sampleRules": policy_analyzer_data.get('consolidationCandidates', {}).get('sampleRules', [])[:5]
                                    }
                                }
                            else:
                                context["policyAnalysis"] = {
                                    "found": False
                                }
                            
                            return JSONResponse({
                                "success": True,
                                "response": str(llm_response),
                                "model": llm_config['model'],
                                "context": context
                            })
                        else:
                            logging.error(f"❌ Empty response from Gemini API. Full response: {result}")
                            raise HTTPException(
                                status_code=500,
                                detail="Gemini API returned empty response. Check server logs for details."
                            )
                    else:
                        error_text = response.text[:500] if hasattr(response, 'text') else str(response.status_code)
                        try:
                            error_json = response.json()
                            error_detail = error_json.get("error", {}).get("message", error_text)
                        except:
                            error_detail = error_text
                        logging.error(f"❌ Gemini API error: {response.status_code} - {error_detail}")
                        raise HTTPException(
                            status_code=500,
                            detail=f"Gemini API error: {response.status_code}. {error_detail}"
                        )
                
                # Try OpenAI-compatible format first (LM Studio, LocalAI, etc.)
                # Since base_url now includes /v1, we can directly use /chat/completions
                headers = {
                    "Content-Type": "application/json"
                }
                
                if llm_config['api_key']:
                    headers["Authorization"] = f"Bearer {llm_config['api_key']}"
                
                openai_url = f"{llm_config['base_url']}/chat/completions"
                openai_payload = {
                    "model": llm_config['model'],
                    "messages": [
                        {"role": "system", "content": "You are an expert firewall rule analyst."},
                        {"role": "user", "content": prompt}
                    ],
                    "temperature": llm_config['temperature'],
                    "max_tokens": llm_config['max_tokens']
                }
                
                logging.info(f"🤖 Trying OpenAI-compatible format at {openai_url}...")
                response = await client.post(openai_url, json=openai_payload, headers=headers)
                
                if response.status_code == 200:
                    result = response.json()
                    logging.info(f"📥 Raw OpenAI-format response structure: {list(result.keys())}")
                    
                    # Try to extract response from OpenAI format
                    choices = result.get("choices", [])
                    if choices and len(choices) > 0:
                        message = choices[0].get("message", {})
                        llm_response = message.get("content") or message.get("text") or ""
                    else:
                        # Try alternative structures
                        llm_response = (
                            result.get("text") or 
                            result.get("content") or
                            result.get("response") or
                            ""
                        )
                    
                    if llm_response and llm_response.strip() != "":
                        logging.info(f"✅ LLM response received ({len(str(llm_response))} chars)")
                        
                        # Build context with policy analyzer data
                        context = {
                            "totalRules": analysis_data.get('totalRules', 0),
                            "averageScore": analysis_data.get('averageScore', 0),
                            "riskLevels": {
                                "critical": dashboard_data.get('riskDashboard', {}).get('critical', 0) if dashboard_data.get('riskDashboard', {}).get('found') else 0,
                                "high": dashboard_data.get('riskDashboard', {}).get('high', 0) if dashboard_data.get('riskDashboard', {}).get('found') else 0,
                                "medium": dashboard_data.get('riskDashboard', {}).get('medium', 0) if dashboard_data.get('riskDashboard', {}).get('found') else 0,
                                "low": dashboard_data.get('riskDashboard', {}).get('low', 0) if dashboard_data.get('riskDashboard', {}).get('found') else 0,
                                "none": dashboard_data.get('riskDashboard', {}).get('none', 0) if dashboard_data.get('riskDashboard', {}).get('found') else 0
                            },
                            "topRiskyRules": analysis_data.get('riskyRules', [])[:5] if analysis_data.get('riskyRules') else []
                        }
                        
                        # Add policy analyzer data to context if available
                        if policy_analyzer_data.get('found'):
                            context["policyAnalysis"] = {
                                "found": True,
                                "shadowRules": {
                                    "count": policy_analyzer_data.get('shadowRules', {}).get('count', 0),
                                    "sampleRules": policy_analyzer_data.get('shadowRules', {}).get('sampleRules', [])[:5]
                                },
                                "partialShadowRules": {
                                    "count": policy_analyzer_data.get('partialShadowRules', {}).get('count', 0),
                                    "sampleRules": policy_analyzer_data.get('partialShadowRules', {}).get('sampleRules', [])[:5]
                                },
                                "redundantRules": {
                                    "count": policy_analyzer_data.get('redundantRules', {}).get('count', 0),
                                    "sampleRules": policy_analyzer_data.get('redundantRules', {}).get('sampleRules', [])[:5]
                                },
                                "generalizationRisks": {
                                    "count": policy_analyzer_data.get('generalizationRisks', {}).get('count', 0),
                                    "sampleRules": policy_analyzer_data.get('generalizationRisks', {}).get('sampleRules', [])[:5]
                                },
                                "correlationRisks": {
                                    "count": policy_analyzer_data.get('correlationRisks', {}).get('count', 0),
                                    "sampleRules": policy_analyzer_data.get('correlationRisks', {}).get('sampleRules', [])[:5]
                                },
                                "consolidationCandidates": {
                                    "count": policy_analyzer_data.get('consolidationCandidates', {}).get('count', 0),
                                    "sampleRules": policy_analyzer_data.get('consolidationCandidates', {}).get('sampleRules', [])[:5]
                                }
                            }
                        else:
                            context["policyAnalysis"] = {
                                "found": False
                            }
                        
                        return JSONResponse({
                            "success": True,
                            "response": str(llm_response),
                            "model": llm_config['model'],
                            "context": context
                        })
                
                # If OpenAI format didn't work, try Ollama API format as fallback
                error_text = response.text[:500] if hasattr(response, 'text') else str(response.status_code)
                logging.warning(f"⚠️ OpenAI format failed (status {response.status_code}): {error_text}")
                logging.info(f"⚠️ Trying Ollama format as fallback...")
                
                ollama_url = f"{llm_config['base_url'].replace('/v1', '')}/api/generate"
                payload = {
                    "model": llm_config['model'],
                    "prompt": prompt,
                    "stream": False,
                    "options": {
                        "temperature": llm_config['temperature'],
                        "num_predict": llm_config['max_tokens']
                    }
                }
                
                response = await client.post(ollama_url, json=payload, headers=headers)
                
                if response.status_code == 200:
                    result = response.json()
                    logging.info(f"📥 Raw OpenAI-format response structure: {list(result.keys())}")
                    logging.info(f"📥 Raw OpenAI-format response (first 500 chars): {str(result)[:500]}")
                    
                    # Try to extract response from OpenAI format
                    choices = result.get("choices", [])
                    if choices and len(choices) > 0:
                        message = choices[0].get("message", {})
                        llm_response = message.get("content") or message.get("text") or ""
                    else:
                        # Try alternative structures
                        llm_response = (
                            result.get("text") or 
                            result.get("content") or
                            result.get("response") or
                            ""
                        )
                    
                    if not llm_response or llm_response == "":
                        logging.error(f"❌ No response found in OpenAI-format result. Full response: {result}")
                        raise HTTPException(
                            status_code=500,
                            detail=f"LLM returned empty response in OpenAI format. Check server logs for details. Response keys: {list(result.keys())}"
                        )
                    
                    logging.info(f"✅ LLM response received via OpenAI format ({len(str(llm_response))} chars)")
                    return JSONResponse({
                        "success": True,
                        "response": str(llm_response),
                        "model": llm_config['model'],
                        "context": {
                            "totalRules": analysis_data.get('totalRules', 0),
                            "averageScore": analysis_data.get('averageScore', 0),
                            "riskLevels": {
                                "critical": dashboard_data.get('riskDashboard', {}).get('critical', 0) if dashboard_data.get('riskDashboard', {}).get('found') else 0,
                                "high": dashboard_data.get('riskDashboard', {}).get('high', 0) if dashboard_data.get('riskDashboard', {}).get('found') else 0,
                                "medium": dashboard_data.get('riskDashboard', {}).get('medium', 0) if dashboard_data.get('riskDashboard', {}).get('found') else 0,
                                "low": dashboard_data.get('riskDashboard', {}).get('low', 0) if dashboard_data.get('riskDashboard', {}).get('found') else 0,
                                "none": dashboard_data.get('riskDashboard', {}).get('none', 0) if dashboard_data.get('riskDashboard', {}).get('found') else 0
                            },
                            "topRiskyRules": analysis_data.get('riskyRules', [])[:5] if analysis_data.get('riskyRules') else []
                        }
                    })
                else:
                    error_text = response.text[:500] if hasattr(response, 'text') else str(response.status_code)
                    logging.error(f"❌ LLM API error: {response.status_code} - {error_text}")
                    # Try /v1/completions as last resort (LM Studio and some other servers)
                    logging.info(f"⚠️ Trying /v1/completions endpoint...")
                    completions_url = f"{llm_config['base_url']}/v1/completions"
                    completions_payload = {
                        "model": llm_config['model'],
                        "prompt": prompt,
                        "temperature": llm_config['temperature'],
                        "max_tokens": llm_config['max_tokens']
                    }
                    
                    try:
                        response = await client.post(completions_url, json=completions_payload, headers=headers)
                        if response.status_code == 200:
                            result = response.json()
                            logging.info(f"📥 Raw completions response structure: {list(result.keys())}")
                            
                            # Try to extract from completions format
                            choices = result.get("choices", [])
                            if choices and len(choices) > 0:
                                llm_response = choices[0].get("text") or choices[0].get("content") or ""
                            else:
                                llm_response = result.get("text") or result.get("content") or result.get("response") or ""
                            
                            if llm_response and llm_response != "":
                                logging.info(f"✅ LLM response received via /v1/completions format ({len(str(llm_response))} chars)")
                                return JSONResponse({
                                    "success": True,
                                    "response": str(llm_response),
                                    "model": llm_config['model'],
                                    "context": {
                                        "totalRules": analysis_data.get('totalRules', 0),
                                        "averageScore": analysis_data.get('averageScore', 0),
                                        "riskLevels": {
                                            "critical": dashboard_data.get('riskDashboard', {}).get('critical', 0) if dashboard_data.get('riskDashboard', {}).get('found') else 0,
                                            "high": dashboard_data.get('riskDashboard', {}).get('high', 0) if dashboard_data.get('riskDashboard', {}).get('found') else 0,
                                            "medium": dashboard_data.get('riskDashboard', {}).get('medium', 0) if dashboard_data.get('riskDashboard', {}).get('found') else 0,
                                            "low": dashboard_data.get('riskDashboard', {}).get('low', 0) if dashboard_data.get('riskDashboard', {}).get('found') else 0,
                                            "none": dashboard_data.get('riskDashboard', {}).get('none', 0) if dashboard_data.get('riskDashboard', {}).get('found') else 0
                                        },
                                        "topRiskyRules": analysis_data.get('riskyRules', [])[:5] if analysis_data.get('riskyRules') else []
                                    }
                                })
                    except Exception as e:
                        logging.warning(f"⚠️ /v1/completions also failed: {e}")
                    
                    raise HTTPException(
                        status_code=500,
                        detail=f"LLM API error: {response.status_code}. Response: {error_text}. Tried Ollama, OpenAI, and Completions formats. Please check LLM server configuration and ensure it's running at {llm_config['base_url']}. Check backend logs for response structure."
                    )
        
        except httpx.TimeoutException:
            logging.error(f"❌ LLM request timeout after {llm_config['timeout']}s")
            raise HTTPException(status_code=504, detail=f"LLM request timed out. Please try again.")
        except httpx.RequestError as e:
            logging.error(f"❌ LLM request error: {e}")
            raise HTTPException(status_code=503, detail=f"Failed to connect to LLM server. Please check if LLM server is running at {llm_config['base_url']}")
        except Exception as e:
            logging.error(f"❌ Error calling LLM: {e}")
            import traceback
            logging.error(traceback.format_exc())
            raise HTTPException(status_code=500, detail=f"Error processing LLM query: {str(e)}")
    
    except pd.errors.EmptyDataError:
        raise HTTPException(status_code=400, detail="Excel file is empty")
    except Exception as e:
        logging.error(f"❌ Error processing NLM query: {e}")
        import traceback
        logging.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Failed to process query: {str(e)}")


@app.post("/api/explain-cell")
async def explain_cell_endpoint(request: ExplainCellRequest):
    """
    Endpoint to explain why a specific cell in ExcelGrid is marked as True.
    """
    return await explain_cell_value(request)

# ============================================================================
# Advanced Analysis Endpoints
# ============================================================================

@app.options("/api/analyze/stream/{session_id}")
async def analysis_stream_options(session_id: str):
    """Handle OPTIONS preflight for SSE endpoint."""
    from fastapi.responses import Response
    logging.info(f"🌐 OPTIONS preflight request for SSE stream {session_id[:8]}...")
    return Response(
        content="",
        status_code=200,
        headers={
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET, OPTIONS",
            "Access-Control-Allow-Headers": "*",
            "Access-Control-Allow-Credentials": "true",
            "Access-Control-Max-Age": "86400"
        }
    )

@app.post("/api/analyze")
async def analyze(
    file: UploadFile = File(None),
    x_session_id: str = Header(None, alias="X-Session-Id")
):
    """Run advanced policy analysis and return counts plus a download id. 
    If no file is provided, automatically uses Final_output_scored_static.xlsx from data/expansions.
    Overwrites the uploaded Excel with analysis columns."""
    logging.info("=" * 80)
    logging.info(f"🚀 Starting Advanced Policy Analysis")
    
    # Determine file source: uploaded file or static file
    use_static_file = file is None or file.filename is None
    tmp_path = None
    
    if use_static_file:
        logging.info("📁 No file provided, attempting to use static file...")
        static_file_path = _find_static_file()
        if not static_file_path:
            raise HTTPException(
                status_code=404, 
                detail="No file provided and Final_output_scored_static.xlsx not found. Please run scoring first or upload a file."
            )
        tmp_path = static_file_path
        logging.info(f"✅ Using static file: {tmp_path}")
    else:
        logging.info(f"📁 Received file: {file.filename} (size: {file.size} bytes)")
        if not file.filename.lower().endswith((".xlsx", ".xls")):
            logging.error(f"❌ Invalid file type: {file.filename}")
            raise HTTPException(status_code=400, detail="Only Excel files (.xlsx or .xls) are supported")
        
        try:
            logging.info("💾 Saving uploaded file to temporary location...")
            import tempfile
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                content = await file.read()
                tmp_file.write(content)
                tmp_path = pathlib.Path(tmp_file.name)
            logging.info(f"✓ File saved to: {tmp_path}")
        except Exception as e:
            logging.error(f"❌ Error saving uploaded file: {e}")
            raise HTTPException(status_code=500, detail=f"Failed to save uploaded file: {str(e)}")
    
    # Use client-supplied session ID or generate one
    session_id = x_session_id if x_session_id else str(uuid.uuid4())
    if x_session_id:
        logging.info(f"📋 Using client-supplied session ID: {session_id[:8]}...")
    else:
        logging.info(f"🆔 Generated new session ID: {session_id[:8]}...")
    
    # Initialize session immediately (before file processing) so SSE can connect early
    if session_id not in ANALYSIS_PROGRESS:
        activity_log_list = []
        ANALYSIS_PROGRESS[session_id] = {
            "overall_percent": 0,
            "current_stage": "uploading" if not use_static_file else "loading_static",
            "current_firewall": None,
            "firewall_progress": {},
            "firewall_details": {},
            "activity_log": activity_log_list,
            "total_firewalls": 0,
            "current_firewall_index": 0,
            "firewall_rule_counts": {},
            "completed": False,
            "downloadId": None
        }
        SSE_QUEUES[session_id] = Queue()
        logging.info(f"✅ Pre-initialized ANALYSIS_PROGRESS for session {session_id}")
    
    try:
        # Load Excel file
        logging.info("📖 Loading Excel file...")
        try:
            df = pd.read_excel(tmp_path, sheet_name='Raw Data')
            logging.info("✓ Loaded sheet 'Raw Data'")
        except Exception:
            try:
                df = pd.read_excel(tmp_path, sheet_name='Raw_Data')
                logging.info("✓ Loaded sheet 'Raw_Data'")
            except Exception:
                df = pd.read_excel(tmp_path, sheet_name=0)
                logging.info("✓ Loaded first available sheet")
        
        logging.info(f"📊 Loaded DataFrame: {len(df)} rows, {len(df.columns)} columns")
        
        # Detect firewalls
        source_file_col = None
        for col_name in ['Source_File', 'Source File', 'source_file', 'Firewall', 'Device']:
            if col_name in df.columns:
                source_file_col = col_name
                break
        
        unique_firewalls = []
        firewall_rule_counts = {}
        if source_file_col:
            unique_firewalls = df[source_file_col].unique().tolist()
            unique_firewalls = [str(fw) for fw in unique_firewalls if pd.notna(fw) and str(fw).strip()]
            for fw in unique_firewalls:
                firewall_rule_counts[fw] = int((df[source_file_col] == fw).sum())
        else:
            unique_firewalls = ["All Rules"]
            firewall_rule_counts["All Rules"] = len(df)
        
        # Update progress store
        existing_activity_log = ANALYSIS_PROGRESS[session_id].get("activity_log", [])
        if not isinstance(existing_activity_log, list):
            existing_activity_log = []
        
        ANALYSIS_PROGRESS[session_id].update({
            "overall_percent": 0,
            "current_stage": "initializing",
            "current_firewall": None,
            "firewall_progress": {fw: 0 for fw in unique_firewalls},
            "firewall_details": {fw: {"percent": 0, "rules_processed": 0, "total_rules": firewall_rule_counts.get(fw, 0), "pairs_processed": 0, "total_pairs": 0} for fw in unique_firewalls},
            "activity_log": existing_activity_log,
            "total_firewalls": len(unique_firewalls),
            "current_firewall_index": 0,
            "firewall_rule_counts": firewall_rule_counts
        })
        
        if session_id not in SSE_QUEUES:
            SSE_QUEUES[session_id] = Queue()
        
        # Use SessionContext to enable automatic logging capture
        with SessionContext(session_id):
            def add_activity_log(message: str, log_type: str = "info"):
                if session_id not in ANALYSIS_PROGRESS:
                    return
                if "activity_log" not in ANALYSIS_PROGRESS[session_id]:
                    ANALYSIS_PROGRESS[session_id]["activity_log"] = []
                log_entry = {
                    "timestamp": datetime.now().strftime("%H:%M:%S"),
                    "message": message,
                    "type": log_type
                }
                log_list = ANALYSIS_PROGRESS[session_id]["activity_log"]
                log_list.append(log_entry)
                if len(log_list) > 200:
                    log_list.pop(0)
                if session_id in SSE_QUEUES:
                    try:
                        SSE_QUEUES[session_id].put({"type": "log", "data": log_entry})
                    except Exception:
                        pass
            
            def progress_callback(stage: str, message: str, overall_percent: int, 
                            firewall_name: str = None, firewall_percent: int = None,
                            firewall_index: int = None, total_firewalls: int = None,
                            rules_processed: int = None, total_rules: int = None,
                            pairs_processed: int = None, total_pairs: int = None):
                if session_id not in ANALYSIS_PROGRESS:
                    return
                ANALYSIS_PROGRESS[session_id]["current_stage"] = stage
                ANALYSIS_PROGRESS[session_id]["overall_percent"] = overall_percent
                
                msg_lower = message.lower()
                log_type = "info"
                if "✓" in message or "complete" in msg_lower or "success" in msg_lower:
                    log_type = "success"
                elif "⚠" in message or "warning" in msg_lower:
                    log_type = "warning"
                elif "❌" in message or "error" in msg_lower or "failed" in msg_lower:
                    log_type = "error"
                
                log_entry = {
                    "timestamp": datetime.now().strftime("%H:%M:%S"),
                    "message": message,
                    "type": log_type
                }
                
                if "activity_log" not in ANALYSIS_PROGRESS[session_id]:
                    ANALYSIS_PROGRESS[session_id]["activity_log"] = []
                log_list_ref = ANALYSIS_PROGRESS[session_id]["activity_log"]
                log_list_ref.append(log_entry)
                if len(log_list_ref) > 200:
                    log_list_ref.pop(0)
                
                if firewall_name:
                    ANALYSIS_PROGRESS[session_id]["current_firewall"] = firewall_name
                    if firewall_percent is not None:
                        if firewall_name in ANALYSIS_PROGRESS[session_id]["firewall_progress"]:
                            ANALYSIS_PROGRESS[session_id]["firewall_progress"][firewall_name] = firewall_percent
                    if firewall_name in ANALYSIS_PROGRESS[session_id]["firewall_details"]:
                        fw_detail = ANALYSIS_PROGRESS[session_id]["firewall_details"][firewall_name]
                        if rules_processed is not None:
                            fw_detail["rules_processed"] = rules_processed
                        if total_rules is not None:
                            fw_detail["total_rules"] = total_rules
                        if pairs_processed is not None:
                            fw_detail["pairs_processed"] = pairs_processed
                        if total_pairs is not None:
                            fw_detail["total_pairs"] = total_pairs
                        if fw_detail["total_rules"] > 0:
                            rules_progress = (fw_detail["rules_processed"] / fw_detail["total_rules"]) * 60
                            pairs_progress = 0
                            if fw_detail["total_pairs"] > 0:
                                pairs_progress = (fw_detail["pairs_processed"] / fw_detail["total_pairs"]) * 40
                            fw_detail["percent"] = min(100, int(rules_progress + pairs_progress))
                            ANALYSIS_PROGRESS[session_id]["firewall_progress"][firewall_name] = fw_detail["percent"]
                
                if firewall_index is not None:
                    ANALYSIS_PROGRESS[session_id]["current_firewall_index"] = firewall_index
                if total_firewalls is not None:
                    ANALYSIS_PROGRESS[session_id]["total_firewalls"] = total_firewalls
                
                if session_id in SSE_QUEUES:
                    try:
                        SSE_QUEUES[session_id].put({"type": "progress_update", "data": ANALYSIS_PROGRESS[session_id].copy()})
                    except Exception:
                        pass
            
            add_activity_log(f"📊 Loaded DataFrame: {len(df)} rows, {len(df.columns)} columns", "success")
            if source_file_col:
                add_activity_log(f"🔍 Detected {len(unique_firewalls)} unique firewall(s)", "info")
            
            progress_callback("file_loaded", f"📁 Loaded {len(df)} rows from Excel. Detected {len(unique_firewalls)} firewall(s).", 5, 
                            total_firewalls=len(unique_firewalls))
            
            add_activity_log("🔍 Starting policy analysis...", "info")
            progress_callback("analysis_start", "🚀 Starting advanced policy analysis...", 8,
                            total_firewalls=len(unique_firewalls))
            
            analyzed = detect_policy(df, progress_callback=progress_callback)
            add_activity_log("✓ Policy analysis completed", "success")
            
            add_activity_log("💾 Saving analyzed results back to Excel...", "info")
            progress_callback("saving_excel", "💾 Saving analyzed results to Excel file...", 95)
            with pd.ExcelWriter(tmp_path, engine='openpyxl', mode='w') as writer:
                analyzed.to_excel(writer, index=False, sheet_name='Raw Data')
            
            # Apply formatting to the Excel file
            try:
                from openpyxl import load_workbook
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                
                wb = load_workbook(tmp_path)
                ws = wb['Raw Data']
                
                # Define styles
                header_fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')  # Medium grey
                header_font = Font(bold=True, color='FFFFFF', size=11)  # White text
                header_alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
                
                cell_alignment = Alignment(vertical='middle', horizontal='left', wrap_text=True)
                
                true_fill = PatternFill(start_color='4A90E2', end_color='4A90E2', fill_type='solid')  # Medium blue
                true_font = Font(bold=True, color='FFFFFF')  # White text
                true_alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
                
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Format header row (row 1)
                for col_num, header in enumerate(analyzed.columns, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = thin_border
                
                # Format data rows
                for row_num in range(2, len(analyzed) + 2):
                    for col_num, header in enumerate(analyzed.columns, 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.alignment = cell_alignment
                        cell.border = thin_border
                        
                        # Get the value from the dataframe
                        df_value = analyzed.iloc[row_num - 2, col_num - 1]
                        
                        # Highlight True values with medium blue
                        if df_value is True or (isinstance(df_value, str) and df_value.lower() == 'true'):
                            cell.fill = true_fill
                            cell.font = true_font
                            cell.alignment = true_alignment
                            cell.value = 'True'
                        elif df_value is False or (isinstance(df_value, str) and df_value.lower() == 'false'):
                            cell.value = 'False'
                
                # Freeze first row and first column
                ws.freeze_panes = 'B2'
                
                # Auto-adjust column widths
                for col_num, header in enumerate(analyzed.columns, 1):
                    column_letter = get_column_letter(col_num)
                    max_length = len(str(header))
                    for row_num in range(2, len(analyzed) + 2):
                        cell_value = ws.cell(row=row_num, column=col_num).value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)))
                    ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 50)
                
                wb.save(tmp_path)
                logging.info("✓ Excel file formatted with styles")
            except Exception as e:
                logging.warning(f"⚠️ Could not format Excel file: {e}")
                # Continue even if formatting fails
            
            add_activity_log("✓ Excel file updated with analysis results", "success")
            progress_callback("excel_saved", "✓ Excel file updated with analysis results", 98)
            
            # Counts
            shadow_count = int((analyzed.get('Shadow_Rule') == True).sum()) if 'Shadow_Rule' in analyzed.columns else 0
            pshadow_count = int((analyzed.get('Partial_Shadow_Rule') == True).sum()) if 'Partial_Shadow_Rule' in analyzed.columns else 0
            redundant_count = int((analyzed.get('Redundant_Rule') == True).sum()) if 'Redundant_Rule' in analyzed.columns else 0
            general_count = int((analyzed.get('Generalization_Risk') == True).sum()) if 'Generalization_Risk' in analyzed.columns else 0
            correl_count = int((analyzed.get('Correlation_Risk') == True).sum()) if 'Correlation_Risk' in analyzed.columns else 0
            consolid_count = int((analyzed.get('Consolidation_Candidate') == True).sum()) if 'Consolidation_Candidate' in analyzed.columns else 0
            
            logging.info(f"📊 ANALYSIS RESULTS: Shadow={shadow_count}, Partial Shadow={pshadow_count}, Redundant={redundant_count}, Generalization={general_count}, Correlation={correl_count}, Consolidation={consolid_count}")
            
            # Get firewall-level distribution for all metrics
            firewall_analysis_distribution = []
            if source_file_col:
                for firewall_name, group_df in analyzed.groupby(source_file_col):
                    if pd.isna(firewall_name) or str(firewall_name).strip() == '':
                        continue
                    
                    # Calculate counts for each metric
                    shadow = int((group_df.get('Shadow_Rule') == True).sum()) if 'Shadow_Rule' in group_df.columns else 0
                    partial_shadow = int((group_df.get('Partial_Shadow_Rule') == True).sum()) if 'Partial_Shadow_Rule' in group_df.columns else 0
                    total_shadow = shadow + partial_shadow
                    redundant = int((group_df.get('Redundant_Rule') == True).sum()) if 'Redundant_Rule' in group_df.columns else 0
                    generalization = int((group_df.get('Generalization_Risk') == True).sum()) if 'Generalization_Risk' in group_df.columns else 0
                    correlation = int((group_df.get('Correlation_Risk') == True).sum()) if 'Correlation_Risk' in group_df.columns else 0
                    consolidation = int((group_df.get('Consolidation_Candidate') == True).sum()) if 'Consolidation_Candidate' in group_df.columns else 0
                    total = total_shadow + redundant + generalization + correlation + consolidation
                    
                    firewall_display = str(firewall_name).strip()
                    if firewall_display.lower().endswith('.csv') or firewall_display.lower().endswith('.xlsx'):
                        firewall_display = firewall_display.rsplit('.', 1)[0]
                    
                    firewall_analysis_distribution.append({
                        "firewall": firewall_display,
                        "shadow": shadow,
                        "partialShadow": partial_shadow,
                        "totalShadow": total_shadow,
                        "redundant": redundant,
                        "generalization": generalization,
                        "correlation": correlation,
                        "consolidation": consolidation,
                        "total": total
                    })
                
                # Sort by total (all metrics combined) descending
                firewall_analysis_distribution.sort(key=lambda x: x['total'], reverse=True)
            
            # Register for download
            download_id = str(uuid.uuid4())
            ANALYSIS_FILES[download_id] = str(tmp_path)
            add_activity_log("✅ Analysis completed successfully!", "success")
            progress_callback("complete", "✅ Analysis completed successfully! All results saved to Excel.", 100)
            
            # Mark analysis as completed in progress tracker
            ANALYSIS_PROGRESS[session_id]["completed"] = True
            ANALYSIS_PROGRESS[session_id]["downloadId"] = download_id
            
            # Send final progress update to SSE queue if it exists
            if session_id in SSE_QUEUES:
                try:
                    SSE_QUEUES[session_id].put({
                        "type": "progress_update", 
                        "data": ANALYSIS_PROGRESS[session_id].copy()
                    })
                except Exception:
                    pass
            
            return JSONResponse({
                "success": True,
                "downloadId": download_id,
                "sessionId": session_id,
                "shadowRules": {"found": True, "shadowCount": shadow_count, "partialShadowCount": pshadow_count},
                "redundancy": {"found": True, "redundantCount": redundant_count},
                "generalization": {"found": True, "generalizationCount": general_count},
                "correlation": {"found": True, "correlationCount": correl_count},
                "consolidation": {"found": True, "consolidationCount": consolid_count},
                "firewallAnalysisDistribution": firewall_analysis_distribution,
            })
    except Exception as e:
        logging.error(f"❌ Error during analysis: {str(e)}")
        logging.exception("Full traceback:")
        if 'session_id' in locals() and session_id in ANALYSIS_PROGRESS:
            ANALYSIS_PROGRESS[session_id]["activity_log"].append({
                "timestamp": datetime.now().strftime("%H:%M:%S"),
                "message": f"Error: {str(e)}",
                "type": "error"
            })
        if tmp_path and tmp_path.exists() and not use_static_file:
            tmp_path.unlink(missing_ok=True)
        raise HTTPException(status_code=500, detail=f"Analyze failed: {e}")

@app.get("/api/analyze/progress/{session_id}")
def get_analysis_progress(session_id: str):
    """Get progress information for an ongoing or completed analysis."""
    if session_id not in ANALYSIS_PROGRESS:
        logging.warning(f"Session {session_id} not found in ANALYSIS_PROGRESS")
        raise HTTPException(status_code=404, detail="Session not found")
    
    original_data = ANALYSIS_PROGRESS[session_id]
    progress_data = {
        "overall_percent": original_data.get("overall_percent", 0),
        "current_stage": original_data.get("current_stage", "unknown"),
        "current_firewall": original_data.get("current_firewall"),
        "firewall_progress": original_data.get("firewall_progress", {}),
        "total_firewalls": original_data.get("total_firewalls", 0),
        "current_firewall_index": original_data.get("current_firewall_index", 0),
        "completed": original_data.get("completed", False),
        "downloadId": original_data.get("downloadId"),
    }
    
    if "activity_log" in original_data:
        activity_log_raw = original_data["activity_log"]
        if isinstance(activity_log_raw, list):
            copied_logs = []
            for entry in activity_log_raw:
                if isinstance(entry, dict):
                    copied_logs.append({
                        "timestamp": str(entry.get("timestamp", "")),
                        "message": str(entry.get("message", "")),
                        "type": str(entry.get("type", "info"))
                    })
            progress_data["activity_log"] = copied_logs
        else:
            progress_data["activity_log"] = []
    else:
        progress_data["activity_log"] = []
    
    return progress_data

@app.get("/api/analyze/stream/{session_id}")
async def analysis_stream(session_id: str):
    """Stream real-time analysis progress and activity logs using Server-Sent Events."""
    from queue import Empty
    
    logging.info(f"🌐 SSE endpoint hit for session {session_id[:8]}...")
    
    if session_id not in SSE_QUEUES:
        SSE_QUEUES[session_id] = Queue()

    async def event_generator():
        try:
            logging.info(f"📡 Starting event generator for session {session_id[:8]}...")
            
            initial_data = f"data: {json.dumps({'status': 'connecting', 'session': session_id[:8]})}\n\n"
            yield initial_data
            
            connected_event = f"event: connected\ndata: {json.dumps({'message': 'SSE connection established'})}\n\n"
            yield connected_event
            
            max_wait = 10.0
            wait_interval = 0.1
            waited = 0.0
            
            if session_id not in ANALYSIS_PROGRESS:
                yield f"event: waiting\ndata: {json.dumps({'message': 'Waiting for session to be initialized...'})}\n\n"
            
            while session_id not in ANALYSIS_PROGRESS and waited < max_wait:
                await asyncio.sleep(wait_interval)
                waited += wait_interval
                if session_id in ANALYSIS_PROGRESS:
                    break
            
            if session_id not in ANALYSIS_PROGRESS:
                yield f"event: error\ndata: {json.dumps({'message': 'Session not found'})}\n\n"
                return
            
            initial_progress = ANALYSIS_PROGRESS[session_id].copy()
            if "activity_log" in initial_progress and isinstance(initial_progress["activity_log"], list):
                activity_log_copy = []
                for entry in initial_progress["activity_log"]:
                    if isinstance(entry, dict):
                        activity_log_copy.append({
                            "timestamp": str(entry.get("timestamp", "")),
                            "message": str(entry.get("message", "")),
                            "type": str(entry.get("type", "info"))
                        })
                initial_progress["activity_log"] = activity_log_copy
            else:
                initial_progress["activity_log"] = []
            
            initial_event = f"event: initial_state\ndata: {json.dumps(initial_progress)}\n\n"
            yield initial_event

            queue = SSE_QUEUES[session_id]
            last_log_count = len(initial_progress.get("activity_log", []))
            
            while True:
                try:
                    if session_id in ANALYSIS_PROGRESS:
                        progress_data = ANALYSIS_PROGRESS[session_id].copy()
                        if "activity_log" in progress_data and isinstance(progress_data["activity_log"], list):
                            activity_log_copy = []
                            for entry in progress_data["activity_log"]:
                                if isinstance(entry, dict):
                                    activity_log_copy.append({
                                        "timestamp": str(entry.get("timestamp", "")),
                                        "message": str(entry.get("message", "")),
                                        "type": str(entry.get("type", "info"))
                                    })
                            progress_data["activity_log"] = activity_log_copy
                            current_log_count = len(activity_log_copy)
                            
                            if current_log_count != last_log_count:
                                progress_event = f"event: progress_update\ndata: {json.dumps(progress_data)}\n\n"
                                yield progress_event
                                last_log_count = current_log_count
                        else:
                            progress_data["activity_log"] = []
                        
                        if progress_data.get("overall_percent", 0) >= 100:
                            yield f"event: complete\ndata: {json.dumps({'message': 'Analysis complete'})}\n\n"
                            break
                    else:
                        yield f"event: error\ndata: {json.dumps({'message': 'Session not found'})}\n\n"
                        break

                    queue_empty = False
                    while not queue_empty:
                        try:
                            event = queue.get_nowait()
                            log_event = f"event: {event['type']}\ndata: {json.dumps(event['data'])}\n\n"
                            yield log_event
                        except Empty:
                            queue_empty = True

                    await asyncio.sleep(0.1)
                    
                except asyncio.CancelledError:
                    logging.info(f"SSE stream for session {session_id} cancelled by client")
                    break
                except Exception as e:
                    logging.error(f"Error in SSE event_generator for session {session_id}: {e}")
                    yield f"event: error\ndata: {json.dumps({'message': str(e)})}\n\n"
                    break
        finally:
            logging.info(f"SSE stream closed for session {session_id}")

    headers = {
        "Cache-Control": "no-cache, no-transform",
        "Connection": "keep-alive",
        "Content-Type": "text/event-stream; charset=utf-8",
        "X-Accel-Buffering": "no",
        "Access-Control-Allow-Origin": "*",
        "Access-Control-Allow-Credentials": "true",
        "Access-Control-Allow-Methods": "GET, OPTIONS",
        "Access-Control-Allow-Headers": "*",
    }
    
    response = StreamingResponse(
        event_generator(), 
        media_type="text/event-stream",
        headers=headers
    )
    
    return response

@app.get("/api/download-analysis")
def download_analysis(id: str = Query(...)):
    """Download the analyzed Excel file by download ID."""
    logging.info(f"📥 Download request for analysis ID: {id}")
    path = ANALYSIS_FILES.get(id)
    if not path or not os.path.exists(path):
        logging.warning(f"❌ Analysis file not found or expired for ID: {id}")
        raise HTTPException(status_code=404, detail="Analysis file not found or expired")
    filename = f"analysis-{id}.xlsx"
    logging.info(f"✓ Sending file: {filename} ({os.path.getsize(path)} bytes)")
    return FileResponse(path, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=filename)

@app.post("/api/analyze/reorder-suggestions")
async def get_reorder_suggestions(
    file: UploadFile = File(None),
    download_id: str = Query(None),
    save_to_excel: bool = Query(True)
):
    """Get rule reordering suggestions and optionally save Suggested_Order column to Excel."""
    logging.info("🔀 Generating rule reordering suggestions...")
    logging.info(f"   download_id: {download_id}, save_to_excel: {save_to_excel}")
    
    # Determine file source
    tmp_path = None
    is_temp_file = False
    use_static_file = file is None or file.filename is None
    
    if download_id and download_id in ANALYSIS_FILES:
        # Use analyzed file from download_id
        tmp_path = pathlib.Path(ANALYSIS_FILES[download_id])
        if not tmp_path.exists():
            logging.error(f"❌ Analysis file not found at: {tmp_path}")
            raise HTTPException(status_code=404, detail="Analysis file not found")
        logging.info(f"✅ Using analyzed file from download_id: {tmp_path}")
    elif use_static_file:
        # Use static file
        static_file_path = _find_static_file()
        if not static_file_path:
            raise HTTPException(
                status_code=404,
                detail="No file provided and Final_output_scored_static.xlsx not found."
            )
        tmp_path = static_file_path
        logging.info(f"✅ Using static file: {tmp_path}")
    else:
        # Use uploaded file
        if not file.filename.lower().endswith((".xlsx", ".xls")):
            raise HTTPException(status_code=400, detail="Only Excel files are supported")
        
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_path = pathlib.Path(tmp_file.name)
            is_temp_file = True
            logging.info(f"✅ Created temporary file: {tmp_path}")
    
    try:
        # Load Excel file
        logging.info(f"📖 Loading Excel file from: {tmp_path}")
        try:
            df = pd.read_excel(tmp_path, sheet_name='Raw Data')
            sheet_name = 'Raw Data'
            logging.info(f"✅ Loaded sheet 'Raw Data'")
        except Exception as e1:
            logging.warning(f"Could not load 'Raw Data' sheet: {e1}")
            try:
                df = pd.read_excel(tmp_path, sheet_name='Raw_Data')
                sheet_name = 'Raw_Data'
                logging.info(f"✅ Loaded sheet 'Raw_Data'")
            except Exception as e2:
                logging.warning(f"Could not load 'Raw_Data' sheet: {e2}")
                df = pd.read_excel(tmp_path, sheet_name=0)
                # Get actual sheet name
                with pd.ExcelFile(tmp_path) as xls:
                    sheet_name = xls.sheet_names[0] if xls.sheet_names else 'Sheet1'
                logging.info(f"✅ Loaded first sheet: {sheet_name}")
        
        logging.info(f"📊 Loaded DataFrame: {len(df)} rows, {len(df.columns)} columns")
        
        # Check if analysis columns exist
        if "Rule_Order" not in df.columns:
            logging.error(f"❌ Rule_Order column not found. Available columns: {list(df.columns)[:10]}...")
            raise HTTPException(
                status_code=400,
                detail="File must have Rule_Order column. Please run analysis first."
            )
        
        # Generate reordering suggestions (returns both suggestions and updated DataFrame)
        logging.info("🔍 Generating reordering suggestions...")
        from app.policy_analyzer import suggest_rule_reordering
        suggestions, df_updated = suggest_rule_reordering(df)
        logging.info(f"✅ Generated {len(suggestions)} suggestions")
        
        # Save updated DataFrame back to Excel if requested
        if save_to_excel:
            logging.info(f"💾 Saving Suggested_Order column to Excel...")
            try:
                # Try to preserve other sheets by reading the workbook first
                from openpyxl import load_workbook
                try:
                    wb = load_workbook(tmp_path)
                    sheet_names_to_preserve = [s for s in wb.sheetnames if s != sheet_name]
                    wb.close()
                    
                    # If there are other sheets, use openpyxl to preserve them
                    if sheet_names_to_preserve:
                        logging.info(f"   Preserving {len(sheet_names_to_preserve)} other sheet(s)")
                        # Read other sheets
                        other_sheets_data = {}
                        for sname in sheet_names_to_preserve:
                            try:
                                other_sheets_data[sname] = pd.read_excel(tmp_path, sheet_name=sname)
                            except Exception as e:
                                logging.warning(f"   Could not read sheet '{sname}': {e}")
                        
                        # Write all sheets including updated one
                        with pd.ExcelWriter(tmp_path, engine='openpyxl', mode='w') as writer:
                            # Write updated sheet
                            df_updated.to_excel(writer, index=False, sheet_name=sheet_name)
                            # Write preserved sheets
                            for sname, sdata in other_sheets_data.items():
                                sdata.to_excel(writer, index=False, sheet_name=sname)
                        logging.info(f"✅ Saved with preserved sheets")
                    else:
                        # No other sheets, just write the updated one
                        with pd.ExcelWriter(tmp_path, engine='openpyxl', mode='w') as writer:
                            df_updated.to_excel(writer, index=False, sheet_name=sheet_name)
                        logging.info(f"✅ Saved updated sheet")
                except Exception as e:
                    logging.warning(f"   Could not preserve other sheets: {e}, using simple write")
                    # Fallback: simple write (overwrites entire file)
                    with pd.ExcelWriter(tmp_path, engine='openpyxl', mode='w') as writer:
                        df_updated.to_excel(writer, index=False, sheet_name=sheet_name)
                    logging.info(f"✅ Saved using simple write")
                
                logging.info(f"✅ Saved Suggested_Order column to Excel at: {tmp_path}")
            except Exception as e:
                logging.error(f"❌ Error saving to Excel: {e}")
                logging.exception("Full traceback:")
                raise HTTPException(status_code=500, detail=f"Failed to save to Excel: {str(e)}")
        
        logging.info(f"✅ Generated {len(suggestions)} reordering suggestions")
        
        return JSONResponse({
            "success": True,
            "suggestions": suggestions,
            "total_suggestions": len(suggestions),
            "saved_to_excel": save_to_excel,
            "file_path": str(tmp_path) if save_to_excel else None
        })
    
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"❌ Error generating reordering suggestions: {e}")
        logging.exception("Full traceback:")
        raise HTTPException(status_code=500, detail=f"Failed to generate suggestions: {str(e)}")
    finally:
        # Only clean up if we created a temporary file (not from download_id or static file)
        if is_temp_file and tmp_path and tmp_path.exists():
            try:
                tmp_path.unlink()
                logging.info(f"🗑️ Cleaned up temporary file: {tmp_path}")
            except Exception as e:
                logging.warning(f"Could not delete temp file: {e}")
