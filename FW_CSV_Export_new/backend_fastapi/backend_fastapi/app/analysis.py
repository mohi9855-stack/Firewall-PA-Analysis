"""
Analysis module for firewall rule analysis dashboard.
Handles data processing and analysis calculations directly from Excel files.
"""

import pandas as pd
import pathlib
import logging
from typing import Dict, List, Any, Optional
from fastapi import HTTPException, UploadFile
import openpyxl
from copy import copy
import re
from datetime import datetime

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class AnalysisEngine:
    """Main analysis engine for firewall rule analysis."""
    
    def __init__(self, data_root: pathlib.Path):
        self.data_root = data_root
        self.expansions_dir = data_root / "expansions"
        
    def extract_value_from_formula(self, formula_text: str, col_name: str = "") -> str:
        """Extract actual value from Excel formula.
        Formulas are like: IF(FALSE,"True," & ...,"False,0")
        The last quoted string is typically the actual value when FALSE.
        """
        if not formula_text or not isinstance(formula_text, str):
            return ""
        
        # Extract all quoted strings from the formula
        quoted_strings = re.findall(r'"([^"]+)"', formula_text)
        
        if not quoted_strings:
            return ""
        
        # For IF(FALSE,...) formulas, the last quoted string is usually the actual value
        # For IF(TRUE,...) formulas, we need to parse differently
        formula_upper = formula_text.upper()
        
        if formula_upper.startswith("IF(TRUE"):
            # For TRUE case, look for values like "True,5" or "True,10"  
            for qs in quoted_strings:
                if qs.lower().startswith("true,"):
                    return qs
        elif formula_upper.startswith("IF(FALSE"):
            # For FALSE case, the last quoted string is the value
            return quoted_strings[-1]
        
        return ""
    
    def _get_risky_rules(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """Get top 10 risky rules."""
        risky_rules = []
        
        # Check for Score_Total column
        if 'Score_Total' not in df.columns:
            logging.warning("Score_Total column not found in DataFrame. Available columns: " + ", ".join(df.columns.tolist()[:10]))
            return risky_rules
        
        logging.debug("Found Score_Total column - proceeding to get risky rules")
        logging.debug(f"📋 All available columns in DataFrame ({len(df.columns)} total): {', '.join(df.columns.tolist())}")
        
        # Try to find the rule name column - check multiple possible names
        # Prioritize 'Rule_Name' since that's the actual column name in the Excel file
        rule_name_col = None
        possible_name_cols = ['Rule_Name', 'Name', 'Rule Name', 'RuleName', 'name', 'rule_name', 'Rule', 'rule']
        
        # Also check case-insensitive matching
        df_columns_lower = {col.lower(): col for col in df.columns}
        for col in possible_name_cols:
            if col in df.columns:
                rule_name_col = col
                logging.debug(f"✅ Found rule name column (exact match): '{col}'")
                break
            elif col.lower() in df_columns_lower:
                rule_name_col = df_columns_lower[col.lower()]
                logging.debug(f"✅ Found rule name column (case-insensitive match): '{col}' -> '{rule_name_col}'")
                break
        
        if not rule_name_col:
            logging.warning(f"Rule name column not found. Available columns: {', '.join(df.columns.tolist()[:10])}")
            # Use first column as fallback if no name column found
            if len(df.columns) > 0:
                rule_name_col = df.columns[0]
                logging.info(f"Using first column '{rule_name_col}' as rule name")
            else:
                return risky_rules
        
        try:
            # Convert Score_Total to numeric for sorting
            df_temp = df.copy()
            df_temp['Score_Total_numeric'] = pd.to_numeric(df_temp['Score_Total'], errors='coerce').fillna(0)
            
            logging.debug(f"Total rows in DataFrame: {len(df_temp)}")
            logging.debug(f"Score_Total column statistics: min={df_temp['Score_Total_numeric'].min()}, max={df_temp['Score_Total_numeric'].max()}, mean={df_temp['Score_Total_numeric'].mean()}")
            
            # First, try to get rules with Score_Total > 0
            df_with_scores = df_temp[df_temp['Score_Total_numeric'] > 0]
            
            if len(df_with_scores) > 0:
                # Sort by Score_Total in descending order (highest to lowest) and get top 10
                # Use sort_values to ensure all columns are preserved
                if rule_name_col in df_with_scores.columns:
                    top_risky = df_with_scores.sort_values(
                        by=['Score_Total_numeric', rule_name_col], 
                        ascending=[False, True]
                    ).head(10).copy()  # Use .copy() to ensure we have a proper DataFrame
                else:
                    top_risky = df_with_scores.sort_values(by='Score_Total_numeric', ascending=False).head(10).copy()
                logging.debug(f"Found {len(df_with_scores)} rules with Score_Total > 0, selecting top 10 (sorted descending by Score_Total)")
            else:
                # If no rules have Score_Total > 0, get top 10 rules regardless of score
                # This ensures we always show something, even if all scores are 0
                # Still sort by Score_Total (even if all are 0, this maintains consistent ordering)
                if rule_name_col in df_temp.columns:
                    top_risky = df_temp.sort_values(
                        by=['Score_Total_numeric', rule_name_col], 
                        ascending=[False, True]
                    ).head(10).copy()
                else:
                    top_risky = df_temp.sort_values(by='Score_Total_numeric', ascending=False).head(10).copy()
                logging.warning(f"No rules with Score_Total > 0 found. Showing top 10 rules regardless of score (all may be 0)")
                logging.info(f"Total rules in DataFrame: {len(df_temp)}")
            
            # Reset index to ensure clean iteration
            top_risky = top_risky.reset_index(drop=True)
            logging.debug(f"Top risky rules DataFrame shape: {top_risky.shape}, Columns: {', '.join(top_risky.columns.tolist())}")
            
            if len(top_risky) == 0:
                logging.warning("No rules found at all in DataFrame")
                return risky_rules
            
            # Verify sorting - log the scores in order and verify they're descending
            scores_list = top_risky['Score_Total_numeric'].tolist()
            is_descending = all(scores_list[i] >= scores_list[i+1] for i in range(len(scores_list)-1))
            logging.debug(f"Top 10 risky rules scores (in order): {scores_list}")
            logging.debug(f"Scores are in descending order: {is_descending}")
            if not is_descending:
                logging.error("⚠️ WARNING: Scores are NOT in descending order! Sorting may have failed.")
            logging.debug(f"Using column '{rule_name_col}' for rule names.")
            
            # Verify rule_name_col exists in top_risky
            if rule_name_col not in top_risky.columns:
                logging.error(f"⚠️ ERROR: Column '{rule_name_col}' not found in top_risky DataFrame!")
                logging.error(f"Available columns in top_risky: {', '.join(top_risky.columns.tolist())}")
                # Try to use the original column name
                if 'Rule_Name' in top_risky.columns:
                    rule_name_col = 'Rule_Name'
                    logging.info(f"Using 'Rule_Name' directly")
                else:
                    logging.error("Cannot find Rule_Name column. Using index as fallback.")
            
            # Log first few rule names and scores for verification using direct column access
            logging.debug(f"🔍 Verifying rule name extraction using column: '{rule_name_col}'")
            for i in range(min(3, len(top_risky))):
                # Use direct column access via iloc for reliability
                rule_name_raw = top_risky.iloc[i][rule_name_col] if rule_name_col in top_risky.columns else None
                score_raw = top_risky.iloc[i]['Score_Total'] if 'Score_Total' in top_risky.columns else 0
                rule_name_str = str(rule_name_raw) if pd.notna(rule_name_raw) and rule_name_raw is not None else "N/A"
                sample_score = int(pd.to_numeric(score_raw, errors='coerce') or 0)
                logging.debug(f"Sample rule {i+1}: Raw value={rule_name_raw}, Name='{rule_name_str}', Score={sample_score}")
            
            # Iterate through the sorted rules using iloc for reliable access
            for idx in range(len(top_risky)):
                # Use iloc for direct row access - more reliable than iterrows
                row = top_risky.iloc[idx]
                score = int(pd.to_numeric(row['Score_Total'], errors='coerce') or 0)
                
                # Get rule name from Rule_Name column using direct column access
                if rule_name_col in top_risky.columns:
                    rule_name_raw = row[rule_name_col]
                else:
                    logging.error(f"Column '{rule_name_col}' not in row! Available: {row.index.tolist()}")
                    rule_name_raw = None
                
                # Convert to string, handling None and NaN
                if rule_name_raw is None or pd.isna(rule_name_raw) or str(rule_name_raw).strip() == '' or str(rule_name_raw).lower() == 'nan':
                    rule_name = f"Rule {idx + 1}"
                    logging.warning(f"Rule {idx + 1}: Rule name is empty/None, using fallback: '{rule_name}'")
                else:
                    rule_name = str(rule_name_raw).strip()
                    if idx < 3:  # Log first 3 for debugging
                        logging.debug(f"✅ Rule {idx + 1}: Successfully extracted Name='{rule_name}', Score={score}")
                
                # Determine risk level based on Score_Total ranges
                if score >= 175:
                    risk_level = "Critical"
                elif score >= 100:
                    risk_level = "High"
                elif score >= 50:
                    risk_level = "Medium"
                else:
                    risk_level = "Low"
                
                risky_rules.append({
                    "name": rule_name,
                    "score": score,
                    "riskLevel": risk_level
                })
            
            logging.debug(f"Returning {len(risky_rules)} risky rules")
        except Exception as e:
            logging.error(f"Error getting risky rules: {e}")
            logging.exception("Full traceback:")
        
        return risky_rules
    
    def _get_internet_rules(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """Get internet rules."""
        internet_rules = []
        if 'Risky_Inbound' in df.columns and 'Risky_Outbound' in df.columns:
            risky_mask = (df['Risky_Inbound'] == True) | (df['Risky_Outbound'] == True)
            internet_df = df[risky_mask]
            
            for _, row in internet_df.head(20).iterrows():
                rule_name = str(row.get('Name', 'Unknown'))
                internet_rules.append({
                    "name": rule_name,
                    "riskyInbound": bool(row.get('Risky_Inbound', False)),
                    "riskyOutbound": bool(row.get('Risky_Outbound', False)),
                    "score": int(pd.to_numeric(row.get('Score_Total', 0), errors='coerce') or 0)
                })
        
        return internet_rules
    
    def _count_insecure_ports(self, df: pd.DataFrame) -> int:
        """Count rules with insecure ports."""
        insecure_port_count = 0
        if 'Service_Insecure_Match' in df.columns:
            col = df['Service_Insecure_Match'].astype(str)
            logging.debug(f"Service_Insecure_Match column found. Sample values: {col.head(10).tolist()}")
            logging.debug(f"Unique values (first 20): {col.unique()[:20].tolist()}")
            # Check for "True" in the values
            insecure_port_count = col.str.contains('True', case=False, na=False).sum()
            logging.debug(f"Counted {insecure_port_count} rules with insecure ports")
        else:
            logging.warning("Service_Insecure_Match column not found in dataframe")
        
        return insecure_port_count
    
    def _count_source_user_not_used(self, df: pd.DataFrame) -> int:
        """Count rules where source user is not used (0 points)."""
        source_user_not_used = 0
        if 'Source_User_Scoring' in df.columns:
            col = df['Source_User_Scoring'].astype(str)
            # Count where value is "False,0" (0 points awarded)
            source_user_not_used = col.str.contains('False,0', case=False, na=False).sum()
        
        return source_user_not_used
    
    def get_analysis_data_from_excel(self, excel_file: pathlib.Path) -> Dict[str, Any]:
        """Get analysis data directly from Excel file without SQLite."""
        if not excel_file.exists():
            raise HTTPException(status_code=404, detail=f"Excel file not found: {excel_file}")
        
        try:
            # Read the Excel file directly - use data_only=True to get calculated values
            # Force reload by closing any existing workbook connections
            # Note: openpyxl doesn't cache, but we ensure fresh read each time
            wb = openpyxl.load_workbook(excel_file, data_only=True, read_only=False)
            
            # Try to find the correct sheet - prioritize 'Raw Data'
            sheet_name = None
            if 'Raw Data' in wb.sheetnames:
                sheet_name = 'Raw Data'
            elif 'Raw_Data' in wb.sheetnames:
                sheet_name = 'Raw_Data'
            elif len(wb.sheetnames) > 0:
                sheet_name = wb.sheetnames[0]
            else:
                raise HTTPException(status_code=400, detail="No sheets found in Excel file")
            
            ws = wb[sheet_name]
            
            # Read data into DataFrame
            data = []
            headers = [cell.value for cell in ws[1]]
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_data = {}
                for idx, cell_value in enumerate(row):
                    if idx < len(headers):
                        col_name = headers[idx]
                        row_data[col_name] = str(cell_value) if cell_value is not None else ""
                data.append(row_data)
            
            wb.close()
            
            # Convert to DataFrame
            df = pd.DataFrame(data)
            
            if len(df) == 0:
                return {
                    "totalRules": 0,
                    "averageScore": 0,
                    "highRisk": 0,
                    "mediumRisk": 0,
                    "lowRisk": 0,
                    "riskyRules": [],
                    "internetRules": [],
                    "insecurePortCount": 0,
                    "sourceUserNotUsed": 0
                }
            
            logging.debug(f"Read {len(df)} rows directly from Excel")
            
            # Calculate risk statistics
            total_rules = len(df)
            
            # Score column - convert from string to numeric for calculations
            if 'Score_Total' in df.columns:
                score_col = pd.to_numeric(df['Score_Total'], errors='coerce').fillna(0)
            else:
                score_col = pd.Series([0] * len(df))
            avg_score = score_col.mean()
            
            # Risk distribution
            high_risk = (score_col >= 76).sum()
            medium_risk = ((score_col >= 26) & (score_col < 76)).sum()
            low_risk = (score_col < 26).sum()
            
            # Top 10 risky rules
            risky_rules = self._get_risky_rules(df)
            
            # Internet rules
            internet_rules = self._get_internet_rules(df)
            
            # Count insecure ports from Service_Insecure_Match column
            insecure_port_count = self._count_insecure_ports(df)
            
            # Source User not Used - count rules with 0 points for Source_User_Scoring
            source_user_not_used = self._count_source_user_not_used(df)
            
            return {
                "totalRules": total_rules,
                "averageScore": round(avg_score, 2),
                "highRisk": int(high_risk),
                "mediumRisk": int(medium_risk),
                "lowRisk": int(low_risk),
                "riskyRules": risky_rules,
                "internetRules": internet_rules,
                "insecurePortCount": int(insecure_port_count),
                "sourceUserNotUsed": int(source_user_not_used)
            }
            
        except Exception as e:
            logging.error(f"Error reading Excel file: {e}")
            raise HTTPException(status_code=500, detail=f"Failed to read Excel file: {str(e)}")
    
    def get_excel_sheet_info(self, excel_file: pathlib.Path) -> Dict[str, Any]:
        """Get information about sheets in an Excel file."""
        if not excel_file.exists():
            raise HTTPException(status_code=404, detail=f"Excel file not found: {excel_file}")
        
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=False)
            sheet_info = {}
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Get basic sheet info
                max_row = ws.max_row
                max_col = ws.max_column
                
                # Get headers if available
                headers = []
                if max_row > 0:
                    headers = [cell.value for cell in ws[1] if cell.value is not None]
                
                sheet_info[sheet_name] = {
                    'max_row': max_row,
                    'max_column': max_col,
                    'header_count': len(headers),
                    'headers': headers[:10],  # First 10 headers
                    'has_data': max_row > 1
                }
            
            wb.close()
            
            return {
                'file_path': str(excel_file),
                'sheet_count': len(wb.sheetnames),
                'sheets': sheet_info,
                'recommended_sheet': self._get_recommended_sheet(wb.sheetnames)
            }
            
        except Exception as e:
            logging.error(f"Error reading Excel file info: {e}")
            raise HTTPException(status_code=500, detail=f"Failed to read Excel file info: {str(e)}")
    
    def _get_recommended_sheet(self, sheet_names: List[str]) -> str:
        """Get the recommended sheet name based on available sheets."""
        # Priority order for sheet names
        priority_names = ['Raw Data', 'Raw_Data', 'Data', 'Sheet1']
        
        for name in priority_names:
            if name in sheet_names:
                return name
        
        # Return first sheet if no priority match
        return sheet_names[0] if sheet_names else None
    
    def _count_true_values(self, df: pd.DataFrame, column_name: str) -> int:
        """Count True values in a column. Handles various formats: True/False, "True,20"/"False,0", etc."""
        if column_name not in df.columns:
            return 0
        
        col = df[column_name].astype(str)
        # Count if contains "True" (case-insensitive) - handles "True", "True,20", etc.
        # Convert to native Python int for JSON serialization
        return int(col.str.contains('True', case=False, na=False).sum())
    
    def get_firewall_risk_distribution(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """Get risk distribution grouped by firewall (Source_File column).
        
        Returns a list of dictionaries with risk counts per firewall, sorted by Critical count.
        """
        # Check if Source_File column exists
        source_file_col = None
        for col_name in ['Source_File', 'Source File', 'source_file', 'Firewall', 'Device', 'Source']:
            if col_name in df.columns:
                source_file_col = col_name
                break
        
        if not source_file_col or 'Score_Total' not in df.columns:
            # If no Source_File column or Score_Total, return empty list
            return []
        
        # Convert Score_Total to numeric
        df_temp = df.copy()
        df_temp['Score_Total_numeric'] = pd.to_numeric(df_temp['Score_Total'], errors='coerce').fillna(0)
        
        # Group by firewall and calculate risk counts
        firewall_risk_data = []
        
        for firewall_name, group_df in df_temp.groupby(source_file_col):
            # Skip NaN/empty firewall names
            if pd.isna(firewall_name) or str(firewall_name).strip() == '':
                continue
            
            score_col = group_df['Score_Total_numeric']
            
            # Calculate risk counts using same thresholds as Risk Dashboard
            critical = int((score_col >= 175).sum())
            high = int(((score_col >= 100) & (score_col < 175)).sum())
            medium = int(((score_col >= 50) & (score_col < 100)).sum())
            low = int(((score_col >= 1) & (score_col < 50)).sum())
            none = int((score_col == 0).sum())
            
            # Remove .csv extension from firewall name for display
            firewall_display = str(firewall_name).strip()
            if firewall_display.lower().endswith('.csv'):
                firewall_display = firewall_display[:-4]
            
            firewall_risk_data.append({
                "firewall": firewall_display,
                "critical": critical,
                "high": high,
                "medium": medium,
                "low": low,
                "none": none
            })
        
        # Sort by risk levels in priority order: Critical, High, Medium, Low, No Risk (all descending)
        firewall_risk_data.sort(
            key=lambda x: (
                x['critical'],    # Primary: Critical count
                x['high'],       # Secondary: High count
                x['medium'],     # Tertiary: Medium count
                x['low'],        # Quaternary: Low count
                x['none']        # Quinary: No Risk count
            ),
            reverse=True
        )
        
        return firewall_risk_data
    
    def get_excel_grid_data_from_excel(self, excel_file: pathlib.Path) -> Dict[str, Any]:
        """Get Excel grid data with computed columns for ExcelGrid page."""
        if not excel_file.exists():
            raise HTTPException(status_code=404, detail=f"Excel file not found: {excel_file}")
        
        try:
            # Read Excel file
            wb = openpyxl.load_workbook(excel_file, data_only=True, read_only=False)
            
            # Find the correct sheet
            sheet_name = None
            if 'Raw Data' in wb.sheetnames:
                sheet_name = 'Raw Data'
            elif 'Raw_Data' in wb.sheetnames:
                sheet_name = 'Raw_Data'
            elif len(wb.sheetnames) > 0:
                sheet_name = wb.sheetnames[0]
            else:
                raise HTTPException(status_code=400, detail="No sheets found in Excel file")
            
            ws = wb[sheet_name]
            
            # Read data into DataFrame
            data = []
            headers = [cell.value for cell in ws[1]]
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_data = {}
                for idx, cell_value in enumerate(row):
                    if idx < len(headers):
                        col_name = headers[idx]
                        row_data[col_name] = str(cell_value) if cell_value is not None else ""
                data.append(row_data)
            
            wb.close()
            
            # Convert to DataFrame
            df = pd.DataFrame(data)
            
            if len(df) == 0:
                return {
                    "success": True,
                    "data": [],
                    "totalRows": 0,
                    "columns": []
                }
            
            # Helper function to check if value contains "True" (case-insensitive)
            def is_true(value):
                if value is None or pd.isna(value):
                    return False
                return str(value).strip().lower() == 'true' or 'true' in str(value).lower()
            
            # Helper function to convert scoring value to True/False
            def scoring_to_bool(value):
                if value is None or pd.isna(value):
                    return False
                try:
                    num_value = pd.to_numeric(value, errors='coerce')
                    return not pd.isna(num_value) and num_value > 0
                except:
                    return False
            
            # Process each row and compute columns
            processed_data = []
            for idx, row in df.iterrows():
                row_dict = {}
                
                # Rule_Name
                rule_name_col = None
                for col in ['Rule_Name', 'Name', 'Rule Name', 'RuleName', 'name', 'rule_name', 'Rule', 'rule']:
                    if col in df.columns:
                        rule_name_col = col
                        break
                
                if rule_name_col and rule_name_col in row:
                    row_dict['Rule_Name'] = str(row[rule_name_col]) if pd.notna(row[rule_name_col]) else f"Rule {idx + 1}"
                else:
                    row_dict['Rule_Name'] = f"Rule {idx + 1}"
                
                # Source_File (Firewall)
                source_file_col = None
                for col in ['Source_File', 'Source File', 'source_file']:
                    if col in df.columns:
                        source_file_col = col
                        break
                
                if source_file_col and source_file_col in row:
                    firewall_name = str(row[source_file_col]) if pd.notna(row[source_file_col]) else "Unknown"
                    # Remove .csv extension if present
                    if firewall_name.lower().endswith('.csv'):
                        firewall_name = firewall_name[:-4]
                    row_dict['Source_File'] = firewall_name
                else:
                    row_dict['Source_File'] = "Unknown"
                
                # Overpermissive Source: OR of Src_IsAny, Src_CIDR_Le_16, Src_CIDR_17_22
                overpermissive_source = False
                for col in ['Src_IsAny', 'Src_CIDR_Le_16', 'Src_CIDR_17_22']:
                    if col in row and is_true(row[col]):
                        overpermissive_source = True
                        break
                row_dict['Overpermissive_Source'] = overpermissive_source
                
                # Overpermissive Destination: OR of Dst_IsAny, Dst_CIDR_Le_16, Dst_CIDR_17_22
                overpermissive_destination = False
                for col in ['Dst_IsAny', 'Dst_CIDR_Le_16', 'Dst_CIDR_17_22']:
                    if col in row and is_true(row[col]):
                        overpermissive_destination = True
                        break
                row_dict['Overpermissive_Destination'] = overpermissive_destination
                
                # Service & App: OR of Service_Any_OR_RangeGt1000, App_Any_OR_RangeGt1000
                service_app = False
                for col in ['Service_Any_OR_RangeGt1000', 'App_Any_OR_RangeGt1000']:
                    if col in row and is_true(row[col]):
                        service_app = True
                        break
                row_dict['Service_App'] = service_app
                
                # Shadow Rule: OR of Shadow_Rule or Partial_Shadow_Rule
                shadow_rule = False
                for col in ['Shadow_Rule', 'Partial_Shadow_Rule']:
                    if col in row and is_true(row[col]):
                        shadow_rule = True
                        break
                row_dict['Shadow_Rule'] = shadow_rule
                
                # Service_Insecure_Match
                if 'Service_Insecure_Match' in row:
                    row_dict['Insecure_Ports'] = is_true(row['Service_Insecure_Match'])
                else:
                    row_dict['Insecure_Ports'] = False
                
                # Risky_Inbound (True/False)
                if 'Risky_Inbound' in row:
                    row_dict['Risky_Inbound'] = is_true(row['Risky_Inbound'])
                else:
                    row_dict['Risky_Inbound'] = False
                
                # Risky_Outbound (True/False)
                if 'Risky_Outbound' in row:
                    row_dict['Risky_Outbound'] = is_true(row['Risky_Outbound'])
                else:
                    row_dict['Risky_Outbound'] = False
                
                # Rule Usage (True/False from Rule_Usage_Scoring)
                if 'Rule_Usage_Scoring' in row:
                    row_dict['Rule_Usage'] = scoring_to_bool(row['Rule_Usage_Scoring'])
                else:
                    row_dict['Rule_Usage'] = False
                
                # Last Used (Rule Usage Apps Seen)
                if 'Rule Usage Apps Seen' in row:
                    row_dict['Last_Used'] = str(row['Rule Usage Apps Seen']) if pd.notna(row['Rule Usage Apps Seen']) else ""
                else:
                    row_dict['Last_Used'] = ""
                
                # Rule Description (True/False from Rule_Usage_Description_Scoring)
                if 'Rule_Usage_Description_Scoring' in row:
                    row_dict['Rule_Description'] = scoring_to_bool(row['Rule_Usage_Description_Scoring'])
                else:
                    row_dict['Rule_Description'] = False
                
                # Source User (True/False from Source_User_Scoring)
                if 'Source_User_Scoring' in row:
                    row_dict['Source_User'] = scoring_to_bool(row['Source_User_Scoring'])
                else:
                    row_dict['Source_User'] = False
                
                # Security Profile (True/False from Profile_Scoring)
                if 'Profile_Scoring' in row:
                    row_dict['Security_Profile'] = scoring_to_bool(row['Profile_Scoring'])
                else:
                    row_dict['Security_Profile'] = False
                
                # Total Score (Score_Total - numeric value)
                if 'Score_Total' in row:
                    try:
                        score = pd.to_numeric(row['Score_Total'], errors='coerce')
                        row_dict['Total_Score'] = int(score) if not pd.isna(score) else 0
                    except:
                        row_dict['Total_Score'] = 0
                else:
                    row_dict['Total_Score'] = 0
                
                # Redundant Rule (True/False)
                if 'Redundant_Rule' in row:
                    row_dict['Redundant_Rule'] = is_true(row['Redundant_Rule'])
                else:
                    row_dict['Redundant_Rule'] = False
                
                # Consolidation (True/False from Consolidation_Candidate)
                if 'Consolidation_Candidate' in row:
                    row_dict['Consolidation'] = is_true(row['Consolidation_Candidate'])
                else:
                    row_dict['Consolidation'] = False
                
                # Add raw fields for LLM context
                for field in ['Source', 'Destination', 'Service', 'Application', 'Action', 
                              'Src_IsAny', 'Dst_IsAny', 'Service_Any_OR_RangeGt1000', 
                              'Service_Insecure_Match']:
                    if field in df.columns:
                         row_dict[field] = str(row[field]) if pd.notna(row[field]) else ""

                processed_data.append(row_dict)
            
            return {
                "success": True,
                "data": processed_data,
                "totalRows": len(processed_data),
                "columns": [
                    "Rule_Name", "Source_File", "Overpermissive_Source", "Overpermissive_Destination",
                    "Service_App", "Insecure_Ports", "Risky_Inbound", "Risky_Outbound",
                    "Rule_Usage", "Last_Used", "Rule_Description", "Source_User",
                    "Security_Profile", "Total_Score", "Shadow_Rule", "Redundant_Rule", "Consolidation"
                ]
            }
            
        except Exception as e:
            logging.error(f"Error getting excel grid data: {e}")
            logging.exception("Full traceback:")
            raise HTTPException(
                status_code=500,
                detail=f"Failed to get excel grid data: {str(e)}"
            )
    
    def get_dashboard_data_from_excel(self, excel_file: pathlib.Path) -> Dict[str, Any]:
        """Get dashboard data in the format expected by DashboardPage."""
        if not excel_file.exists():
            raise HTTPException(status_code=404, detail=f"Excel file not found: {excel_file}")
        
        try:
            # Read Excel file
            wb = openpyxl.load_workbook(excel_file, data_only=True, read_only=False)
            
            # Find the correct sheet
            sheet_name = None
            if 'Raw Data' in wb.sheetnames:
                sheet_name = 'Raw Data'
            elif 'Raw_Data' in wb.sheetnames:
                sheet_name = 'Raw_Data'
            elif len(wb.sheetnames) > 0:
                sheet_name = wb.sheetnames[0]
            else:
                raise HTTPException(status_code=400, detail="No sheets found in Excel file")
            
            ws = wb[sheet_name]
            
            # Read data into DataFrame
            data = []
            headers = [cell.value for cell in ws[1]]
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_data = {}
                for idx, cell_value in enumerate(row):
                    if idx < len(headers):
                        col_name = headers[idx]
                        row_data[col_name] = str(cell_value) if cell_value is not None else ""
                data.append(row_data)
            
            wb.close()
            
            # Convert to DataFrame
            df = pd.DataFrame(data)
            
            if len(df) == 0:
                return {
                    "success": True,
                    "totalRows": 0,
                    "overpermissive": {},
                    "riskDashboard": None,
                    "combinedRiskInbound": None,
                    "combinedRiskOutbound": None,
                    "firewallRiskDistribution": [],
                    "sheetUsed": sheet_name
                }
            
            total_rows = len(df)
            
            # Helper function to count True values in a column
            def count_column(col_name: str, display_name: str = None) -> Dict[str, Any]:
                count = self._count_true_values(df, col_name)
                return {
                    "count": int(count),  # Ensure native Python int for JSON serialization
                    "found": col_name in df.columns,
                    "columnName": str(col_name),  # Ensure string type
                    "displayName": str(display_name or col_name)  # Ensure string type
                }
            
            # Helper to try multiple column names
            def try_count_column(primary_col: str, fallback_col: str = None, display_name: str = None) -> Dict[str, Any]:
                if primary_col in df.columns:
                    return count_column(primary_col, display_name)
                elif fallback_col and fallback_col in df.columns:
                    return count_column(fallback_col, display_name)
                else:
                    return {
                        "count": 0,
                        "found": False,
                        "columnName": primary_col,
                        "displayName": display_name or primary_col
                    }
            
            # Build overpermissive object
            overpermissive = {
                "sourceAny": count_column("Src_IsAny", "Source Any"),
                "srcCidrLe16": try_count_column("Src_CIDR_Le_16", "Src_CIDR Less 16", "Src_CIDR_Le_16"),
                "srcCidr17_22": try_count_column("Src_CIDR_17_22", "Src_CIDR_le_22", "Src_CIDR_17_22"),
                "destinationAny": count_column("Dst_IsAny", "Destination Any"),
                "dstCidrLe16": try_count_column("Dst_CIDR_Le_16", "Dst_CIDR Less 16", "Dst_CIDR_Le_16"),
                "dstCidr17_22": try_count_column("Dst_CIDR_17_22", "Dst_CIDR_le_22", "Dst_CIDR_17_22"),
                "serviceBroad": count_column("Service_Any_OR_RangeGt1000", "Service is Broad"),
                "appBroad": count_column("App_Any_OR_RangeGt1000", "Application is Broad"),
                "insecurePorts": count_column("Service_Insecure_Match", "Insecure Ports"),
                "riskyInbound": count_column("Risky_Inbound", "Risky inbound internet"),
                "riskyOutbound": count_column("Risky_Outbound", "Risky outbound internet"),
                "migrateInsecure": try_count_column("Migrate_Insecure", "Migrate_Insecure_Score", "Migrate to App-ID (Insecure Ports)"),
                "migrateOtherPorts": count_column("Migrate_Other_ports_Score", "Migrate Other Ports"),
                "srcZoneIsAny": count_column("SrcZone_IsAny", "Source zone"),
                "dstZoneIsAny": count_column("DstZone_IsAny", "Destination zone"),
                "ruleUsageScoring": count_column("Rule_Usage_Scoring", "Unused Rules"),
                "profileScoring": count_column("Profile_Scoring", "Missing Security Profile"),
                "optionsScoring": count_column("Options_Scoring", "Missing Log Forwarding"),
                "ruleUsageDescriptionScoring": count_column("Rule_Usage_Description_Scoring", "Missing Description"),
                "sourceUserScoring": count_column("Source_User_Scoring", "Source User")
            }
            
            # Risk Dashboard - count by Score_Total ranges
            risk_dashboard = None
            if 'Score_Total' in df.columns:
                score_col = pd.to_numeric(df['Score_Total'], errors='coerce').fillna(0)
                # Convert all numpy/pandas int64 to native Python int for JSON serialization
                risk_dashboard = {
                    "found": True,
                    "columnName": "Score_Total",
                    "critical": int((score_col >= 175).sum()),
                    "high": int(((score_col >= 100) & (score_col < 175)).sum()),
                    "medium": int(((score_col >= 50) & (score_col < 100)).sum()),
                    "low": int(((score_col >= 1) & (score_col < 50)).sum()),
                    "none": int((score_col == 0).sum())
                }
            else:
                risk_dashboard = {
                    "found": False,
                    "columnName": "Score_Total",
                    "critical": 0,
                    "high": 0,
                    "medium": 0,
                    "low": 0,
                    "none": 0
                }
            
            # Combined Risk Inbound - count rules with risky inbound + overpermissive
            combined_risk_inbound = None
            if 'Risky_Inbound' in df.columns:
                risky_inbound_col = df['Risky_Inbound'].astype(str).str.contains('True', case=False, na=False)
                
                # Count overpermissive flags
                overpermissive_flags = (
                    (df.get('Src_IsAny', pd.Series([False] * len(df))).astype(str).str.contains('True', case=False, na=False)) |
                    (df.get('Dst_IsAny', pd.Series([False] * len(df))).astype(str).str.contains('True', case=False, na=False)) |
                    (df.get('Service_Any_OR_RangeGt1000', pd.Series([False] * len(df))).astype(str).str.contains('True', case=False, na=False))
                ).astype(int)
                
                # Count rules with risky inbound + number of overpermissive flags
                # Convert all numpy/pandas int64 to native Python int for JSON serialization
                combined_risk_inbound = {
                    "found": True,
                    "critical": int((risky_inbound_col & (overpermissive_flags >= 3)).sum()),
                    "high": int((risky_inbound_col & (overpermissive_flags == 2)).sum()),
                    "medium": int((risky_inbound_col & (overpermissive_flags == 1)).sum())
                }
            else:
                combined_risk_inbound = {
                    "found": False,
                    "critical": 0,
                    "high": 0,
                    "medium": 0
                }
            
            # Combined Risk Outbound - same logic but for Risky_Outbound
            combined_risk_outbound = None
            if 'Risky_Outbound' in df.columns:
                risky_outbound_col = df['Risky_Outbound'].astype(str).str.contains('True', case=False, na=False)
                
                # Count overpermissive flags
                overpermissive_flags = (
                    (df.get('Src_IsAny', pd.Series([False] * len(df))).astype(str).str.contains('True', case=False, na=False)) |
                    (df.get('Dst_IsAny', pd.Series([False] * len(df))).astype(str).str.contains('True', case=False, na=False)) |
                    (df.get('Service_Any_OR_RangeGt1000', pd.Series([False] * len(df))).astype(str).str.contains('True', case=False, na=False))
                ).astype(int)
                
                # Count rules with risky outbound + number of overpermissive flags
                # Convert all numpy/pandas int64 to native Python int for JSON serialization
                combined_risk_outbound = {
                    "found": True,
                    "critical": int((risky_outbound_col & (overpermissive_flags >= 3)).sum()),
                    "high": int((risky_outbound_col & (overpermissive_flags == 2)).sum()),
                    "medium": int((risky_outbound_col & (overpermissive_flags == 1)).sum())
                }
            else:
                combined_risk_outbound = {
                    "found": False,
                    "critical": 0,
                    "high": 0,
                    "medium": 0
                }
            
            # Get firewall risk distribution
            firewall_risk_distribution = self.get_firewall_risk_distribution(df)
            
            # Get top 10 risky rules
            risky_rules = self._get_risky_rules(df)
            logging.info(f"📊 Dashboard: Retrieved {len(risky_rules)} risky rules")
            if len(risky_rules) > 0:
                logging.info(f"📊 Dashboard: First risky rule: {risky_rules[0]}")
            
            # Ensure all numeric values are native Python types for JSON serialization
            result = {
                "success": True,
                "totalRows": int(total_rows),
                "overpermissive": overpermissive,
                "riskDashboard": risk_dashboard,
                "combinedRiskInbound": combined_risk_inbound,
                "combinedRiskOutbound": combined_risk_outbound,
                "firewallRiskDistribution": firewall_risk_distribution,
                "riskyRules": risky_rules,
                "sheetUsed": str(sheet_name) if sheet_name else "Unknown"
            }
            logging.info(f"📊 Dashboard: Returning response with {len(risky_rules)} risky rules")
            return result
            
        except Exception as e:
            logging.error(f"Error reading Excel file for dashboard: {e}")
            logging.exception("Full traceback:")
            raise HTTPException(status_code=500, detail=f"Failed to read Excel file: {str(e)}")
    
    


    def get_firewall_analysis_distribution(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """
        Get firewall-level distribution of advanced analysis metrics.
        Returns list of firewalls with counts for shadow, redundant, generalization, correlation, consolidation.
        """
        # Determine firewall column
        firewall_col = 'Source_File'
        if 'Source_File' not in df.columns:
            if 'Source File' in df.columns:
                firewall_col = 'Source File'
            elif 'Firewall' in df.columns:
                firewall_col = 'Firewall'
            elif 'Device' in df.columns:
                firewall_col = 'Device'
            elif 'Source' in df.columns:
                firewall_col = 'Source'
            else:
                return []
        
        firewall_data = {}
        
        # Check which columns exist
        has_shadow = 'Shadow_Rule' in df.columns
        has_partial_shadow = 'Partial_Shadow_Rule' in df.columns
        has_redundant = 'Redundant_Rule' in df.columns
        has_generalization = 'Generalization_Risk' in df.columns
        has_correlation = 'Correlation_Risk' in df.columns
        has_consolidation = 'Consolidation_Candidate' in df.columns
        
        # Check for risk columns
        has_score = 'Score' in df.columns
        has_risk_level = 'Risk Level' in df.columns
        
        # Group by firewall
        for _, row in df.iterrows():
            firewall = str(row.get(firewall_col, 'Unknown'))
            if firewall == 'Unknown' or pd.isna(row.get(firewall_col)):
                continue
            
            # Remove .csv or .xlsx extension for display
            firewall_clean = firewall.replace('.csv', '').replace('.xlsx', '')
            
            if firewall_clean not in firewall_data:
                firewall_data[firewall_clean] = {
                    'firewall': firewall_clean,
                    'shadow': 0,
                    'partialShadow': 0,
                    'totalShadow': 0,
                    'redundant': 0,
                    'generalization': 0,
                    'correlation': 0,
                    'consolidation': 0,
                    'total_rules': 0,
                    'high_risk_count': 0,
                    'medium_risk_count': 0,
                    'low_risk_count': 0,
                    'total_score': 0,
                    'average_score': 0.0,
                    'high_risk_percentage': 0.0
                }
            
            # Increment total rules
            firewall_data[firewall_clean]['total_rules'] += 1
            
            # Count risk levels
            if has_risk_level:
                risk_level = str(row.get('Risk Level', '')).lower()
                if 'high' in risk_level or 'critical' in risk_level:
                    firewall_data[firewall_clean]['high_risk_count'] += 1
                elif 'medium' in risk_level:
                    firewall_data[firewall_clean]['medium_risk_count'] += 1
                elif 'low' in risk_level:
                    firewall_data[firewall_clean]['low_risk_count'] += 1
            elif has_score:
                # Fallback to score if Risk Level not present
                try:
                    score = float(row.get('Score', 0))
                    if score >= 76:
                        firewall_data[firewall_clean]['high_risk_count'] += 1
                    elif score >= 26:
                        firewall_data[firewall_clean]['medium_risk_count'] += 1
                    else:
                        firewall_data[firewall_clean]['low_risk_count'] += 1
                except:
                    pass
            
            # Sum scores for average calculation
            if has_score:
                try:
                    firewall_data[firewall_clean]['total_score'] += float(row.get('Score', 0))
                except:
                    pass

            # Count policy issues
            if has_shadow and row.get('Shadow_Rule') == True:
                firewall_data[firewall_clean]['shadow'] += 1
                firewall_data[firewall_clean]['totalShadow'] += 1
            
            if has_partial_shadow and row.get('Partial_Shadow_Rule') == True:
                firewall_data[firewall_clean]['partialShadow'] += 1
                firewall_data[firewall_clean]['totalShadow'] += 1
            
            if has_redundant and row.get('Redundant_Rule') == True:
                firewall_data[firewall_clean]['redundant'] += 1
            
            if has_generalization and row.get('Generalization_Risk') == True:
                firewall_data[firewall_clean]['generalization'] += 1
            
            # Count correlation risks
            if has_correlation and row.get('Correlation_Risk') == True:
                firewall_data[firewall_clean]['correlation'] += 1
            
            # Count consolidation candidates
            if has_consolidation and row.get('Consolidation_Candidate') == True:
                firewall_data[firewall_clean]['consolidation'] += 1
        
        # Calculate totals and averages
        result = []
        for fw_data in firewall_data.values():
            fw_data['total'] = (
                fw_data['totalShadow'] +
                fw_data['redundant'] +
                fw_data['generalization'] +
                fw_data['correlation'] +
                fw_data['consolidation']
            )
            
            # Calculate average score
            if fw_data['total_rules'] > 0:
                fw_data['average_score'] = fw_data['total_score'] / fw_data['total_rules']
                fw_data['high_risk_percentage'] = (fw_data['high_risk_count'] / fw_data['total_rules']) * 100
            
            result.append(fw_data)
        
        # Sort by high risk count (descending)
        result.sort(key=lambda x: x['high_risk_count'], reverse=True)
        
        return result

# Factory function to create analysis engine
def create_analysis_engine(data_root: pathlib.Path) -> AnalysisEngine:
    """Create and return an AnalysisEngine instance."""
    return AnalysisEngine(data_root)
